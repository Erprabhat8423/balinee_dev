import sys
import os
import json
from django.core import serializers
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import getConfigurationResult,getModelColumnById
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings

# Create your views here.

# roleManagement View
@login_required
def index(request):
    context = {}

    permissions = SpPermissions.objects.filter(status=1) 
    modules = SpModules.objects.filter(status=1)
    for module in modules : 
        module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)

    page = request.GET.get('page')
    roles = SpRoles.objects.all().order_by('-id')
    paginator = Paginator(roles, getConfigurationResult('page_limit'))

    try:
        roles = paginator.page(page)
    except PageNotAnInteger:
        roles = paginator.page(1)
    except EmptyPage:
        roles = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = int(total_pages) % paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
        
        
    role_details = SpRoles.objects.order_by('-id').first()


    context['roles'] = roles
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    context['role_details'] = role_details
    context['permissions'] = permissions
    context['modules'] = modules
    context['page_title'] = "Manage Roles & Permissions"
    context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
    context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
    context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')
    

    template = 'role-permission/role-management/roles.html'

    return render(request, template, context)


@login_required
def roleDetails(request,role_id):

    context = {}

    permissions = SpPermissions.objects.filter(status=1) 
    modules = SpModules.objects.filter(status=1)
    for module in modules : 
        module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
    role_details = SpRoles.objects.get(id=role_id)

    context['role_details'] = role_details
    context['permissions'] = permissions
    context['modules'] = modules
    context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
    context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
    context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')

    template = 'role-permission/role-management/role-details.html'

    return render(request, template, context)


@login_required
def ajaxRoleList(request):
    page = request.GET.get('page')
    roles = SpRoles.objects.all().order_by('-id')
    paginator = Paginator(roles, getConfigurationResult('page_limit'))

    try:
        roles = paginator.page(page)
    except PageNotAnInteger:
        roles = paginator.page(1)
    except EmptyPage:
        roles = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpRoles.objects.order_by('-id').first()

    return render(request, 'role-permission/role-management/ajax-roles.html', {'roles': roles, 'total_pages':total_pages, 'organization_details': organization_details})


@login_required
def ajaxRoleLists(request):
    page = request.GET.get('page')

    roles = SpRoles.objects.all().order_by('-id')
    paginator = Paginator(roles, getConfigurationResult('page_limit'))

    try:
        roles = paginator.page(page)
    except PageNotAnInteger:
        roles = paginator.page(1)
    except EmptyPage:
        roles = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    return render(request, 'role-permission/role-management/ajax-organization-lists.html', {'roles': roles, 'total_pages':total_pages})







@login_required
def addRole(request):

    if request.method == "POST":
        response = {}
        if SpRoles.objects.filter(department_id=request.POST['department_id'], role_name=request.POST['role_name']).exists():
            response['message'] = "Role already exist"
            response['flag'] = False
        else:
            role = SpRoles()
            role.organization_id = request.POST['organization_id']
            role.organization_name = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
            role.department_id = request.POST['department_id']
            role.department_name = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
            role.role_name = request.POST['role_name']
            role.responsibilities = request.POST['responsibilities']

            if request.POST['reporting_role_id'] != "" :
                if int(request.POST['reporting_role_id']) > 0 :
                    reporting_department_id = getModelColumnById(SpRoles,request.POST['reporting_role_id'],'department_id')
                    role.reporting_department_id = reporting_department_id
                    role.reporting_department_name = getModelColumnById(SpDepartments,reporting_department_id,'department_name')
                else:
                    role.reporting_department_id = None
                    role.reporting_department_name = None

                role.reporting_role_id = request.POST['reporting_role_id']
                
                if int(request.POST['reporting_role_id']) > 0 :
                    role.reporting_role_name = getModelColumnById(SpRoles,request.POST['reporting_role_id'],'role_name')
                else :
                    role.reporting_role_name = "Super User"

            else :
                role.reporting_department_id = None
                role.reporting_department_name = None
                role.reporting_role_id = None
                role.reporting_role_name = None

            role.status = 1
            role.save()
            if role.id != "" :
                response['role_id'] = role.id
                response['message'] = "Record has been saved successfully"
                response['flag'] = True
            else:
                response['message'] = "Failed to saved"
                response['flag'] = False

        return JsonResponse(response)

    else:

        context = {}
        permissions = SpPermissions.objects.filter(status=1)
        organizations = SpOrganizations.objects.filter(status=1)
        departments = SpDepartments.objects.filter(status=1)

        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        
        for department in departments : 
            department.roles = SpRoles.objects.filter(status=1,department_id=department.id)

        context['permissions'] = permissions
        context['organizations'] = organizations
        context['departments'] = departments
        context['modules'] = modules
        template = 'role-permission/role-management/add-role.html'
        return render(request,template , context)

@login_required
def getAddRolePermission(request,role_id):
    try:
        context = {}
        role = SpRoles.objects.get(id=role_id)

        other_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for department in other_departments : 
            department.other_roles = SpRoles.objects.filter(status=1,department_id=department.id).exclude(id=role.id)
        
        permissions = SpPermissions.objects.filter(status=1)
        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        
        context['permissions'] = permissions
        context['modules'] = modules
        context['role'] = role
        context['other_departments'] = other_departments
        context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
        context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
        context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')
        template = 'role-permission/role-management/add-role-permission.html'
        return render(request,template ,context)
    except SpRoles.DoesNotExist:
        return HttpResponse('role not found')

@login_required
def addRolePermission(request):
    if request.method == "POST":
        role = SpRoles.objects.get(id=request.POST['role_id'])
        permissions = SpPermissions.objects.filter(status=1)
        sub_modules = SpSubModules.objects.filter(status=1)
        for sub_module in sub_modules :
            for permission in permissions :
                var_name = 'permission_'+ str(sub_module.id) + '_' + str(permission.id)
                if var_name in request.POST:
                    role_permission = SpRolePermissions()
                    role_permission.role_id = role.id
                    role_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                    role_permission.sub_module_id = sub_module.id
                    role_permission.permission_id = permission.id
                    role_permission.save()
                    total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                    if request.POST[total_work_flows_var] :

                        role_permission.workflow = request.POST[total_work_flows_var]
                        role_permission.save()

                        total_work_flows = json.loads(request.POST[total_work_flows_var])
                        
                        SpRoleWorkflowPermissions.objects.filter(role_id=role.id,sub_module_id=sub_module.id,permission_id=permission.id).delete()
                        
                        for total_work_flow in total_work_flows :
                            role_permission_level = SpRoleWorkflowPermissions()
                            role_permission_level.role_id = role.id
                            role_permission_level.sub_module_id = sub_module.id
                            role_permission_level.permission_id = permission.id
                            role_permission_level.level_id = total_work_flow['level_id']
                            role_permission_level.level = getModelColumnById(SpWorkflowLevels,total_work_flow['level_id'],'level')
                            role_permission_level.description = total_work_flow['description']
                            if int(total_work_flow['role_id']) > 0 :
                                role_permission_level.workflow_level_dept_id = getModelColumnById(SpRoles,total_work_flow['role_id'],'department_id')
                            else:
                                role_permission_level.workflow_level_dept_id = None

                            role_permission_level.workflow_level_role_id = total_work_flow['role_id']
                            role_permission_level.status = 1
                            role_permission_level.save()

        response = {}
        response['flag'] = True
        response['message'] = "Record updated successfully"
        return JsonResponse(response)
    else:
        return HttpResponse('Method not allowed')
    


def orgDepartmentOption(request,organization_id):
    response = {}
    options = '<option value="" selected>Select</option>'
    departments = SpDepartments.objects.filter(status=1,organization_id=organization_id)
    for department in departments : 
        options += "<option value="+str(department.id)+">"+department.department_name+"</option>"

    response['options'] = options
    return JsonResponse(response)

def orgRoleOption(request,organization_id):
    response = {}
    options = '<option value="">Select</option>'
    options += '<option value="0">Super Admin</option>'
    departments = SpDepartments.objects.filter(status=1,organization_id=organization_id)
    for department in departments : 
        roles = SpRoles.objects.filter(status=1,department_id=department.id)
        if roles:
            options += '<optgroup label="' + department.department_name + '">'
            for role in roles : 
                options += "<option value="+str(role.id)+">"+role.role_name+"</option>"
            options += '</optgroup>'
    

    response['options'] = options
    return JsonResponse(response)

def departmentRoleOption(request,department_id):
    response = {}
    options = '<option value="" selected>Select</option>'
    roles = SpRoles.objects.filter(status=1,department_id=department_id)
    for role in roles : 
        options += "<option value="+str(role.id)+">"+role.role_name+"</option>"

    response['options'] = options
    return JsonResponse(response)



@login_required
def editRole(request,role_id):
    if request.method == "POST":
        response = {}
        if SpRoles.objects.filter(role_name=request.POST['role_name'],department_id=request.POST['department_id']).exclude(id=role_id).exists() :
            response['flag'] = False
            response['message'] = "Role name already exist"
        else:
            role = SpRoles.objects.get(id=role_id)
            role.organization_id = request.POST['organization_id']
            role.organization_name = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
            role.department_id = request.POST['department_id']
            role.department_name = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
            role.role_name = request.POST['role_name']
            role.responsibilities = request.POST['responsibilities']

            if request.POST['reporting_role_id'] != "" :
                if int(request.POST['reporting_role_id']) > 0 :
                    reporting_department_id = getModelColumnById(SpRoles,request.POST['reporting_role_id'],'department_id')
                    role.reporting_department_id = reporting_department_id
                    role.reporting_department_name = getModelColumnById(SpDepartments,reporting_department_id,'department_name')
                else:
                    role.reporting_department_id = None
                    role.reporting_department_name = None

                role.reporting_role_id = request.POST['reporting_role_id']
                
                if int(request.POST['reporting_role_id']) > 0 :
                    role.reporting_role_name = getModelColumnById(SpRoles,request.POST['reporting_role_id'],'role_name')
                else :
                    role.reporting_role_name = "Super User"

            else :
                role.reporting_department_id = None
                role.reporting_department_name = None
                role.reporting_role_id = None
                role.reporting_role_name = None

            role.save()

            permissions = SpPermissions.objects.filter(status=1)
            sub_modules = SpSubModules.objects.filter(status=1)

            SpRolePermissions.objects.filter(role_id=role.id).delete()
            SpRoleWorkflowPermissions.objects.filter(role_id=role.id).delete()

            for sub_module in sub_modules :
                for permission in permissions :
                    var_name = 'permission_'+ str(sub_module.id) + '_' + str(permission.id)
                    if var_name in request.POST:
                        role_permission = SpRolePermissions()
                        role_permission.role_id = role.id
                        role_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                        role_permission.sub_module_id = sub_module.id
                        role_permission.permission_id = permission.id
                        role_permission.save()
                        total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                        if request.POST[total_work_flows_var] :

                            role_permission.workflow = request.POST[total_work_flows_var]
                            role_permission.save()

                            total_work_flows = json.loads(request.POST[total_work_flows_var])
                            
                            SpRoleWorkflowPermissions.objects.filter(role_id=role.id,sub_module_id=sub_module.id,permission_id=permission.id).delete()
                            for total_work_flow in total_work_flows :
                                role_permission_level = SpRoleWorkflowPermissions()
                                role_permission_level.role_id = role.id
                                role_permission_level.sub_module_id = sub_module.id
                                role_permission_level.permission_id = permission.id
                                role_permission_level.level_id = total_work_flow['level_id']
                                role_permission_level.level = getModelColumnById(SpWorkflowLevels,total_work_flow['level_id'],'level')
                                role_permission_level.description = total_work_flow['description']
                                if int(total_work_flow['role_id']) > 0 :
                                    role_permission_level.workflow_level_dept_id = getModelColumnById(SpRoles,total_work_flow['role_id'],'department_id')
                                else:
                                    role_permission_level.workflow_level_dept_id = None
                                role_permission_level.workflow_level_role_id = total_work_flow['role_id']
                                if 'status' in total_work_flow :
                                    role_permission_level.status = total_work_flow['status']
                                else:
                                    role_permission_level.status = 1
                                role_permission_level.save()
            response['flag']    = True
            response['message'] = "Record has been updated"
        return JsonResponse(response)
    else:
        context = {}
        role = SpRoles.objects.get(id=role_id)
        other_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for department in other_departments : 
            department.other_roles = SpRoles.objects.filter(status=1,department_id=department.id).exclude(id=role.id)

        permissions = SpPermissions.objects.filter(status=1)
        organizations = SpOrganizations.objects.filter(status=1)
        departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        
        reporting_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for reporting_department in reporting_departments:
            reporting_department.roles = SpRoles.objects.filter(status=1,department_id=reporting_department.id).exclude(id=role_id)
            
        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        

        context['role'] = role
        context['other_departments'] = other_departments
        context['permissions'] = permissions
        context['organizations'] = organizations
        context['departments'] = departments
        context['reporting_departments'] = reporting_departments
        context['modules'] = modules
        context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
        context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
        context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')
        template = 'role-permission/role-management/edit-role.html'
    return render(request,template , context)


@login_required
def updateOrganization(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            organization_name = request.POST.get('organization_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
            else:
                land_line_code = request.POST.get('land_line_code')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')
            address = request.POST.get('address')
            pincode = request.POST.get('pincode')

            error_count = 0
            if organization_name == '':
                error_count = 1
                error_response['organization_name_error'] = "Please enter organization name"
            if SpRoles.objects.filter(organization_name=organization_name).exclude(id=id).exists():
                error_count = 1
                error_response['organization_name_error'] = "Organization name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpRoles.objects.filter(email_id=email_id).exclude(id=id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"    
            if address == '':
                error_count = 1
                error_response['address_error'] = "Please enter address"
            if pincode == '':
                error_count = 1
                error_response['pincode_error'] = "Please enter pincode"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                data = SpRoles.objects.get(id=id)
                data.organization_name = organization_name
                data.land_line_code = land_line_code
                data.phone_code = phone_code
                data.landline_no = landline_no
                data.mobile_code = mobile_code
                data.mobile_no = mobile_no
                data.email_id = email_id
                data.address = address
                data.pincode = pincode
                data.save()
                response['error'] = False
                response['message'] = "Data updated successfully"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organization')


@login_required
def getOrganizationRecord(request):
    id = request.GET.get('id')
    organization = SpRoles.objects.get(id=id)
    departments = Departments.objects.all().filter(organization_id__icontains=id).order_by('-id')
    return render(request, 'role-permission/role-management/get-organization-record.html', {'organization_details': organization, 'departments': departments})


@login_required
def updateRoleStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpRoles.objects.get(id=id)
            data.status = is_active
            data.save()
            response['error'] = False
            response['message'] = "Data updated successfully"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/roles')



def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
def exportToPDF(request):
    context = {}
    roles = SpRoles.objects.all().values().order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('role-permission/role-management/organization_pdf_template.html', {'roles': roles, 'url': baseurl})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'organization.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
def exportToXlsx(request):
    roles = SpRoles.objects.all().order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=roles.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'roles'
    
    # Define the titles for columns
    columns = [ 'Organization Name', 'Landline No.', 'Mobile No.', 'Email Id', 'Address']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for organization in roles:
        row_num += 1
        # Define the data for each cell in the row 
        row = [
            organization.organization_name,
            organization.land_line_code + ' ' + organization.phone_code + ' ' + organization.landline_no,
            organization.mobile_code + ' ' + organization.mobile_no,
            organization.email_id,
            organization.address + ', ' + organization.pincode, 
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response