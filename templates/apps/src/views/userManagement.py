import sys
import os
import time
import xlrd
import MySQLdb
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.core import serializers

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.hashers import make_password

# Create your views here.

# User List View
@login_required
def index(request):
    page = request.GET.get('page')
    users = SpUsers.objects.all().filter(user_type=2).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 
    
    if(paginator.count == 0):
        paginator.count = 1

          
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    
    context = {}
    context['users'] = users
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')

    #non operational
    page = request.GET.get('non_page')
    users = SpUsers.objects.all().filter(user_type=3).order_by('-id')
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
        
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages

    #employee
    page = request.GET.get('employee_page')
    users = SpUsers.objects.all().filter(user_type=1).order_by('-id')
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
        
    context['employee_users'] = users
    context['employee_total_pages'] = total_pages

    first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
        ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id left join contract_type on contract_type.id=sp_basic_details.contract_type
        left join sp_addresses on sp_addresses.user_id = sp_users.id where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id=%s order by id desc LIMIT 1 ''',[1,'correspondence', first_employee_id.id])

    if first_employee :
        context['first_employee'] = first_employee[0]
        first_employee_permanent_address = SpAddresses.objects.get(user_id=first_employee[0].id,type='permanent')
    else : 
        context['first_employee'] = []
        first_employee_permanent_address = None

    context['first_employee_permanent_address'] = first_employee_permanent_address
    context['total_distributor'] = SpUsers.objects.filter(is_distributor=1).count()
    context['total_super_stockist'] = SpUsers.objects.filter(is_super_stockist=1).count()
    context['total_retailer'] = SpUsers.objects.filter(is_retailer=1).count()

    context['total_tagged_distributor'] = SpUsers.objects.filter(is_distributor=1,is_tagged=1).count()
    context['total_tagged_super_stockist'] = SpUsers.objects.filter(is_super_stockist=1,is_tagged=1).count()
    context['total_tagged_retailer'] = SpUsers.objects.filter(is_retailer=1,is_tagged=1).count()

    town_data = []
    towns = SpTowns.objects.all()
    for town in towns:

        distributors = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as distributor_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_distributor = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]
        town.distributor_count = distributors.distributor_count

        super_stockist = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as super_stockist_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_super_stockist = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]

        retailers = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as retailers_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_retailer = %s and sp_user_area_allocations.town_id = %s ''',[1,town.id])[0]

        town.retailer_count = retailers.retailers_count
    
        town_data.append(town)

    context['towns'] = town_data
    context['page_title'] = "Manage User"
    template = 'user-management/index.html'
    return render(request, template, context)

#ajax operational user list
@login_required
def ajaxOperationalUsersList(request):
    page = request.GET.get('page')

    users = SpUsers.objects.all().filter(user_type=2).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['users'] = users
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    template = 'user-management/ajax-operational-users-list.html'
    return render(request, template, context)

#ajax non operational user list
@login_required
def ajaxNonOperationalUsersList(request):
    page = request.GET.get('non_page')

    users = SpUsers.objects.all().filter(user_type=3).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages
    template = 'user-management/ajax-non-operational-users-list.html'
    return render(request, template, context)


#ajax employee user list
@login_required
def ajaxEmployeeUsersList(request):
    page = request.GET.get('employee_page')
    
    users = SpUsers.objects.all().filter(user_type=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['employee_users'] = users
    context['employee_total_pages'] = total_pages
    template = 'user-management/ajax-employee-users-list.html'
    
    return render(request, template, context)


# User basic details View
@login_required
def addUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types']    = contact_types
    context['countries']        = countries
    context['country_codes']    = country_codes

    template = 'user-management/add-user-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = SpUsers.objects.make_random_password()
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password    

            error_count = 0
            if request.POST['last_user_id'] != '':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            else:
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
  
            if user_exists:
                error_count = 1
                error_response['official_email_error'] = "Email already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if bool(request.FILES.get('store_image', False)) == True:
                    if request.POST['previous_store_image'] != '':
                        deleteMediaFile(request.POST['previous_store_image'])
                    uploaded_store_image = request.FILES['store_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    store_image_name = uploaded_store_image.name
                    temp = store_image_name.split('.')
                    store_image_name = 'store_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    store_image = storage.save(store_image_name, uploaded_store_image)
                    store_image = storage.url(store_image)

                else:
                    if request.POST['previous_store_image'] != '':
                        store_image = request.POST['previous_store_image'] 
                    else:
                        store_image = None    
                
                if bool(request.FILES.get('profile_image', False)) == True:
                    if request.POST['previous_profile_image'] != '':
                        deleteMediaFile(request.POST['previous_profile_image'])
                    uploaded_profile_image = request.FILES['profile_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    profile_image_name = uploaded_profile_image.name
                    temp = profile_image_name.split('.')
                    profile_image_name = 'profile_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    profile_image = storage.save(profile_image_name, uploaded_profile_image)
                    profile_image = storage.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None        
                
                if request.POST['last_user_id'] != '':
                    SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactPersons.objects.filter(user_id=request.POST['last_user_id']).delete()

                    user = SpUsers.objects.get(id=request.POST['last_user_id'])
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    if request.POST['store_name']:
                        user.user_type = 2
                        user.is_distributor = 1
                    user.save()
                    last_user_id = request.POST['last_user_id']
                else:
                    user = SpUsers()
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    user.password       = make_password(str(password))
                    user.plain_password = str(password)
                    if request.POST['store_name']:
                        user.user_type = 2
                        user.is_distributor = 1
                    user.save()
                    last_user_id = user.id
                    sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    user_contact_no         = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

            
                contact_person_names = request.POST.getlist('contact_person_name[]')
                designations = request.POST.getlist('designation[]')
                contact_numbers = request.POST.getlist('contact_number[]')

                for id, val in enumerate(contact_person_names):
                    user_contact_person         = SpContactPersons()
                    user_contact_person.user_id = last_user_id
                    if contact_person_names[id]!='':
                        user_contact_person.contact_person_name = contact_person_names[id]
                    if designations[id]!='':
                        user_contact_person.designation         = designations[id]
                    if contact_numbers[id]!='':
                        user_contact_person.contact_number      = contact_numbers[id]      
                    user_contact_person.save()

                if request.POST['last_user_id'] != '':
                    basic = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save()
                else:
                    basic = SpBasicDetails()
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save() 

                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id           = last_user_id
                permanent.type              = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations    = SpOrganizations.objects.filter(status=1)
                working_shifts  = SpWorkingShifts.objects.all()
                zones           = SpZones.objects.all()
                routes          = SpRoutes.objects.all()
                user_details    = SpUsers.objects.get(id=last_user_id)

                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=last_user_id)
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None
                
                try:
                    user_basic_details = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                
                try:
                    departments = SpDepartments.objects.filter(organization_id=user_details.organization_id)
                except SpDepartments.DoesNotExist:
                    departments = None

                try:
                    roles = SpRoles.objects.filter(department_id=user_details.department_id)
                except SpRoles.DoesNotExist:
                    roles = None
                
                if user_area_allocations is None:
                    towns = None
                else:
                    towns = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)

                context = {}
                context['oganizations']             = oganizations
                context['working_shifts']           = working_shifts
                context['zones']                    = zones
                context['routes']                   = routes
                context['user_details']             = user_details
                context['user_area_allocations']     = user_area_allocations
                context['user_basic_details']       = user_basic_details
                context['departments']              = departments
                context['roles']                    = roles
                context['towns']                    = towns
                context['last_user_id']             = last_user_id
                
                template = 'user-management/add-user-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template, context)

# Edit User basic details View
@login_required
def editUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    try:
        user_basic_details = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    except SpBasicDetails.DoesNotExist:
        user_basic_details = None

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_contact_details    = SpContactNumbers.objects.filter(user_id=request.GET['last_user_id'])
    user_basic_details      = user_basic_details
    user_store_address      = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='correspondence').first()
    user_permanent_address  = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='permanent').first()
    user_contact_person     = SpContactPersons.objects.filter(user_id=request.GET['last_user_id'])

    if user_store_address:
        store_states = SpStates.objects.filter(country_id=user_store_address.country_id)
        store_cities = SpCities.objects.filter(state_id=user_store_address.state_id)

        permanent_states = SpStates.objects.filter(country_id=user_permanent_address.country_id)
        permanent_cities = SpCities.objects.filter(state_id=user_permanent_address.state_id)
    else:
        store_states = None
        store_cities = None
        permanent_states = None
        permanent_cities = None

    
    context = {}
    context['contact_types']            = contact_types
    context['countries']                = countries
    context['country_codes']            = country_codes
    context['user_details']             = user_details
    context['user_contact_details']     = user_contact_details
    context['user_basic_details']       = user_basic_details
    context['user_store_address']       = user_store_address
    context['user_permanent_address']   = user_permanent_address
    context['user_contact_person']      = user_contact_person
    context['store_states']             = store_states
    context['store_cities']             = store_cities
    context['permanent_states']         = permanent_states
    context['permanent_cities']         = permanent_cities
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-basic-detail.html'
    
    return render(request, template, context)    

# User offical details View
@login_required
def addUserOfficalDetail(request):
    template = 'user-management/add-user-offical-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
            else:
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "SAP ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                user_data.save()

                

                try:
                    user_basic_detail = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                except SpBasicDetails.DoesNotExist:
                    user_basic_detail = None

                if user_basic_detail is None:        
                    user_basic_details                  = SpBasicDetails()
                else:
                    user_basic_details                  = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])

                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.cin                  = request.POST['cin']
                user_basic_details.gstin                = request.POST['gstin']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_basic_details.outstanding_amount   = request.POST['outstanding_amount']
                user_basic_details.opening_crates       = request.POST['opening_crates']
                user_basic_details.save()
                
                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None

                if user_area_allocations is None:        
                    area_allocation = SpUserAreaAllocations()
                else:
                    area_allocation = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])

                area_allocation.user_id     = request.POST['last_user_id']
                area_allocation.zone_id     = request.POST['zone_id']
                area_allocation.zone_name   = getModelColumnById(SpZones,request.POST['zone_id'],'zone')
                area_allocation.town_id     = request.POST['town_id']
                area_allocation.town_name   = getModelColumnById(SpTowns,request.POST['town_id'],'town')
                area_allocation.route_id    = request.POST['route_id']
                area_allocation.route_name  = getModelColumnById(SpRoutes,request.POST['route_id'],'route')
                area_allocation.save()
                
                try:
                    user_variants = SpUserProductVariants.objects.filter(user_id=request.POST['last_user_id'])
                except SpUserProductVariants.DoesNotExist:
                    user_variants = None

                user_details = SpUsers.objects.filter(id=request.POST['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

                context                  = {}
                context['last_user_id']  = request.POST['last_user_id']
                context['user_variants'] = user_variants
                context['user_details']  = user_details
                template                 = 'user-management/add-user-product-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)

# Edit User basic details View
@login_required
def editUserOfficalDetail(request):
    oganizations    = SpOrganizations.objects.filter(status=1)
    working_shifts  = SpWorkingShifts.objects.all()
    zones           = SpZones.objects.all()
    routes          = SpRoutes.objects.all()

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_area_allocations   = SpUserAreaAllocations.objects.get(user_id=request.GET['last_user_id'])
    user_basic_details      = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    
    departments     = SpDepartments.objects.filter(organization_id=user_details.organization_id)
    roles           = SpRoles.objects.filter(department_id=user_details.department_id)
    towns           = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)
    
    context = {}
    context['oganizations']             = oganizations
    context['working_shifts']           = working_shifts
    context['zones']                    = zones
    context['routes']                   = routes
    context['user_details']             = user_details
    context['user_area_allocations']     = user_area_allocations
    context['user_basic_details']       = user_basic_details
    context['departments']              = departments
    context['roles']                    = roles
    context['towns']                    = towns
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-offical-detail.html'
    
    return render(request, template, context)      

# User product details View
@login_required
def addUserProductDetail(request):
    context = {}
    template = 'user-management/add-user-product-detail.html'
    if request.method == "POST":
        try:
            user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
        except SpUserDocuments.DoesNotExist:
            user_documents = None
        context['user_documents']   = user_documents
        context['last_user_id']     = request.POST['last_user_id']
        template = 'user-management/add-user-document-detail.html'
    return render(request, template, context)

# User product details View
@login_required
def editUserProductDetail(request):
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=request.GET['last_user_id'])
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    user_details = SpUsers.objects.filter(id=request.GET['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

    context = {}
    context['last_user_id']  = request.GET['last_user_id']
    context['user_variants'] = user_variants
    context['user_details']  = user_details
    template = 'user-management/add-user-product-detail.html'
    return render(request, template, context) 

# User document details View
@login_required
def addUserDocumentDetail(request):
    template = 'user-management/add-user-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhaar_card'] != '':
                        deleteMediaFile(request.POST['previous_aadhaar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                if request.POST['previous_aadhaar_card'] != '':
                        aadhaar_card = request.POST['previous_aadhaar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card'] != '':
                        deleteMediaFile(request.POST['previous_pan_card'])        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                        pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None

            if bool(request.FILES.get('cin', False)) == True:
                if request.POST['previous_cin'] != '':
                        deleteMediaFile(request.POST['previous_cin'])
                uploaded_cin = request.FILES['cin']
                cins = FileSystemStorage()
                cin = cins.save(uploaded_cin.name, uploaded_cin)
                cin = cins.url(cin)
            else:
                if request.POST['previous_cin'] != '':
                        cin = request.POST['previous_cin'] 
                else:
                    cin = None

            if bool(request.FILES.get('gstin', False)) == True:
                if request.POST['previous_gstin'] != '':
                        deleteMediaFile(request.POST['previous_gstin'])
                uploaded_gstin = request.FILES['gstin']
                gst = FileSystemStorage()
                gstin = gst.save(uploaded_gstin.name, uploaded_gstin)
                gstin = gst.url(gstin)
            else:
                if request.POST['previous_gstin'] != '':
                        gstin = request.POST['previous_gstin'] 
                else:
                    gstin = None
            
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None

            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.save()

                response['error'] = False
                response['message'] = "Record has been successfully saved"
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.save()

                response['error'] = False
                response['message'] = "Record has been successfully updated"

            return JsonResponse(response)
        except Exception as e:
            response['error']            = True
            response['message']          = e
            return HttpResponse(e)
    return render(request, template)
    

# Employee basic details View
@login_required
def addEmployeeBasicDetail(request):
    contact_types = SpContactTypes.objects.filter(status=1)
    countries = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types'] = contact_types
    context['countries']     = countries
    context['country_codes'] = country_codes
    template = 'user-management/add-employee-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = SpUsers.objects.make_random_password()
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password
            
            error_count = 0 
            user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
                return JsonResponse(response)
            else:    
                if request.FILES['profile_image']:
                    uploaded_profile_image = request.FILES['profile_image']
                    pfs = FileSystemStorage()
                    profile_image = pfs.save(uploaded_profile_image.name, uploaded_profile_image)

                user = SpUsers()
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                if profile_image:
                    user.profile_image = pfs.url(profile_image)
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.user_type = 1
                user.password       = make_password(str(password))
                user.plain_password = str(password)
                user.save()
                last_user_id = user.id
                sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()

                basic = SpBasicDetails()
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations                = SpOrganizations.objects.filter(status=1)
                working_shifts              = SpWorkingShifts.objects.all()
                zones                       = SpZones.objects.all()
                routes                      = SpRoutes.objects.all()

                context                     = {}
                context['oganizations']     = oganizations
                context['working_shifts']   = working_shifts
                context['zones']            = zones
                context['routes']           = routes
                context['last_user_id']     = last_user_id

                template = 'user-management/add-employee-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    return render(request, template, context)

# Employee offical details View
@login_required
def addEmployeeOfficalDetail(request):
    template = 'user-management/add-employee-offical-detail.html'
    
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else: 
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name') 

                user_data.save()

                user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_basic_details.save()

                towns    = request.POST.getlist('town_id[]')
                for id, val in enumerate(towns):
                    area_allocation = SpUserAreaAllocations()
                    area_allocation.user_id = request.POST['last_user_id']
                    if towns[id] != '':
                        zone_id = getModelColumnById(SpTowns,towns[id],'zone_id')
                        area_allocation.zone_id                 =   zone_id
                        area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
                        area_allocation.town_id = towns[id]
                        area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
                    area_allocation.save()

                context = {}
                context['last_user_id'] = request.POST['last_user_id']
                
                distributors = SpUsers.objects.raw(''' select sp_users.id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                from sp_user_area_allocations 
                left join sp_users on sp_users.id = sp_user_area_allocations.user_id
                where sp_user_area_allocations.town_id in 
                (select town_id from sp_user_area_allocations as sura where sura.user_id = %s ) 
                and  (sp_users.is_distributor = %s or sp_users.is_super_stockist = %s)
                ''',[request.POST['last_user_id'], 1,1])

                # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
                # from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
                if distributors:
                    context['distributors'] = distributors
                else:
                    context['distributors'] = None
                template = 'user-management/add-employee-attendance.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template,response)

# Employee attendance details View
@login_required
def addEmployeeAttendanceDetail(request):
    if request.method == "POST":

        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = request.POST['last_user_id']
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()

        context = {}
        context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-document-detail.html'
        return render(request, template, context)
    else:
        context = {}
        distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
            from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        if distributors:
            context['distributors'] = distributors
        else:
            context['distributors'] = None
            context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-attendance.html'
        return render(request, template, context)

    

# Employee document details View
@login_required
def addEmployeeDocumentDetail(request):
    template = 'user-management/add-employee-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                pan_card = None

            documents               = SpUserDocuments()
            documents.user_id       = request.POST['last_user_id']
            documents.aadhaar_card  = aadhaar_card
            documents.pan_card      = pan_card
            documents.save()
            response['error'] = False
            response['message'] = "Record has been successfully updated"

            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)



# Employee basic details View
@login_required
def editEmployeeBasicDetail(request, employee_id):
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            else:
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
  
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
            
                if bool(request.FILES.get('profile_image', False)) == True:
                    uploaded_profile_image = request.FILES['profile_image']
                    pfs = FileSystemStorage()
                    profile_image = pfs.save(uploaded_profile_image.name, uploaded_profile_image)
                    profile_image = pfs.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None

                SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()

                user = SpUsers.objects.get(id=request.POST['last_user_id'])
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.profile_image = profile_image
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.save()
                last_user_id = request.POST['last_user_id']

            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
            
                basic                       = SpBasicDetails.objects.get(user_id=last_user_id)
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                response['error'] = False
                response['last_user_id'] = last_user_id
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))

    else : 
        contact_types = SpContactTypes.objects.filter(status=1)
        countries = SpCountries.objects.all()
        country_codes   = SpCountryCodes.objects.filter(status=1)

        context = {}
        context['contact_types'] = contact_types
        context['countries']     = countries
        context['country_codes'] = country_codes

        employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
        sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        where sp_users.id = %s''',[employee_id])
        if employee:
            context['employee'] = employee[0]
            context['employee_correspondence_address']  = employee_correspondence_address = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
            context['employee_permanent_address']       = employee_permanent_address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
            context['user_contacts']                    = SpContactNumbers.objects.filter(user_id=employee_id)
            context['user_areas']                       = SpUserAreaAllocations.objects.filter(user_id=employee_id)
            context['store_states']                     = SpStates.objects.filter(country_id=employee_correspondence_address.country_id)
            context['store_cities']                     = SpCities.objects.filter(state_id=employee_correspondence_address.state_id)
            context['permanent_states']                 = SpStates.objects.filter(country_id=employee_permanent_address.country_id)
            context['permanent_cities']                 = SpCities.objects.filter(state_id=employee_permanent_address.state_id)
            context['last_user_id']                     = employee_id
            try:
                user_documents                              = SpUserDocuments.objects.get(user_id=employee_id)
                context['user_documents'] = user_documents
            except SpUserDocuments.DoesNotExist:
                context['user_documents'] = None

            context['user_attendance_locations'] = SpUserAttendanceLocations.objects.filter(user_id=employee_id,status=1)
            template = 'user-management/edit-employee/employee-basic-detail.html'
            return render(request, template, context)
        else:
            return HttpResponse('Employee not found')

# Employee offical details View
@login_required
def editEmployeeOfficalDetail(request,employee_id):
    error_response = {}
    response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['SAPID_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:  
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)
                    
                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name')
                user_data.emp_sap_id            = request.POST['emp_sap_id']  
                

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name') 

                user_data.save()

                if request.POST['last_user_id'] != '':
                    user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    user_basic_details.pan_number           = request.POST['pan_number']
                    user_basic_details.working_shift_id     = request.POST['working_shift_id']
                    user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    user_basic_details.save()
                else:
                    user_basic_details                      = SpBasicDetails()
                    user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    user_basic_details.pan_number           = request.POST['pan_number']
                    user_basic_details.working_shift_id     = request.POST['working_shift_id']
                    user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    user_basic_details.save()
                
                SpUserAreaAllocations.objects.filter(user_id=request.POST['last_user_id']).delete()
                
                towns    = request.POST.getlist('town_id[]')
                for id, val in enumerate(towns):
                    
                    if towns[id] != '':
                        area_allocation                         = SpUserAreaAllocations()
                        area_allocation.user_id                 = request.POST['last_user_id']
                        zone_id                                 = getModelColumnById(SpTowns,towns[id],'zone_id')
                        area_allocation.zone_id                 =   zone_id
                        area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
                        area_allocation.town_id                 = towns[id]
                        area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
                        area_allocation.save()

                
                response = {}
                response['error'] = False
                response['last_user_id'] = request.POST['last_user_id']
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    else:
        
        template = 'user-management/edit-employee/employee-offical-detail.html'
        context = {}

        employee_details = SpUsers.objects.get(id=employee_id)
        employee_area_allocations = SpUserAreaAllocations.objects.filter(user_id=employee_id)
        emp_zones = []
        emp_towns = []
        for area_allocation in employee_area_allocations:
            emp_zones.append(area_allocation.zone_id)
            emp_towns.append(area_allocation.town_id)
        
        context['emp_zones'] = emp_zones = list(set(emp_zones))
        context['emp_towns'] = emp_towns

        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        working_shifts = SpWorkingShifts.objects.all()
        departments = SpDepartments.objects.filter(organization_id=employee_details.organization_id)
        roles = SpRoles.objects.filter(department_id=employee_details.department_id)
        reporting_users = SpUsers.objects.filter(role_id=employee_details.role_id)
        zones = SpZones.objects.filter()
        for zone in zones:
            zone.towns = SpTowns.objects.filter(zone_id=zone.id)

        oganizations = SpOrganizations.objects.filter(status=1)
        working_shifts = SpWorkingShifts.objects.all()

        

        context['oganizations']                 = oganizations
        context['working_shifts']               = working_shifts
        context['employee_details']             = employee_details
        context['employee_area_allocations']    = employee_area_allocations
        context['employee_basic_details']       = employee_basic_details
        context['departments']                  = departments
        context['roles']                        = roles
        context['reporting_users']              = reporting_users 
        context['zones']                        = zones
        context['last_user_id']                 = employee_id
        
        return render(request, template,context)

# Employee attendance details View
@login_required
def editEmployeeAttendanceDetail(request,employee_id):
    if request.method == "POST":
        SpUserAttendanceLocations.objects.filter(user_id=request.POST['last_user_id']).delete()
        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = employee_id
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()
                
        response = {}
        response['error'] = False
        response['last_user_id'] = request.POST['last_user_id']
        return JsonResponse(response)
    else:
        context = {}
        user_attendance_locations = SpUserAttendanceLocations.objects.filter(user_id=employee_id)
        if user_attendance_locations:
            user_attendance_locations = user_attendance_locations
        else:
            user_attendance_locations = None
        context['user_attendance_locations'] = user_attendance_locations
        distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
            from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        if distributors:
            context['distributors'] = distributors
        else:
            context['distributors'] = None

        context['last_user_id'] = employee_id
        context['user_attendance_locations'] = user_attendance_locations
        template = 'user-management/edit-employee/employee-attendance.html'
        return render(request, template, context)

# Employee document details View
@login_required
def editEmployeeDocumentDetail(request,employee_id):
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhar_card']:
                    deleteMediaFile(request.POST['previous_aadhar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                if request.POST['previous_aadhar_card'] != '':
                    aadhaar_card = request.POST['previous_aadhar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card']:
                    deleteMediaFile(request.POST['previous_pan_card'])        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                    pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None
                    
           
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None
            
            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been successfully saved"
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been successfully updated"
            
            return JsonResponse(response)

        except Exception as e:
            return HttpResponse(e)
    else:
        try:
            employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
        except SpUserDocuments.DoesNotExist:
            employee_documents = None

        template = 'user-management/edit-employee/employee-document-detail.html'
        response = {}
        response['last_user_id'] = employee_id
        response['employee_documents'] = employee_documents
        return render(request, template,response)




@login_required
def userShortDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender,
    sp_basic_details.outstanding_amount,sp_basic_details.opening_crates,
    sp_addresses.address_line_1,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,
    sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_addresses on sp_addresses.user_id = sp_users.id   
    where sp_addresses.type=%s  and sp_users.id = %s ''',['correspondence',user_id])
    if user :
        context['user'] = user[0]
        context['contact_persons'] = SpContactPersons.objects.filter(user_id=user_id)
    else : 
        context['user'] = []
    template = 'user-management/user-short-details.html'
    return render(request, template,context)

@login_required
def userDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.personal_email,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.date_of_joining,sp_basic_details.working_shift_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[user_id])[0]
    
    try:
        documents = SpUserDocuments.objects.get(user_id=user_id)
    except SpUserDocuments.DoesNotExist:
        documents = None

    try:
        allocation = SpUserAreaAllocations.objects.get(user_id=user_id)
    except SpUserAreaAllocations.DoesNotExist:
        allocation = None
            

    context['user'] = user
    context['user_correspondence_address'] = SpAddresses.objects.get(user_id=user_id,type='correspondence')
    context['user_permanent_address']      = SpAddresses.objects.get(user_id=user_id,type='permanent')
    context['contact_persons']             = SpContactPersons.objects.filter(user_id=user_id)
    context['user_contacts']               = SpContactNumbers.objects.filter(user_id=user_id)
    context['area_allocated']              = allocation 
    context['user_documents']              = documents
    context['user_attendance_locations']   = SpUserAttendanceLocations.objects.filter(user_id=user_id,status=1)
    template = 'user-management/user-details.html'

    return render(request, template,context)


@login_required
def employeeShortDetail(request,employee_id):
    context = {}
    employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type,sp_basic_details.gender,sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join contract_type on contract_type.id=sp_basic_details.contract_type
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id = %s ''',[1,'correspondence',employee_id])
    
    if employee :
        context['employee'] = employee[0]
    else : 
        context['employee'] = []

    context['employee_permanent_address'] = SpAddresses.objects.get(user_id=employee_id,type='permanent')    
    template = 'user-management/employee-short-details.html'
    return render(request, template,context)

@login_required
def employeeDetail(request,employee_id):
    context = {}
    context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[employee_id])[0]

    context['employee_correspondence_address'] = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
    context['employee_permanent_address'] = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    context['user_contacts'] = SpContactNumbers.objects.filter(user_id=employee_id)
    context['user_areas'] = SpUserAreaAllocations.objects.filter(user_id=employee_id)
    try:
        documents = SpUserDocuments.objects.get(user_id=employee_id)
    except SpUserDocuments.DoesNotExist:
        documents = None
    context['user_documents'] = documents
    context['user_attendance_locations'] = SpUserAttendanceLocations.objects.filter(user_id=employee_id,status=1)
    template = 'user-management/employee-details.html'

    return render(request, template,context)

def getGroupedTownOptions(request):
    options = ''
    zone_ids = request.POST['zone_ids'].split(',')
    zones = SpZones.objects.raw(''' select * from sp_zones where id in %s ''',[zone_ids])
    for zone in zones:
        towns = SpTowns.objects.filter(zone_id=zone.id)
        if towns:
            options += '<optgroup label="' + zone.zone + '">'
            for town in towns : 
                options += "<option value="+str(town.id)+">"+town.town+"</option>"
            options += '</optgroup>'
    
    return HttpResponse(options)

def getReportingUserOptions(request, role_id):
    options = '<option value="">Select role</option>'
    reporting_users = SpUsers.objects.raw(''' select id,first_name,last_name from sp_users where role_id = %s and user_type = %s ''',[role_id,1])
    for reporting_user in reporting_users:
        options += "<option value="+str(reporting_user.id)+">"+reporting_user.first_name+" "+reporting_user.last_name+"</option>"
    
    return HttpResponse(options)

#update user status
@login_required
def updateUserStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUsers.objects.get(id=id)
            data.status = is_active
            data.save()
            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#update user status
@login_required
def updateUserVariantPrice(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            price = request.POST.get('price')
            user_type = request.POST.get('user_type')
            is_distributor = request.POST.get('is_distributor')
            
            data = SpUserProductVariants.objects.get(id=id)
            if user_type == '2':
                if is_distributor == '1':
                    data.sp_distributor = price
                else:
                    data.sp_superstockist = price    
            else:
                data.sp_employee = price 
    
            data.save()

            data = SpUserProductVariants.objects.get(id=id)

            response['error'] = False
            response['message'] = "Record has been successfully updated"
            response['id'] = id
            response['price'] = price
            response['user_type'] = user_type
            response['is_distributor'] = is_distributor
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#export to excel operational user list
@login_required
def exportOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        try:
            user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)
        except SpBasicDetails.DoesNotExist:
            user.outstanding_amount = None
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Operational User'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]

    if 'outstanding_amount' in column_list:
        columns += [ 'Outstanding Amount' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        if user.outstanding_amount:
            outstanding_amount = user.outstanding_amount.outstanding_amount 
        else:
            outstanding_amount = ''   
        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]

        if 'outstanding_amount' in column_list:
            row += [ outstanding_amount ]           
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#export to pdf operational user list
@login_required
def exportOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#export to excel non-operational user list
@login_required
def exportNonOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=non-operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Non-Operational-Users'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]         
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

#export to pdf non operational user list
@login_required
def exportNonOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/non_operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Non-Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#export to excel employee list
@login_required
def exportEmployeeToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employees'
    
    # Define the titles for columns
    columns = []

    if 'employee_name' in column_list:
        columns += [ 'Employee Name' ]

    if 'employee_role' in column_list:
        columns += [ 'Role' ]
    
    if 'employee_dep_org' in column_list:
        columns += [ 'Dept./Org.' ] 

    if 'employee_platform' in column_list:
        columns += [ 'Platform(web/mobile)' ]    

    if 'employee_last_sign_in' in column_list:
        columns += [ 'Last Login' ]

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        if user.last_login is not None:
            if user.auth_token is None:
                employee_platform = 'Web'
            else:
                employee_platform = 'APP'
        else:
            employee_platform = ''               
         
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'employee_name' in column_list:
            row += [ user.first_name + user.middle_name + user.last_name ]

        if 'employee_role' in column_list:
            row += [ user.role_name ]
        
        if 'employee_dep_org' in column_list:
            row += [ user.department_name + '/' + user.organization_name ] 

        if 'employee_platform' in column_list:
            row += [ employee_platform ]

        if 'employee_last_sign_in' in column_list:
            row += [ user.last_login ]             
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response

    #export to pdf non operational user list
@login_required
def exportEmployeeToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/employee_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Employees.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

@login_required
def getUserMap(request):
    user_coordinates = SpUsers.objects.filter(id=request.GET.get('distributor_id')).values('latitude', 'longitude').first()
    
    context = {}
    context['user_coordinates'] = user_coordinates
    context['periphery']        = request.GET.get('periphery')
    template = 'user-management/user-map.html'
    return render(request, template, context) 

@login_required
def importProductVariant(request):
    # workbook object is created 
    # wb_obj = load_workbook('media/operatonal-users.xlsx')

    # sheet_obj = wb_obj.active 
    # m_row = sheet_obj.max_row 
    
    # # Loop will print all values 
    # # of first column  
    
    # for i in range(1, m_row + 1): 
    #     row = [cell.value for cell in sheet_obj[i]] 
    #     print(row)
    #     template = ProductVariantTemplate()
    #     template.store_name = row[0]
    #     template.role = row[1]
    #     template.save()
            
    return HttpResponse('row')     


def updateUserRole(user_id,params):
    role_permissions = SpRolePermissions.objects.filter(role_id=params.POST['role_id'])
    if len(role_permissions):
        SpUserRolePermissions.objects.filter(user_id=user_id).delete()
        for role_permission in role_permissions:
            user_role_permission = SpUserRolePermissions()
            user_role_permission.user_id = user_id
            user_role_permission.role_id = params.POST['role_id']
            user_role_permission.module_id = role_permission.module_id
            user_role_permission.sub_module_id = role_permission.sub_module_id
            user_role_permission.permission_id = role_permission.permission_id
            user_role_permission.workflow = role_permission.workflow
            user_role_permission.save()

        
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=params.POST['role_id'])
        if len(role_permission_workflows):
            SpUserRoleWorkflowPermissions.objects.filter(user_id=user_id).delete()
            for role_permission_workflow in role_permission_workflows : 
                user_role_permission_wf = SpUserRoleWorkflowPermissions()
                user_role_permission_wf.user_id = user_id
                user_role_permission_wf.role_id = role_permission_workflow.role_id
                user_role_permission_wf.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_wf.permission_id = role_permission_workflow.permission_id
                user_role_permission_wf.level_id = role_permission_workflow.level_id
                user_role_permission_wf.level = role_permission_workflow.level
                user_role_permission_wf.description = role_permission_workflow.description
                user_role_permission_wf.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                user_role_permission_wf.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_wf.status = role_permission_workflow.status
                user_role_permission_wf.save()


