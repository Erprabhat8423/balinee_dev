import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings

# Create your views here.

# Organizations View
@login_required
def index(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    if organization_details:
        departments = SpDepartments.objects.all().filter(organization_id=organization_details.id).order_by('-id')
    else:
        departments = {}
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['page_limit']             = getConfigurationResult('page_limit')
    context['organization_details']   = organization_details
    context['departments']            = departments
    context['page_title'] = "Manage Oragnizations"

    template = 'role-permission/organizations.html'
    return render(request, template, context)

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
def exportToPDF(request, columns):
    column_list = columns.split (",")
    context = {}
    organizations = SpOrganizations.objects.all().values().order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('role-permission/organization_pdf_template.html', {'organizations': organizations, 'url': baseurl, 'columns' : column_list})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'organizations.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
def exportToXlsx(request, columns):
    column_list = columns.split (",")
    organizations = SpOrganizations.objects.all().order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=organizations.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Organizations'
    
    # Define the titles for columns
    columns = []

    if 'org_name' in column_list:
        columns += [ 'Organization Name' ]

    if 'landline_no' in column_list:
        columns += [ 'Landline No.' ]
    
    if 'mobile_no' in column_list:
        columns += [ 'Mobile No.' ] 

    if 'email_id' in column_list:
        columns += [ 'Email Id' ]

        columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for organization in organizations:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'org_name' in column_list:
            row += [ organization.organization_name ]

        if 'landline_no' in column_list:
            row += [ organization.landline_country_code + ' ' + organization.landline_state_code + ' ' + organization.landline_number ]
        
        if 'mobile_no' in column_list:
            row += [ organization.mobile_country_code + ' ' + organization.mobile_number ] 

        if 'email_id' in column_list:
            row += [ organization.email ]       
       
        row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response


@login_required
def ajaxOrganizationList(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['organization_details']   = organization_details

    template = 'role-permission/ajax-organizations.html'
    return render(request, template, context)


@login_required
def ajaxOrganizationLists(request):
    page = request.GET.get('page')

    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    
    context = {}
    context['organizations']     = organizations
    context['total_pages']       = total_pages

    template = 'role-permission/ajax-organization-lists.html'
    return render(request, template, context)


@login_required
def addOrganization(request):
    template = 'role-permission/add-organization.html'
    return render(request, template)


@login_required
def saveOrganization(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            organization_name = request.POST.get('organization_name')
            
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = ''
            else:
                land_line_code = request.POST.get('land_line_code')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')
            address = request.POST.get('address')
            pincode = request.POST.get('pincode')

            error_count = 0
            if organization_name == '':
                error_count = 1
                error_response['organization_name_error'] = "Please enter organization name"
            if SpOrganizations.objects.filter(organization_name=organization_name).exists():
                error_count = 1
                error_response['organization_name_error'] = "Organization name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpOrganizations.objects.filter(email=email_id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"    
            if address == '':
                error_count = 1
                error_response['address_error'] = "Please enter address"
            if pincode == '':
                error_count = 1
                error_response['pincode_error'] = "Please enter pincode"    
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                SpOrganizations.objects.create(
                    organization_name=organization_name,
                    landline_country_code=land_line_code,
                    landline_state_code=phone_code,
                    landline_number=landline_no,
                    mobile_country_code=mobile_code,
                    mobile_number=mobile_no,
                    email=email_id,
                    address=address,
                    pincode=pincode,
                    status=1
                )
                response['error'] = False
                response['message'] = "Record has been successfully saved"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
def editOrganization(request):
    id = request.GET.get('id')

    context = {}
    context['organization']     = SpOrganizations.objects.get(id=id)

    template = 'role-permission/edit-organization.html'
    return render(request, template, context)


@login_required
def updateOrganization(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            organization_name = request.POST.get('organization_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
            else:
                land_line_code = request.POST.get('land_line_code')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')
            address = request.POST.get('address')
            pincode = request.POST.get('pincode')

            error_count = 0
            if organization_name == '':
                error_count = 1
                error_response['organization_name_error'] = "Please enter organization name"
            if SpOrganizations.objects.filter(organization_name=organization_name).exclude(id=id).exists():
                error_count = 1
                error_response['organization_name_error'] = "Organization name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpOrganizations.objects.filter(email=email_id).exclude(id=id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"    
            if address == '':
                error_count = 1
                error_response['address_error'] = "Please enter address"
            if pincode == '':
                error_count = 1
                error_response['pincode_error'] = "Please enter pincode"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                data = SpOrganizations.objects.get(id=id)
                data.organization_name = organization_name
                data.landline_country_code = land_line_code
                data.landline_state_code = phone_code
                data.landline_number = landline_no
                data.mobile_country_code = mobile_code
                data.mobile_number = mobile_no
                data.email = email_id
                data.address = address
                data.pincode = pincode
                data.save()
                response['error'] = False
                response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
def getOrganizationRecord(request):
    id = request.GET.get('id')
    
    context = {}
    context['organization_details']     = SpOrganizations.objects.get(id=id)
    context['departments']              = SpDepartments.objects.all().filter(organization_id=id).order_by('-id')

    template = 'role-permission/get-organization-record.html'
    return render(request, template, context)


@login_required
def updateOrganizationStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpOrganizations.objects.get(id=id)
            data.status = is_active
            data.save()
            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
def addDepartment(request):
    context = {}
    context['organization_id']          = request.GET.get('organization_id')
    context['organization_name']        = request.GET.get('organization_name')
    context['landline']                 = request.GET.get('landline')
    context['country_code']             = request.GET.get('country_code')
    context['code']                     = request.GET.get('code')
    context['country_codes']            = SpCountryCodes.objects.filter(status=1)

    template = 'role-permission/add-department.html'
    return render(request, template, context)


@login_required
def saveDepartment(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            organization_id = request.POST.get('organization_id')
            department_name = request.POST.get('department_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
                dept_ext = '';
            else:
                land_line_code = request.POST.get('land_line_code')
                dept_ext = request.POST.get('dept_ext')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')

            error_count = 0
            if department_name == '':
                error_count = 1
                error_response['department_name_error'] = "Please enter department name"
            if SpDepartments.objects.filter(department_name=department_name).filter(Q(organization_id = organization_id)).exists():
                error_count = 1
                error_response['department_name_error'] = "Department name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpDepartments.objects.filter(email=email_id).filter(Q(organization_id = organization_id)).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists" 
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                SpDepartments.objects.create(
                    organization_id=organization_id,
                    organization_name=getModelColumnById(SpOrganizations,organization_id,'organization_name'),
                    department_name=department_name,
                    landline_country_code=land_line_code,
                    landline_state_code=phone_code,
                    landline_number=landline_no,
                    extension_number=dept_ext,
                    mobile_country_code=mobile_code,
                    mobile_number=mobile_no,
                    email=email_id,
                    status=1
                )
                response['error'] = False
                response['message'] = "Record has been successfully saved"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
def editDepartment(request):
    organization_id = request.GET.get('organization_id')
    organization_name = request.GET.get('organization_name')
    department_name = request.GET.get('department_name')
    id = request.GET.get('id')
    department = SpDepartments.objects.get(id=id)
    return render(request, 'role-permission/edit-department.html', {'organization_id': organization_id, 'department_name': department_name, 'department': department, 'organization_name':organization_name})


@login_required
def updateDepartment(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            organization_id = request.POST.get('organization_id')
            department_name = request.POST.get('department_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
                dept_ext = '';
            else:
                land_line_code = request.POST.get('land_line_code')
                dept_ext = request.POST.get('dept_ext')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')

            error_count = 0
            if department_name == '':
                error_count = 1
                error_response['department_name_error'] = "Please enter department name"
            if SpDepartments.objects.filter(department_name=department_name).filter(Q(organization_id = organization_id)).exclude(id=id).exists():
                error_count = 1
                error_response['department_name_error'] = "Department name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpDepartments.objects.filter(email=email_id).filter(Q(organization_id = organization_id)).exclude(id=id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                data = SpDepartments.objects.get(id=id)
                data.organization_id = organization_id
                data.organization_name = getModelColumnById(SpOrganizations,organization_id,'organization_name')
                data.department_name = department_name
                data.landline_country_code = land_line_code
                data.extension_number = dept_ext
                data.landline_state_code = phone_code
                data.landline_number = landline_no
                data.mobile_country_code = mobile_code
                data.mobile_number = mobile_no
                data.email = email_id
                data.save()
                response['error'] = False
                response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
def updateDepartmentStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpDepartments.objects.get(id=id)
            data.status = is_active
            data.save()
            response['error'] = False
            response['message'] = "Record has been successfully updated"
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')
