import sys
import shutil, os
import time,json
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
import openpyxl
from decimal import Decimal
from datetime import timedelta,date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.core import serializers
from math import sin, cos, sqrt, atan2, radians
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.hashers import make_password


from PIL import Image

# Create your views here.


def mapUserLeavess(request,role_id):
    user_id = SpUsers.objects.filter(role_id  = role_id,status = 1)
    for i in user_id:
        mapUserLeaves(role_id,i.id)
    
    
# User List View
@login_required
def index(request):
    #employee
    page = request.GET.get('employee_page')
    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    all_users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    context = {}   
    context['all_users'] = all_users
    context['employee_users'] = users
    context['employee_total_pages'] = total_pages

    first_employee_id = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id').first()

    if first_employee_id is not None:
        # first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
        # ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
        # FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
        # left join sp_addresses on sp_addresses.user_id = sp_users.id  
        # where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id=%s order by id desc LIMIT 1 ''',[1,'correspondence', first_employee_id.id])
        first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type ,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
        ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id left join contract_type on contract_type.id=sp_basic_details.contract_type
        left join sp_addresses on sp_addresses.user_id = sp_users.id where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id=%s order by id desc LIMIT 1 ''',[1,'correspondence', first_employee_id.id])
        
        if first_employee :
            context['first_employee'] = first_employee[0]
            try:
                first_employee_permanent_address = SpAddresses.objects.get(user_id=first_employee[0].id,type='permanent')
            except SpAddresses.DoesNotExist:
                first_employee_permanent_address = None
            
        else : 
            context['first_employee'] = []
            first_employee_permanent_address = None

        context['first_employee_permanent_address'] = first_employee_permanent_address
  
    if request.user.role_id == 0:
        organizations = SpOrganizations.objects.all()
        departments = SpDepartments.objects.all()
        roles = SpRoles.objects.all()
        
    town_data = []
    towns = SpTowns.objects.all()
    context['jobs']  = ContractType.objects.all()
    context['organizations']  = organizations
    context['departments']  = SpDepartments.objects.all()
    context['roles']  = roles

    context['towns'] = town_data
    context['page_title'] = "Manage Employees"
    template = 'user-management/index.html'
    return render(request, template, context)

#ajax operational user list
@login_required
def ajaxOperationalUsersList(request):
    page = request.GET.get('page')

    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['users'] = users
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    template = 'user-management/ajax-operational-users-list.html'
    return render(request, template, context)

#ajax non operational user list
@login_required
def ajaxNonOperationalUsersList(request):
    page = request.GET.get('non_page')

    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages
    template = 'user-management/ajax-non-operational-users-list.html'
    return render(request, template, context)


# #ajax employee user list
# @login_required
# def ajaxEmployeeUsersList(request):
#     page = request.GET.get('employee_page')
    
#     users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
#     paginator = Paginator(users, getConfigurationResult('page_limit'))

#     try:
#         users = paginator.page(page)
#     except PageNotAnInteger:
#         users = paginator.page(1)
#     except EmptyPage:
#         users = paginator.page(paginator.num_pages)  
#     if page is not None:
#           page = page
#     else:
#           page = 1

#     total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
#     if(paginator.count == 0):
#         paginator.count = 1
        
#     temp = total_pages%paginator.count
#     if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
#         total_pages = total_pages+1
#     else:
#         total_pages = total_pages

#     context = {}
#     context['employee_users'] = users
#     context['employee_total_pages'] = total_pages
#     template = 'user-management/ajax-employee-users-list.html'
    
#     return render(request, template, context)


#ajax employee user list
@login_required
def ajaxEmployeeUsersList(request):
    page = request.GET.get('employee_page')
    search = request.GET.get('search')
    jobs = request.GET.getlist('jobs[]')
    departments = request.GET.getlist('depts[]')
    roles = request.GET.getlist('roles[]')


    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    if search:
        users = users.filter(id = search)
    if len(departments) > 0:
        users = users.filter(department_id__in = departments)
    if len(roles) > 0:
        users = users.filter(role_id__in = roles)
    
    if len(jobs) > 0:
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type__in = jobs):
                users = users.exclude(id = user.id)

    users_count = users.count()
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['employee_users'] = users
    context['employee_count'] = users_count
    context['employee_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'user-management/ajax-employee-users-list.html'
    
    return render(request, template, context)



@login_required
def userGeoTagged(request):
    context = {}
  
    try:
        user_coordinates = SpUsers.objects.get(id=request.GET['id'])
    except SpUserAttendanceLocations.DoesNotExist:
        user_coordinates = None
           
    context['user_coordinates'] = user_coordinates
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-geo-tagged.html'

    return render(request, template,context)

# User basic details View
@login_required
def addUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types']    = contact_types
    context['countries']        = countries
    context['country_codes']    = country_codes

    template = 'user-management/add-user-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password    

            error_count = 0
            if request.POST['last_user_id'] != '' and request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            elif request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            else:
                user_exists = 0

            if user_exists:
                error_count = 1
                error_response['official_email_error'] = "Email already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if bool(request.FILES.get('store_image', False)) == True:
                    if request.POST['previous_store_image'] != '':
                        deleteMediaFile(request.POST['previous_store_image'])
                    uploaded_store_image = request.FILES['store_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    store_image_name = uploaded_store_image.name
                    temp = store_image_name.split('.')
                    store_image_name = 'store_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    store_image = storage.save(store_image_name, uploaded_store_image)
                    store_image = storage.url(store_image)

                else:
                    if request.POST['previous_store_image'] != '':
                        store_image = request.POST['previous_store_image'] 
                    else:
                        store_image = None    
                
                if bool(request.FILES.get('profile_image', False)) == True:
                    if request.POST['previous_profile_image'] != '':
                        deleteMediaFile(request.POST['previous_profile_image'])
                    uploaded_profile_image = request.FILES['profile_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    profile_image_name = uploaded_profile_image.name
                    temp = profile_image_name.split('.')
                    profile_image_name = 'profile_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    profile_image = storage.save(profile_image_name, uploaded_profile_image)
                    profile_image = storage.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None        
                
                if request.POST['last_user_id'] != '':
                    SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactPersons.objects.filter(user_id=request.POST['last_user_id']).delete()

                    user = SpUsers.objects.get(id=request.POST['last_user_id'])
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = request.POST['last_user_id']
                    user_inserted = 1
                else:
                    user = SpUsers()
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    user.password       = make_password(str(password))
                    user.plain_password = str(password)
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = user.id
                    user_inserted = 0
                    #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    user_contact_no         = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

            
                contact_person_names = request.POST.getlist('contact_person_name[]')
                designations = request.POST.getlist('designation[]')
                contact_numbers = request.POST.getlist('contact_number[]')

                for id, val in enumerate(contact_person_names):
                    user_contact_person         = SpContactPersons()
                    user_contact_person.user_id = last_user_id
                    if contact_person_names[id]!='':
                        user_contact_person.contact_person_name = contact_person_names[id]
                    if designations[id]!='':
                        user_contact_person.designation         = designations[id]
                    if contact_numbers[id]!='':
                        user_contact_person.contact_number      = contact_numbers[id]      
                    user_contact_person.save()

                if request.POST['last_user_id'] != '':
                    basic = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save()
                else:
                    basic = SpBasicDetails()
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save() 

                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id           = last_user_id
                permanent.type              = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations    = SpOrganizations.objects.filter(status=1)
                working_shifts  = TblClWorkingShifts.objects.all()
                zones           = SpZones.objects.all()
                routes          = SpRoutes.objects.all()
                user_details    = SpUsers.objects.get(id=last_user_id)

                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=last_user_id)
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None
                
                try:
                    user_basic_details = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                
                try:
                    departments = SpDepartments.objects.filter(organization_id=user_details.organization_id)
                except SpDepartments.DoesNotExist:
                    departments = None

                try:
                    roles = SpRoles.objects.filter(department_id=user_details.department_id)#.filter(id__in=[8,9])
                except SpRoles.DoesNotExist:
                    roles = None
                
                if user_area_allocations is None:
                    towns = None
                else:
                    towns = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)

                if user_inserted == 0:
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
                else:
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
                
                context = {}
                context['oganizations']             = oganizations
                context['working_shifts']           = working_shifts
                context['zones']                    = zones
                context['routes']                   = routes
                context['user_details']             = user_details
                context['user_area_allocations']     = user_area_allocations
                context['user_basic_details']       = user_basic_details
                context['departments']              = departments
                context['roles']                    = roles
                context['towns']                    = towns
                context['last_user_id']             = last_user_id
                
                template = 'user-management/add-user-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template, context)

# Edit User basic details View
@login_required
def editUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.all()
    country_codes   = SpCountryCodes.objects.filter(status=1)

    try:
        user_basic_details = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    except SpBasicDetails.DoesNotExist:
        user_basic_details = None

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_contact_details    = SpContactNumbers.objects.filter(user_id=request.GET['last_user_id'])
    user_basic_details      = user_basic_details
    user_store_address      = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='correspondence').first()
    user_permanent_address  = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='permanent').first()
    user_contact_person     = SpContactPersons.objects.filter(user_id=request.GET['last_user_id'])

    if user_store_address:
        store_states = SpStates.objects.filter(country_id=user_store_address.country_id)
        store_cities = SpCities.objects.filter(state_id=user_store_address.state_id)
        if user_permanent_address:
            permanent_states = SpStates.objects.filter(country_id=user_permanent_address.country_id)
            permanent_cities = SpCities.objects.filter(state_id=user_permanent_address.state_id)
        else:
            permanent_states = None
            permanent_cities = None    
    else:
        store_states = None
        store_cities = None
        permanent_states = None
        permanent_cities = None

    context = {}
    context['contact_types']            = contact_types
    context['countries']                = countries
    context['country_codes']            = country_codes
    context['user_details']             = user_details
    context['user_contact_details']     = user_contact_details
    context['user_basic_details']       = user_basic_details
    context['user_store_address']       = user_store_address
    context['user_permanent_address']   = user_permanent_address
    context['user_contact_person']      = user_contact_person
    context['store_states']             = store_states
    context['store_cities']             = store_cities
    context['permanent_states']         = permanent_states
    context['permanent_cities']         = permanent_cities
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-basic-detail.html'
    
    return render(request, template, context)    

# User offical details View
@login_required
def addUserOfficalDetail(request):
    template = 'user-management/add-user-offical-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
            else:
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "SAP ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role & map product
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                    mapProductToUser(user_data.id)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                if request.POST['role_id'] == '8':
                    user_data.is_distributor = 1
                    user_data.is_super_stockist = 0
                else:
                    user_data.is_distributor = 0
                    user_data.is_super_stockist = 1    
                user_data.save()

                

                try:
                    user_basic_detail = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                except SpBasicDetails.DoesNotExist:
                    user_basic_detail = None

                if user_basic_detail is None:        
                    user_basic_details                  = SpBasicDetails()
                else:
                    user_basic_details                  = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])

                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.cin                  = request.POST['cin']
                user_basic_details.gstin                = request.POST['gstin']
                user_basic_details.fssai                = request.POST['fssai']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.order_timing         = request.POST['order_timing']
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['date_of_relieving']:
                    user_basic_details.date_of_relieving      = datetime.strptime(request.POST['date_of_relieving'], '%d/%m/%Y').strftime('%Y-%m-%d') 
                user_basic_details.outstanding_amount   = request.POST['outstanding_amount']
                user_basic_details.security_amount      = request.POST['security_amount']
                user_basic_details.opening_crates       = request.POST['opening_crates']
                user_basic_details.save()
                
                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None

                if user_area_allocations is None:        
                    area_allocation = SpUserAreaAllocations()
                else:
                    area_allocation = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])

                area_allocation.user_id     = request.POST['last_user_id']
                area_allocation.state_id     = getModelColumnById(SpZones,request.POST['zone_id'],'state_id')
                area_allocation.state_name   = getModelColumnById(SpZones,request.POST['zone_id'],'state_name')
                area_allocation.zone_id     = request.POST['zone_id']
                area_allocation.zone_name   = getModelColumnById(SpZones,request.POST['zone_id'],'zone')
                area_allocation.town_id     = request.POST['town_id']
                area_allocation.town_name   = getModelColumnById(SpTowns,request.POST['town_id'],'town')
                area_allocation.route_id    = request.POST['route_id']
                area_allocation.route_name  = getModelColumnById(SpRoutes,request.POST['route_id'],'route')
                area_allocation.save()
                
                try:
                    user_variants = SpUserProductVariants.objects.filter(user_id=request.POST['last_user_id'])
                except SpUserProductVariants.DoesNotExist:
                    user_variants = None

                user_details = SpUsers.objects.filter(id=request.POST['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]
                    
                context                  = {}
                context['last_user_id']  = request.POST['last_user_id']
                context['user_variants'] = user_variants
                context['user_details']  = user_details
                template                 = 'user-management/add-user-product-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)

# Edit User basic details View
@login_required
def editUserOfficalDetail(request):
    oganizations    = SpOrganizations.objects.filter(status=1)
    working_shifts  = TblClWorkingShifts.objects.all()
    zones           = SpZones.objects.all()
    routes          = SpRoutes.objects.all()

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_area_allocations   = SpUserAreaAllocations.objects.get(user_id=request.GET['last_user_id'])
    user_basic_details      = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    
    departments     = SpDepartments.objects.filter(organization_id=user_details.organization_id)
    roles           = SpRoles.objects.filter(department_id=user_details.department_id)#.filter(id__in=[8,9])
    towns           = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)
    
    context = {}
    context['oganizations']             = oganizations
    context['working_shifts']           = working_shifts
    context['zones']                    = zones
    context['routes']                   = routes
    context['user_details']             = user_details
    context['user_area_allocations']     = user_area_allocations
    context['user_basic_details']       = user_basic_details
    context['departments']              = departments
    context['roles']                    = roles
    context['towns']                    = towns
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-offical-detail.html'
    
    return render(request, template, context)      

# User product details View
@login_required
def addUserProductDetail(request):
    context = {}
    template = 'user-management/add-user-product-detail.html'
    if request.method == "POST":
        try:
            user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
        except SpUserDocuments.DoesNotExist:
            user_documents = None
        context['user_documents']   = user_documents
        context['last_user_id']     = request.POST['last_user_id']
        template = 'user-management/add-user-document-detail.html'
    return render(request, template, context)

# User product details View
@login_required
def editUserProductDetail(request):
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=request.GET['last_user_id'])
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    user_details = SpUsers.objects.filter(id=request.GET['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

    context = {}
    context['last_user_id']  = request.GET['last_user_id']
    context['user_variants'] = user_variants
    context['user_details']  = user_details
    template = 'user-management/add-user-product-detail.html'
    return render(request, template, context) 

# User document details View
@login_required
def addUserDocumentDetail(request):
    template = 'user-management/add-user-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhaar_card'] != '':
                        deleteMediaFile(request.POST['previous_aadhaar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                aadhaar_card_name = uploaded_aadhaar_card.name
                temp = aadhaar_card_name.split('.')
                aadhaar_card_name = 'aadhaar_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                aadhaar_card = storage.save(aadhaar_card_name, uploaded_aadhaar_card)
                aadhaar_card = storage.url(aadhaar_card)
            else:
                if request.POST['previous_aadhaar_card'] != '':
                        aadhaar_card = request.POST['previous_aadhaar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card'] != '':
                        deleteMediaFile(request.POST['previous_pan_card'])        
                uploaded_pan_card = request.FILES['pan_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                pan_card_name = uploaded_pan_card.name
                temp = pan_card_name.split('.')
                pan_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                pan_card = storage.save(pan_card_name, uploaded_pan_card)
                pan_card = storage.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                        pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None

            if bool(request.FILES.get('cin', False)) == True:
                if request.POST['previous_cin'] != '':
                        deleteMediaFile(request.POST['previous_cin'])
                uploaded_cin = request.FILES['cin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                cin_name = uploaded_cin.name
                temp = cin_name.split('.')
                cin_name = 'cin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                cin = storage.save(cin_name, uploaded_cin)
                cin = storage.url(cin)
            else:
                if request.POST['previous_cin'] != '':
                        cin = request.POST['previous_cin'] 
                else:
                    cin = None

            if bool(request.FILES.get('gstin', False)) == True:
                if request.POST['previous_gstin'] != '':
                        deleteMediaFile(request.POST['previous_gstin'])
                uploaded_gstin = request.FILES['gstin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                gstin_name = uploaded_gstin.name
                temp = gstin_name.split('.')
                gstin_name = 'gstin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                gstin = storage.save(gstin_name, uploaded_gstin)
                gstin = storage.url(gstin)
            else:
                if request.POST['previous_gstin'] != '':
                        gstin = request.POST['previous_gstin'] 
                else:
                    gstin = None

            if bool(request.FILES.get('fssai', False)) == True:
                if request.POST['previous_fssai'] != '':
                        deleteMediaFile(request.POST['previous_fssai'])
                uploaded_fssai = request.FILES['fssai']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                fssai_name = uploaded_fssai.name
                temp = fssai_name.split('.')
                fssai_name = 'fssai_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                fssai = storage.save(fssai_name, uploaded_fssai)
                fssai = storage.url(fssai)
            else:
                if request.POST['previous_fssai'] != '':
                        fssai = request.POST['previous_fssai'] 
                else:
                    fssai = None        
            
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None

            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error']            = True
            response['message']          = e
            return HttpResponse(e)
    return render(request, template)
    


# Employee basic details View
@login_required
def addEmployeeBasicDetail(request):
    contact_types = SpContactTypes.objects.filter(status=1)
    countries = TblCountry.objects.filter()
    country_codes   = SpCountryCodes.objects.filter(status=1)
    #print(countries)
    context = {}
    context['contact_types'] = contact_types
    context['countries']     = countries
    context['country_codes'] = country_codes
    template = 'user-management/add-employee-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password
            
            error_count = 0 
            user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
           
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
                return JsonResponse(response)
            else:
                
                user = SpUsers()
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.user_type = 1
                user.password       = make_password(str(password))
                user.plain_password = str(password)
                user.save()

                if request.POST['previous_profile_image']:
                    profile_image = request.POST['previous_profile_image']
                    im = Image.open(BytesIO(base64.b64decode(profile_image.split(",")[1])))
                    im.save("media/profileImage/employee_photo_"+str(user.id)+".png", 'PNG')
                    filePath="media/profileImage/employee_photo_" +str(user.id) + ".png"
                    profile_image = filePath 
                else:
                    profile_image = None
                

                user.profile_image = profile_image
                user.save()
                last_user_id = user.id
                #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()

                basic = SpBasicDetails()
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(TblCountry, request.POST['store_country_id'],'country_name')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(TblStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(TblNewDistrict, request.POST['store_city_id'],'district_name')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(TblCountry, request.POST['permanent_country_id'],'country_name')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(TblStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(TblNewDistrict, request.POST['permanent_city_id'],'district_name')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()


                oganizations                = SpOrganizations.objects.filter(status=1)
                working_shifts              = TblClWorkingShifts.objects.all()
                
                context                     = {}
                context['banks'] =   SpBanks.objects.all()
                contract_type    = ContractType.objects.all()
                context['working_shifts']   = working_shifts
                context['contract_type']    = contract_type
                context['oganizations']     = oganizations
                
                context['last_user_id']     = last_user_id

                template = 'user-management/add-employee-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    return render(request, template, context)




# # Employee basic details View
# @login_required
# def addEmployeeBasicDetail(request):
#     contact_types = SpContactTypes.objects.filter(status=1)
#     countries = TblCountry.objects.filter()
#     country_codes   = SpCountryCodes.objects.filter(status=1)
#     print(countries)
#     context = {}
#     context['contact_types'] = contact_types
#     context['countries']     = countries
#     context['country_codes'] = country_codes
#     template = 'user-management/add-employee-basic-detail.html'
#     response = {}
#     error_response = {}
#     if request.method == "POST":
#         try:
#             password = SpUsers.objects.make_random_password()
#             user_context = {}
#             user_context['first_name']      = request.POST['first_name']
#             user_context['middle_name']     = request.POST['middle_name']
#             user_context['last_name']       = request.POST['last_name']
#             user_context['official_email']  = request.POST['official_email']
#             user_context['password']        = password
            
#             error_count = 0 
#             user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
#             if user_exists:
#                 error_count = 1
#                 error_response['emailId_error'] = "Email already exists"
#             if(error_count > 0):
#                 response['error'] = True
#                 response['message'] = error_response
#                 return JsonResponse(response)
#             else:    
#                 # if request.FILES['profile_image']:
#                 #     uploaded_profile_image = request.FILES['profile_image']
#                 #     pfs = FileSystemStorage()
#                 #     profile_image = pfs.save(uploaded_profile_image.name, uploaded_profile_image)

#                 user = SpUsers()
#                 user.salutation = request.POST['salutation']
#                 user.first_name = request.POST['first_name']
#                 user.middle_name = request.POST['middle_name']
#                 user.last_name = request.POST['last_name']
#                 user.official_email = request.POST['official_email']
#                 user.user_type = 1
#                 user.password       = make_password(str(password))
#                 user.plain_password = str(password)
#                 user.save()
#                 if request.POST['previous_profile_image']:
#                     profile_image = request.POST['previous_profile_image']
#                     im = Image.open(BytesIO(base64.b64decode(profile_image.split(",")[1])))
#                     im.save("media/profileImage/employee_photo_"+str(user.id)+".png", 'PNG')
#                     filePath="media/profileImage/employee_photo_" +str(user.id) + ".png"
#                     profile_image = filePath 
#                 else:
#                     profile_image = None

#                 user.profile_image = profile_image
#                 user.save()

#                 last_user_id = user.id
#                 #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
            
#                 country_codes       = request.POST.getlist('country_code[]') 
#                 contact_person_name = request.POST.getlist('contact_person_name[]')
#                 contact_types       = request.POST.getlist('contact_type[]')
#                 contact_nos         = request.POST.getlist('contact_no[]')
#                 is_primary          = request.POST.getlist('primary_contact[]')

#                 for id, val in enumerate(contact_nos):
#                     if is_primary[id] == '1':
#                         user_data = SpUsers.objects.get(id=last_user_id)
#                         user_data.primary_contact_number = contact_nos[id]
#                         user_data.save()

#                     user_contact_no = SpContactNumbers()
#                     user_contact_no.user_id = last_user_id
#                     if country_codes[id]!='':
#                         user_contact_no.country_code = country_codes[id]
#                     if contact_types[id]!='':    
#                         user_contact_no.contact_type = contact_types[id]
#                         user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
#                     if contact_nos[id]!='':    
#                         user_contact_no.contact_number = contact_nos[id]
#                     if is_primary[id]!='':    
#                         user_contact_no.is_primary = is_primary[id]
#                     user_contact_no.save()

#                 basic = SpBasicDetails()
#                 basic.user_id               = last_user_id
#                 basic.father_name           = request.POST['father_name']
#                 basic.mother_name           = request.POST['mother_name']
#                 basic.gender                = request.POST['user_gender']
#                 basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
#                 basic.blood_group           = request.POST['blood_group']
#                 basic.save()

            
#                 correspondence = SpAddresses()
#                 correspondence.user_id          = last_user_id
#                 correspondence.type             = 'correspondence'
#                 correspondence.address_line_1   = request.POST['store_address_line_1']
#                 correspondence.address_line_2   = request.POST['store_address_line_2']
#                 correspondence.country_id       = request.POST['store_country_id']
#                 correspondence.country_name     = getModelColumnById(TblCountry, request.POST['store_country_id'],'country_name')
#                 correspondence.state_id         = request.POST['store_state_id']
#                 correspondence.state_name       = getModelColumnById(TblStates, request.POST['store_state_id'],'state')
#                 correspondence.city_id          = request.POST['store_city_id']
#                 correspondence.city_name        = getModelColumnById(TblNewDistrict, request.POST['store_city_id'],'district_name')
#                 correspondence.pincode          = request.POST['store_pincode']
#                 correspondence.save()

#                 permanent = SpAddresses()
#                 permanent.user_id = last_user_id
#                 permanent.type = 'permanent'
#                 permanent.address_line_1    = request.POST['permanent_address_line_1']
#                 permanent.address_line_2    = request.POST['permanent_address_line_2']
#                 permanent.country_id        = request.POST['permanent_country_id']
#                 permanent.country_name      = getModelColumnById(TblCountry, request.POST['permanent_country_id'],'country_name')
#                 permanent.state_id          = request.POST['permanent_state_id']
#                 permanent.state_name        = getModelColumnById(TblStates, request.POST['permanent_state_id'],'state')
#                 permanent.city_id           = request.POST['permanent_city_id']
#                 permanent.city_name         = getModelColumnById(TblNewDistrict, request.POST['permanent_city_id'],'district_name')
#                 permanent.pincode           = request.POST['permanent_pincode']
#                 permanent.save()


#                 oganizations                = SpOrganizations.objects.filter(status=1)
#                 working_shifts              = SpWorkingShifts.objects.all()
               

#                 context                     = {}
#                 context['oganizations']     = oganizations
#                 context['working_shifts']   = working_shifts
#                 context['last_user_id']     = last_user_id

#                 template = 'user-management/add-employee-offical-detail.html'
#                 return render(request, template, context)
#         except Exception as e:
#             response['error'] = True
#             response['message'] = e
#             return HttpResponse(str(e))
#     return render(request, template, context)


# # Employee offical details View
# @login_required
# def addEmployeeOfficalDetail(request):
#     template = 'user-management/add-employee-offical-detail.html'
    
#     response = {}
#     error_response = {}
#     if request.method == "POST":
#         try:
#             error_count = 0
#             emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
#             if emp_sap_id_exists:
#                 error_count = 1
#                 error_response['emp_sap_id_error'] = "Employee ID already exists"

#             if(error_count > 0):
#                 response['error'] = True
#                 response['message'] = error_response

#                 return JsonResponse(response)
#             else: 
#                 user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

#                 # update user role
#                 if user_data.role_id is None :
#                     updateUserRole(user_data.id,request)
#                 if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
#                     updateUserRole(user_data.id,request)

#                 user_data.organization_id       = request.POST['organization_id']
                
#                 user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
#                 user_data.department_id         = request.POST['department_id']
#                 user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
#                 user_data.role_id               = request.POST['role_id']
#                 user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
#                 user_data.emp_sap_id            = request.POST['emp_sap_id']

#                 if request.POST['reporting_to_id']:
#                     user_data.reporting_to_id   = request.POST['reporting_to_id']
#                     user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

#                 user_data.save()

#                 user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
#                 user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
#                 user_basic_details.pan_number           = request.POST['pan_number']
#                 user_basic_details.working_shift_id     = request.POST['working_shift_id']
#                 user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
#                 user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
#                 user_basic_details.save()

#                 if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
#                     banks = SpBankDetails()
#                     banks.user_id = request.POST['last_user_id']
#                     banks.bank_id = request.POST['bank_name']
#                     banks.bank_name = getModelColumnById(
#                         SpBanks, request.POST['bank_name'], 'bank_name')
#                     banks.bank_account_no = request.POST['bank_account_no']
#                     banks.ifsc_code = request.POST['ifsc_code']
#                     banks.bank_address = request.POST['bank_address']
#                     banks.save()
#                 # towns    = request.POST.getlist('town_id[]')
#                 # for id, val in enumerate(towns):
#                 #     area_allocation = SpUserAreaAllocations()
#                 #     area_allocation.user_id = request.POST['last_user_id']
#                 #     if towns[id] != '':
#                 #         zone_id = getModelColumnById(SpTowns,towns[id],'zone_id')
#                 #         area_allocation.zone_id                 =   zone_id
#                 #         area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
#                 #         area_allocation.state_id     = getModelColumnById(SpZones,zone_id,'state_id')
#                 #         area_allocation.state_name   = getModelColumnById(SpZones,zone_id,'state_name')
#                 #         area_allocation.town_id = towns[id]
#                 #         area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
#                 #     area_allocation.save()

#                 context = {}
#                 context['last_user_id'] = request.POST['last_user_id']
#                 distributors = 0
#                 # distributors = SpUsers.objects.raw(''' select sp_users.id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
#                 # from sp_user_area_allocations 
#                 # left join sp_users on sp_users.id = sp_user_area_allocations.user_id
#                 # where sp_user_area_allocations.town_id in 
#                 # (select town_id from sp_user_area_allocations as sura where sura.user_id = %s ) 
#                 # and  (sp_users.is_distributor = %s or sp_users.is_super_stockist = %s)
#                 # ''',[request.POST['last_user_id'], 1,1])
                
#                 # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
#                 # from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
#                 if distributors:
#                     context['distributors'] = distributors
#                 else:
#                     context['distributors'] = None
#                 template = 'user-management/add-employee-biometric-details.html'
#                 return render(request, template, context)
#         except Exception as e:
#             response['error'] = True
#             response['message'] = e
#             return HttpResponse(e)
#     return render(request, template,response)

# def mapUserLeavess(request,role_id,user_id):
#     mapUserLeaves(role_id,user_id)
    
def mapUserLeaves(role_id,user_id):
    try:
        leave_policy_id = SpRoleEntityMapping.objects.get(role_id = role_id , entity_type = "leave_policy")
        leave_policy_id = leave_policy_id.entity_id
    except SpRoleEntityMapping.DoesNotExist:
        leave_policy_id = None
    if leave_policy_id:
        leave_policy_dettail = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id)
        for policy_detail in leave_policy_dettail:
            leave_polcy_ledger = SpUserLeavePolicyLedger()
            leave_polcy_ledger.user_id = user_id
            leave_polcy_ledger.leave_policy_id = leave_policy_id
            leave_polcy_ledger.leave_type_id = policy_detail.leave_type_id
            current_month = datetime.today().strftime('%m')
            if int(current_month) == 1:
                month_leave_count = policy_detail.year_leave_count / 12
                leave_polcy_ledger.year_leave_count = policy_detail.year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
            else:   
                current_month = 12 - int(current_month)
                sub_leave_count = policy_detail.year_leave_count / 12
                year_leave_count = sub_leave_count*current_month
                month_leave_count = year_leave_count  / current_month
                leave_polcy_ledger.year_leave_count = year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
            
            leave_polcy_ledger.consecutive_leave = policy_detail.consecutive_leave 
            leave_polcy_ledger.save()
  


@login_required
def addEmployeeOfficalDetail(request):
    template = 'user-management/add-employee-offical-detail.html'
    
    response = {}
    error_response = {}
    contract_type    = ContractType.objects.all()
    response['contract_type']    = contract_type

    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else: 
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                user_data.login_status         = request.POST['login_status']
                
                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()
                user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                # user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                # user_basic_details.pan_number           = request.POST['pan_number']
                # user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.contract_type        = request.POST['contract_type']
                # user_basic_details.working_shift_name   = getModelColumnById(TblClWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['date_of_relieving']:
                    user_basic_details.date_of_relieving      = datetime.strptime(request.POST['date_of_relieving'], '%d/%m/%Y').strftime('%Y-%m-%d') 
                # if request.POST['total_leave']!='':
                       
                #         user_basic_details.leave_count    = request.POST['total_leave']
                # else:
                #          user_basic_details.leave_count = 0
                user_basic_details.week_of_day= ','.join([str(elem) for elem in request.POST.getlist('week_of_day[]')]) 
                user_basic_details.save()
                
                role_id = request.POST['role_id']
                if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
                    mapUserLeaves(role_id,request.POST['last_user_id'])
                
                # TblClAllocatedShifts.objects.filter(user_id = request.POST['last_user_id']).delete()
                # for each_shift in request.POST.getlist('working_shift_id[]'):
                #     working_shift = TblClAllocatedShifts()
                #     working_shift.user_id = request.POST['last_user_id']
                #     working_shift.working_shift_id = each_shift
                #     working_shift.save()
                    
                role_id = request.POST['role_id']
                if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
                    mapUserLeaves(role_id,request.POST['last_user_id'])
                    
                    
                if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
                    banks = SpBankDetails()
                    banks.user_id = request.POST['last_user_id']
                    banks.bank_id = request.POST['bank_name']
                    banks.bank_name = getModelColumnById(
                        SpBanks, request.POST['bank_name'], 'bank_name')
                    banks.bank_account_no = request.POST['bank_account_no']
                    banks.ifsc_code = request.POST['ifsc_code']
                    banks.bank_address = request.POST['bank_address']
                    banks.save()
                # towns    = request.POST.getlist('town_id[]')
                # for id, val in enumerate(towns):
                #     area_allocation = SpUserAreaAllocations()
                #     area_allocation.user_id = request.POST['last_user_id']
                #     if towns[id] != '':
                #         zone_id = getModelColumnById(SpTowns,towns[id],'zone_id')
                #         area_allocation.zone_id                 =   zone_id
                #         area_allocation.zone_name               = getModelColumnById(SpZones,zone_id,'zone')
                #         area_allocation.state_id     = getModelColumnById(SpZones,zone_id,'state_id')
                #         area_allocation.state_name   = getModelColumnById(SpZones,zone_id,'state_name')
                #         area_allocation.town_id = towns[id]
                #         area_allocation.town_name               = getModelColumnById(SpTowns,towns[id],'town')
                #     area_allocation.save()

                context = {}
                context['last_user_id'] = request.POST['last_user_id']
                context['working_shifts'] = TblClWorkingShifts.objects.all()
                
                distributors = 0
                # distributors = SpUsers.objects.raw(''' select sp_users.id,sp_users.first_name, sp_users.middle_name, sp_users.last_name 
                # from sp_user_area_allocations 
                # left join sp_users on sp_users.id = sp_user_area_allocations.user_id
                # where sp_user_area_allocations.town_id in 
                # (select town_id from sp_user_area_allocations as sura where sura.user_id = %s ) 
                # and  (sp_users.is_distributor = %s or sp_users.is_super_stockist = %s)
                # ''',[request.POST['last_user_id'], 1,1])
                
                # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
                # from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
              
                
             
                if distributors:
                    context['distributors'] = distributors
                else:
                    context['distributors'] = None
                template = 'user-management/add-employee-biometric-details.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template,response)



# Employee attendance details View
@login_required
def addEmployeeAttendanceDetail(request):
    if request.method == "POST":

        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = request.POST['last_user_id']
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()

        context = {}
        context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-document-detail.html'
        return render(request, template, context)
    else:
        context = {}
        distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
            from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        if distributors:
            context['distributors'] = distributors
        else:
            context['distributors'] = None
            context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-attendance.html'
        return render(request, template, context)

    

# Employee document details View
@login_required
def addEmployeeDocumentDetail(request):
    template = 'user-management/add-employee-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                pan_card = None

            documents               = SpUserDocuments()
            documents.user_id       = request.POST['last_user_id']
            documents.aadhaar_card  = aadhaar_card
            documents.pan_card      = pan_card
            documents.save()
            
            #-----------------------------notify android block-------------------------------#
            organization_id = getModelColumnById(SpUsers, request.POST['last_user_id'], 'organization_id')
            all_users = SpUsers.objects.filter(organization_id = organization_id).exclude(id=request.POST['last_user_id']).values_list("id", flat=True)
            for each_user in all_users:
                user_id = each_user
                if SpUsers.objects.filter(id=user_id, firebase_token__isnull=False).exists():
                    userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
                    employee_name = getUserName(request.user.id)
    
                    message_title = document.document_name+" document scan request."
                    message_body =  document.document_name+" document scan request has been generated by "+employee_name
                    notification_image = ""
        
                    if userFirebaseToken is not None and userFirebaseToken != "" :
                        registration_ids = []
                        registration_ids.append(userFirebaseToken)
                        data_message = {}
                        data_message['id'] = 1
                        data_message['status'] = 'notification'
                        data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image'] = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

            response['error'] = False
            response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)


#csrf_exempt
@login_required
def addEmployeePhoto(request):
    context = {}
    return render(request,"user-management/add-employee-photo.html",context)


# Employee basic details View
@login_required
def editEmployeeBasicDetail(request, employee_id):
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            else:
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
  
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if request.POST['previous_profile_image']:
                    if not SpUsers.objects.filter(id = request.POST['last_user_id'], profile_image = request.POST['previous_profile_image']).exists():
                        profile_image = request.POST['previous_profile_image']
                        im = Image.open(
                            BytesIO(base64.b64decode(profile_image.split(",")[1])))
                        im.save("media/profileImage/employee_photo_" +
                                str(employee_id)+".png", 'PNG')
                        filePath = "media/profileImage/employee_photo_" + \
                            str(employee_id) + ".png"
                        profile_image = filePath
                    else:
                        profile_image = request.POST['previous_profile_image']
                else:
                    profile_image = None

                SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                contact_nos         = request.POST.getlist('contact_no[]')
                user = SpUsers.objects.get(id=request.POST['last_user_id'])
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.profile_image = profile_image
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.primary_contact_number = contact_nos[0]
                user.save()
                last_user_id = request.POST['last_user_id']

            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                try:
                    basic   = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    basic  = None
                if basic:
                    basic                   = SpBasicDetails.objects.get(user_id=last_user_id)
                else:
                    basic                   = SpBasicDetails()

                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(TblCountry, request.POST['store_country_id'],'country_name')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(TblStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(TblNewDistrict, request.POST['store_city_id'],'district_name')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(TblCountry, request.POST['permanent_country_id'],'country_name')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(TblStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(TblNewDistrict, request.POST['permanent_city_id'],'district_name')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                response['error'] = False
                response['last_user_id'] = last_user_id
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))

    else : 
        contact_types = SpContactTypes.objects.filter(status=1)
        countries = TblCountry.objects.all()
        country_codes   = SpCountryCodes.objects.filter(status=1)

        context = {}
        context['contact_types'] = contact_types
        context['countries']     = countries
        context['country_codes'] = country_codes

        employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
        sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        where sp_users.id = %s''',[employee_id])
        
        try:
            employee_correspondence_address = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
        except SpAddresses.DoesNotExist:
            employee_correspondence_address = None

        try:
            employee_permanent_address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
        except SpAddresses.DoesNotExist:
            employee_permanent_address = None

        try:
            employee_contact_numbers = SpContactNumbers.objects.filter(user_id=employee_id)
        except SpContactNumbers.DoesNotExist:
            employee_contact_numbers = None    

        employee_user_allocation = None

        try:
            if employee_correspondence_address is not None:
                employee_store_states = TblStates.objects.filter(country_id=employee_correspondence_address.country_id)
            else:
                employee_store_states = None    
        except TblStates.DoesNotExist:
            employee_store_states = None    

        try:
            if employee_correspondence_address is not None:
                employee_store_cities = TblNewDistrict.objects.filter(state_id=employee_correspondence_address.state_id)
            else:
                employee_store_cities = None
        except TblNewDistrict.DoesNotExist:
            employee_store_cities = None 
        
        try:
            if employee_permanent_address is not None:
                employee_permanent_states = TblStates.objects.filter(country_id=employee_permanent_address.country_id)
            else:
                employee_permanent_states = None    
        except TblStates.DoesNotExist:
            employee_permanent_states = None

        try:
            if employee_permanent_address is not None:
                employee_permanent_cities = TblNewDistrict.objects.filter(state_id=employee_permanent_address.state_id)
            else:
                employee_permanent_cities = None    
        except TblNewDistrict.DoesNotExist:
            employee_permanent_cities = None
        try:
            employee_contact_numbers = SpContactNumbers.objects.filter(user_id=employee_id)
        except SpContactNumbers.DoesNotExist:
            employee_contact_numbers = None   

        context['user_contacts']                    = employee_contact_numbers



        if employee:
            context['employee'] = employee[0]
            context['employee_correspondence_address']  = employee_correspondence_address
            context['employee_permanent_address']       = employee_permanent_address
            context['user_contacts']                    = employee_contact_numbers
            context['user_areas']                       = employee_user_allocation
            context['store_states']                     = employee_store_states
            context['store_cities']                     = employee_store_cities
            context['permanent_states']                 = employee_permanent_states
            context['permanent_cities']                 = employee_permanent_cities
            context['last_user_id']                     = employee_id
            try:
                user_documents                              = SpUserDocuments.objects.get(user_id=employee_id)
                context['user_documents'] = user_documents
            except SpUserDocuments.DoesNotExist:
                context['user_documents'] = None

            context['user_attendance_locations'] = None
            template = 'user-management/edit-employee/employee-basic-detail.html'
            return render(request, template, context)
        else:
            return HttpResponse('Employee not found')




@login_required
def editEmployeeOfficalDetail(request,employee_id):
    error_response = {}
    response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['SAPID_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:  
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)
                    
                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name')
                user_data.emp_sap_id            = request.POST['emp_sap_id']  
                user_data.login_status          = request.POST['login_status']

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()

                if request.POST['last_user_id'] != '':
                    user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    # user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    # user_basic_details.pan_number           = request.POST['pan_number']
                    # user_basic_details.working_shift_id        = request.POST['working_shift_id']
                    user_basic_details.contract_type        = request.POST['contract_type']
                    # user_basic_details.geofencing        = request.POST['geofencing_type']
                    # user_basic_details.working_shift_name   = getModelColumnById(TblClWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    if request.POST['date_of_relieving']:
                        user_basic_details.date_of_relieving      = datetime.strptime(request.POST['date_of_relieving'], '%d/%m/%Y').strftime('%Y-%m-%d') 
                    user_basic_details.save()
                else:
                    user_basic_details                      = SpBasicDetails()
                    # user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    # user_basic_details.pan_number           = request.POST['pan_number']
                    # user_basic_details.working_shift_id        = request.POST['working_shift_id']
                    user_basic_details.contract_type        = request.POST['contract_type']
                    # user_basic_details.working_shift_name   = getModelColumnById(TblClWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    if request.POST['date_of_relieving']:
                        user_basic_details.date_of_relieving      = datetime.strptime(request.POST['date_of_relieving'], '%d/%m/%Y').strftime('%Y-%m-%d') 
                    
                # if request.POST['total_leave']!='' and request.POST['total_leave']:
                       
                #     user_basic_details.leave_count  = request.POST['total_leave']
                # else:
                #     user_basic_details.leave_count = 0 
                
                user_basic_details.week_of_day= ','.join([str(elem) for elem in request.POST.getlist('week_of_day[]')]) 
                    
                user_basic_details.save()
                    
                # TblClAllocatedShifts.objects.filter(user_id = request.POST['last_user_id']).delete()
                # for each_shift in request.POST.getlist('working_shift_id[]'):
                #     working_shift = TblClAllocatedShifts()
                #     working_shift.user_id = request.POST['last_user_id']
                #     working_shift.working_shift_id = each_shift
                #     working_shift.save()
                    
                if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
                    try:
                        banks                   =   SpBankDetails.objects.get(user_id = request.POST['last_user_id'])
                        banks.bank_id           =   request.POST['bank_name']
                        banks.bank_name         =   getModelColumnById(SpBanks,request.POST['bank_name'],'bank_name')
                        banks.bank_account_no   =   request.POST['bank_account_no']
                        banks.ifsc_code         =   request.POST['ifsc_code']
                        banks.bank_address      =   request.POST['bank_address']
                        banks.save()
                    except SpBankDetails.DoesNotExist:
                        if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
                            banks                   =   SpBankDetails()
                            banks.user_id           =   request.POST['last_user_id']
                            banks.bank_id           =   request.POST['bank_name']
                            banks.bank_name         =   getModelColumnById(SpBanks,request.POST['bank_name'],'bank_name')
                            banks.bank_account_no   =   request.POST['bank_account_no']
                            banks.ifsc_code         =   request.POST['ifsc_code']
                            banks.bank_address      =   request.POST['bank_address']
                            banks.save()
                else:
                    SpBankDetails.objects.filter(user_id=request.POST['last_user_id']).delete()
                
                
                response = {}
                response['error'] = False
                response['last_user_id'] = request.POST['last_user_id']
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    else:
        
        template = 'user-management/edit-employee/employee-offical-detail.html'
        context = {}

        employee_details = SpUsers.objects.get(id=employee_id)
        employee_area_allocations = None
        
        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        
        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        try:
            week_of_day = employee_basic_details.week_of_day.split(',')
        except:
            week_of_day = None
        
        # working_shifts = TblClWorkingShifts.objects.all()
        departments = SpDepartments.objects.filter(organization_id=employee_details.organization_id)
        try:
           roles = SpRoles.objects.filter(department_id=employee_details.department_id)
        except:
            roles=None

        try:
            reporting_role = SpRoles.objects.get(id = employee_details.role_id)
        except:
            reporting_role = None

        try:
            if reporting_role is not None:
                reporting_users = SpUsers.objects.filter(role_id=reporting_role.reporting_role_id)
            else:
                reporting_users = None
        except SpUsers.DoesNotExist:
            reporting_users = None
                 
        oganizations = SpOrganizations.objects.filter(status=1)

        contract_type   = ContractType.objects.all()

        try:
            employee_bank_details = SpBankDetails.objects.get(user_id=employee_id)
        except SpBankDetails.DoesNotExist:
            employee_bank_details = None
        banks                                   = SpBanks.objects.all().order_by('bank_name')
        # allocat_shift                           = TblClAllocatedShifts.objects.filter(user_id=employee_id).values_list("working_shift_id", flat=True)

        context['banks']                        = banks
        context['employee_bank_details']        = employee_bank_details
        context['oganizations']                 = oganizations
        # context['working_shifts']               = working_shifts
        # context['allocat_shifts']               = allocat_shift
        context['contract_type']                = contract_type
        context['week_of_days']                 = week_of_day
        context['employee_details']             = employee_details
        context['employee_area_allocations']    = employee_area_allocations
        context['employee_basic_details']       = employee_basic_details
        context['departments']                  = departments
        context['roles']                        = roles
        context['reporting_users']              = reporting_users 
        context['last_user_id']                 = employee_id
        
        return render(request, template,context)




# Employee attendance details View
@login_required
def editEmployeeAttendanceDetail(request,employee_id):
    if request.method == "POST":
        SpUserAttendanceLocations.objects.filter(user_id=request.POST['last_user_id']).delete()
        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = employee_id
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()
                
        response = {}
        response['error'] = False
        response['last_user_id'] = request.POST['last_user_id']
        return JsonResponse(response)
    else:
        context = {}
        # user_attendance_locations = SpUserAttendanceLocations.objects.filter(user_id=employee_id)
        # if user_attendance_locations:
        #     user_attendance_locations = user_attendance_locations
        # else:
        #     user_attendance_locations = None
        # context['user_attendance_locations'] = user_attendance_locations
        # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
        #     from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        # if distributors:
        #     context['distributors'] = distributors
        # else:
        #     context['distributors'] = None

        # context['last_user_id'] = employee_id
        # context['user_attendance_locations'] = user_attendance_locations
        template = 'user-management/edit-employee/employee-biometric-details.html'
        return render(request, template, context)

# Employee document details View
@login_required
def editEmployeeDocumentDetail(request,employee_id):
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhar_card']:
                    deleteMediaFile(request.POST['previous_aadhar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                aadhaar_card_name = uploaded_aadhaar_card.name
                temp = aadhaar_card_name.split('.')
                aadhaar_card_name = 'aadhaar_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                aadhaar_card = storage.save(aadhaar_card_name, uploaded_aadhaar_card)
                aadhaar_card = storage.url(aadhaar_card)
            else:
                if request.POST['previous_aadhar_card'] != '':
                    aadhaar_card = request.POST['previous_aadhar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card']:
                    deleteMediaFile(request.POST['previous_pan_card'])  
                uploaded_pan_card = request.FILES['pan_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                pan_card_name = uploaded_pan_card.name
                temp = pan_card_name.split('.')
                pan_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                pan_card = storage.save(pan_card_name, uploaded_pan_card)
                pan_card = storage.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                    pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None
                    
           
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None
            
            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."
            
            return JsonResponse(response)

        except Exception as e:
            return HttpResponse(e)
    else:
        try:
            employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
        except SpUserDocuments.DoesNotExist:
            employee_documents = None

        template = 'user-management/edit-employee/employee-document-detail.html'
        response = {}
        response['last_user_id'] = employee_id
        response['employee_documents'] = employee_documents
        return render(request, template,response)




@login_required
def userShortDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender,
    sp_basic_details.outstanding_amount,sp_basic_details.opening_crates,
    sp_addresses.address_line_1,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,
    sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_addresses on sp_addresses.user_id = sp_users.id   
    where sp_addresses.type=%s  and sp_users.id = %s ''',['correspondence',user_id])
    if user :
        context['user'] = user[0]
        context['contact_persons'] = SpContactPersons.objects.filter(user_id=user_id)
    else : 
        context['user'] = []
      
    template = 'user-management/user-short-details.html'
    return render(request, template,context)

@login_required
def userDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.personal_email,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.date_of_joining,sp_basic_details.working_shift_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[user_id])[0]
    
    try:
        documents = SpUserDocuments.objects.get(user_id=user_id)
    except SpUserDocuments.DoesNotExist:
        documents = None

    try:
        allocation = SpUserAreaAllocations.objects.get(user_id=user_id)
    except SpUserAreaAllocations.DoesNotExist:
        allocation = None
            
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=user_id)
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    try:
        permanent_address = SpAddresses.objects.get(user_id=user_id,type='permanent')
    except SpAddresses.DoesNotExist:
        permanent_address = None    

    context['user'] = user
    context['user_correspondence_address'] = SpAddresses.objects.get(user_id=user_id,type='correspondence')
    context['user_permanent_address']      = permanent_address
    context['contact_persons']             = SpContactPersons.objects.filter(user_id=user_id)
    context['user_contacts']               = SpContactNumbers.objects.filter(user_id=user_id)
    context['area_allocated']              = allocation 
    context['user_documents']              = documents
    context['user_variants']               = user_variants
    context['user_attendance_locations']   = SpUserAttendanceLocations.objects.filter(user_id=user_id,status=1)
    template = 'user-management/user-details.html'

    return render(request, template,context)


@login_required
def employeeShortDetail(request,employee_id):
    context = {}
    employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type,sp_basic_details.gender,sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join contract_type on contract_type.id=sp_basic_details.contract_type
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id = %s ''',[1,'correspondence',employee_id])
    if employee :
        context['employee'] = employee[0]
    else : 
        context['employee'] = []

    try:
        address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    except SpAddresses.DoesNotExist:
        address = None

    context['employee_permanent_address'] = address    
    template = 'user-management/employee-short-details.html'
    return render(request, template,context)

# @login_required
# def employeeDetail(request,employee_id):
#     context = {}
#     context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
#     sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
#      FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
#     where sp_users.id = %s''',[employee_id])[0]

#     try:
#         documents = SpUserDocuments.objects.get(user_id=employee_id)
#     except SpUserDocuments.DoesNotExist:
#         documents = None

#     try:
#         user_contacts = SpContactNumbers.objects.filter(user_id=employee_id)
#     except SpContactNumbers.DoesNotExist:
#         user_contacts = None
#     try:
#         bank_details = SpBankDetails.objects.get(user_id=employee_id)
#     except SpBankDetails.DoesNotExist:
#         bank_details = None
     
#     context['bank_details']                     = bank_details
#     context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
#     context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
#     context['user_contacts']                    = user_contacts
#     context['user_areas']                       = None
#     context['user_documents']                   = documents
#     context['user_coordinates']                 = None
#     context['user_attendance_locations']        = None
#     context['google_app_key']                   = getConfigurationResult('google_app_key')
#     template = 'user-management/employee-details.html'

#     return render(request, template,context)


# @login_required
# def employeeDetail(request,employee_id):
#     context = {}
#     context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
#     sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
#      FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
#     where sp_users.id = %s''',[employee_id])[0]

#     context['contractval'] = contractval = SpBasicDetails.objects.raw('''SELECT sp_basic_details.*,contract_type.id,contract_type.contract_type as contract_name 
#     FROM sp_basic_details left join contract_type on contract_type.id = sp_basic_details.contract_type
#     where sp_basic_details.user_id  = %s''',[employee_id])

#     try:
#         documents = SpUserDocuments.objects.get(user_id=employee_id)
#     except SpUserDocuments.DoesNotExist:
#         documents = None

#     try:
#         user_contacts = SpContactNumbers.objects.filter(user_id=employee_id)
#     except SpContactNumbers.DoesNotExist:
#         user_contacts = None

#     try:
#         bank_details = SpBankDetails.objects.get(user_id=employee_id)
#     except SpBankDetails.DoesNotExist:
#         bank_details = None
#     try:
#         contract_type = SpBasicDetails.objects.get(user_id=employee_id)
#     except SpBasicDetails.DoesNotExist:
#         contract_type = None

#     try:
#         contracttype =  ContractType.objects.get(id=contract_type.contract_type)
#     except ContractType.DoesNotExist:
#          contracttype = None

#     context['contract_type']                     = contracttype
#     context['bank_details']                     = bank_details
#     context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
#     context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
#     context['user_contacts']                    = user_contacts
#     context['user_areas']                       = None
#     context['user_documents']                   = documents
#     context['user_coordinates']                 = None
#     context['user_attendance_locations']        = None
#     context['google_app_key']                   = getConfigurationResult('google_app_key')
#     template = 'user-management/employee-details.html'

#     return render(request, template,context)
# @login_required
# def employeeDetail(request,employee_id):
#     context = {}
#     context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,sp_basic_details.geofencing,
#     sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
#      FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
#     where sp_users.id = %s''',[employee_id])[0]

#     context['contractval'] = contractval = SpBasicDetails.objects.raw('''SELECT sp_basic_details.*,contract_type.id,contract_type.contract_type as contract_name 
#     FROM sp_basic_details left join contract_type on contract_type.id = sp_basic_details.contract_type
#     where sp_basic_details.user_id  = %s''',[employee_id])
    
#     context['workingshift'] = workingshift = TblClAllocatedShifts.objects.raw('''SELECT tbl_cl_allocated_shifts.*,tbl_cl_working_shifts.id, tbl_cl_working_shifts.working_shift ,tbl_cl_working_shifts.start_timing,tbl_cl_working_shifts.end_timing 
#     FROM tbl_cl_allocated_shifts left join tbl_cl_working_shifts on tbl_cl_working_shifts.id = tbl_cl_allocated_shifts.working_shift_id
#     where tbl_cl_allocated_shifts.user_id  = %s''',[employee_id])
#     hours_sum = timedelta(days=0,seconds=0,minutes=0,hours=0)
#     for workingshifts in workingshift:
#         FMT = '%H:%M:%S'
#         starttime = workingshifts.start_timing.strftime("%H:%M:%S")
#         endtime = workingshifts.end_timing.strftime("%H:%M:%S")
#         tdelta = datetime.strptime(endtime, FMT) - datetime.strptime(starttime,FMT)
#         hours_sum += tdelta
#     totalhours = str(hours_sum).split(', ')

#     try:
#         documents = TblClEmployeeDocuments.objects.filter(employee_id=employee_id)
#     except TblClEmployeeDocuments.DoesNotExist:
#         documents = None

#     try:
#         user_contacts = SpContactNumbers.objects.filter(user_id=employee_id)
#     except SpContactNumbers.DoesNotExist:
#         user_contacts = None

#     try:
#         bank_details = SpBankDetails.objects.get(user_id=employee_id)
#     except SpBankDetails.DoesNotExist:
#         bank_details = None
#     try:
#         contract_type = SpBasicDetails.objects.get(user_id=employee_id)
#     except SpBasicDetails.DoesNotExist:
#         contract_type = None

#     try:
#         contracttype =  ContractType.objects.get(id=contract_type.contract_type)
#     except ContractType.DoesNotExist:
#          contracttype = None

#     context['contract_type']                     = contracttype
#     context['bank_details']                      = bank_details
#     context['working_shift']                     = workingshift
#     if len(totalhours) > 0:
#         context['totalhours']                        = totalhours[1]
#     else:
#         context['totalhours']                        = "-"
#     context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
#     context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
#     context['user_contacts']                    = user_contacts
#     context['user_areas']                       = None
#     context['user_documents']                   = documents
#     context['user_coordinates']                 = None
#     context['user_attendance_locations']        = None
#     context['google_app_key']                   = getConfigurationResult('google_app_key')
#     template = 'user-management/employee-details.html'

#     return render(request, template,context)

@login_required
def employeeDetail(request,employee_id):
    context = {}
    context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,sp_basic_details.geofencing,
    sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.leave_count,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    where sp_users.id = %s''',[employee_id])[0]

    context['contractval'] = contractval = SpBasicDetails.objects.raw('''SELECT sp_basic_details.*,contract_type.id,contract_type.contract_type as contract_name 
    FROM sp_basic_details left join contract_type on contract_type.id = sp_basic_details.contract_type
    where sp_basic_details.user_id  = %s''',[employee_id])
    
    context['workingshift'] = workingshift = TblClAllocatedShifts.objects.raw('''SELECT tbl_cl_allocated_shifts.*,tbl_cl_working_shifts.id, tbl_cl_working_shifts.working_shift ,tbl_cl_working_shifts.start_timing,tbl_cl_working_shifts.end_timing 
    FROM tbl_cl_allocated_shifts left join tbl_cl_working_shifts on tbl_cl_working_shifts.id = tbl_cl_allocated_shifts.working_shift_id
    where tbl_cl_allocated_shifts.user_id  = %s''',[employee_id])
    hours_sum = timedelta(days=0,seconds=0,minutes=0,hours=0)
    for workingshifts in workingshift:
        FMT = '%H:%M:%S'
        starttime = workingshifts.start_timing.strftime("%H:%M:%S")
        endtime = workingshifts.end_timing.strftime("%H:%M:%S")
        tdelta = datetime.strptime(endtime, FMT) - datetime.strptime(starttime,FMT)
        hours_sum += tdelta
    totalhours = str(hours_sum).split(', ')

    try:
        documents = TblClEmployeeFolderFiles.objects.filter(employee_id=employee_id)
    except TblClEmployeeDocuments.DoesNotExist:
        documents = None

    try:
        user_contacts = SpContactNumbers.objects.filter(user_id=employee_id)
    except SpContactNumbers.DoesNotExist:
        user_contacts = None

    try:
        bank_details = SpBankDetails.objects.get(user_id=employee_id)
    except SpBankDetails.DoesNotExist:
        bank_details = None
    try:
        contract_type = SpBasicDetails.objects.get(user_id=employee_id)
    except SpBasicDetails.DoesNotExist:
        contract_type = None

    try:
        contracttype =  ContractType.objects.get(id=contract_type.contract_type)
    except ContractType.DoesNotExist:
         contracttype = None
         
    leave_ledger = SpUserLeavePolicyLedger.objects.filter(user_id = employee_id)
    for leave in leave_ledger:
        leave.leave_policy_name = getModelColumnById(SpLeavePolicies,leave.leave_policy_id,'leave_policy')
        leave.laave_type_name = getModelColumnById(SpLeaveTypes,leave.leave_type_id,'leave_type')
        leave.month_leave_counts = round(leave.month_leave_count,1)

    context['contract_type']                     = contracttype
    context['bank_details']                      = bank_details
    context['working_shift']                     = workingshift
    # if len(totalhours[1]) > 0:
    #     context['totalhours']                        = totalhours[1]
    # else:
    context['totalhours']                        = '-'
    context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
    context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    context['user_contacts']                    = user_contacts
    context['user_areas']                       = None
    context['user_documents']                   = documents
    context['leave_ledger']                     = leave_ledger
    context['user_coordinates']                 = None
    context['user_attendance_locations']        = None
    context['google_app_key']                   = getConfigurationResult('google_app_key')
    template = 'user-management/employee-details.html'

    return render(request, template,context)



def getGroupedTownOptions(request):
    options = ''
    zone_ids = request.POST['zone_ids'].split(',')
    zones = SpZones.objects.raw(''' select * from sp_zones where id in %s ''',[zone_ids])
    for zone in zones:
        towns = SpTowns.objects.filter(zone_id=zone.id)
        if towns:
            options += '<optgroup label="' + zone.zone + '">'
            for town in towns : 
                options += "<option value="+str(town.id)+">"+town.town+"</option>"
            options += '</optgroup>'
    
    return HttpResponse(options)

def getReportingUserOptions(request, role_id):
    reporting_role = SpRoles.objects.get(id=role_id)
    options = '<option value="">Select Reporting to User*</option>'
    reporting_users = SpUsers.objects.raw(''' select id,first_name,last_name from sp_users where role_id = %s and user_type = %s ''',[reporting_role.reporting_role_id,1])
    for reporting_user in reporting_users:
        options += "<option value="+str(reporting_user.id)+">"+reporting_user.first_name+" "+reporting_user.last_name+"</option>"
    
    return HttpResponse(options)

#update user status
@login_required
def updateUserStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUsers.objects.get(id=id)
            data.status = is_active
            data.save()

            if is_active == '1':
                status = 'Unblock'
            else:
                status = 'Block'
                
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status
            activity    = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#update user status
@login_required
def updateUserVariantPrice(request):

    response = {}
    if request.method == "POST":
        try:
            id              = request.POST.get('id')
            price           = request.POST.get('price')
            user_type       = request.POST.get('user_type')
            is_distributor  = request.POST.get('is_distributor')
            
            data = SpUserProductVariants.objects.get(id=id)
            if user_type == '2':
                if is_distributor == '1':
                    data.sp_distributor = price
                    previous_price = data.sp_distributor
                else:
                    data.sp_superstockist = price
                    previous_price = data.sp_superstockist    
            else:
                data.sp_employee = price
                previous_price = data.sp_employee 
    
            data.save()
            
            response['error'] = False
            response['message'] = "Record has been updated successfully."
            response['id'] = id
            response['price'] = price
            response['user_type'] = user_type
            response['is_distributor'] = is_distributor
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#export to excel operational user list
@login_required
def exportOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        try:
            user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)
        except SpBasicDetails.DoesNotExist:
            user.outstanding_amount = None
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Operational User'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]

    if 'outstanding_amount' in column_list:
        columns += [ 'Outstanding Amount' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        if user.outstanding_amount:
            outstanding_amount = user.outstanding_amount.outstanding_amount 
        else:
            outstanding_amount = ''   
        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]

        if 'outstanding_amount' in column_list:
            row += [ outstanding_amount ]           
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#export to pdf operational user list
@login_required
def exportOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#export to excel non-operational user list
@login_required
def exportNonOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=non-operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Non-Operational-Users'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]         
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

#export to pdf non operational user list
@login_required
def exportNonOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/non_operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Non-Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

# #export to excel employee list
# @login_required
# def exportEmployeeToXlsx(request, columns):
#     column_list = columns.split (",")
#     users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')
    
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=employees.xlsx'.format(
#         date=datetime.now().strftime('%Y-%m-%d'),
#     )
#     workbook = Workbook()

#     # Define some styles and formatting that will be later used for cells
#     header_font = Font(name='Calibri', bold=True)
#     centered_alignment = Alignment(horizontal='center')
#     border_bottom = Border(
#         bottom=Side(border_style='medium', color='FF000000'),
#     )
#     wrapped_alignment = Alignment(
#         vertical='top',
#         wrap_text=True
#     )

#     # Get active worksheet/tab
#     worksheet = workbook.active
#     worksheet.title = 'Employees'
    
#     # Define the titles for columns
#     columns = []

#     if 'employee_name' in column_list:
#         columns += [ 'Employee Name' ]

#     if 'employee_role' in column_list:
#         columns += [ 'Role' ]
    
#     if 'employee_dep_org' in column_list:
#         columns += [ 'Dept./Org.' ] 

#     if 'employee_platform' in column_list:
#         columns += [ 'Platform(web/mobile)' ]    

#     if 'employee_last_sign_in' in column_list:
#         columns += [ 'Last Login' ]

#         # columns += [ 'Address' ] 

#     row_num = 1

#     # Assign the titles for each cell of the header
#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title
#         cell.font = header_font
#         cell.alignment = centered_alignment
#         cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#         cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#         column_letter = get_column_letter(col_num)
#         column_dimensions = worksheet.column_dimensions[column_letter]
#         column_dimensions.width = 20

#     # Iterate through all movies
#     for user in users:
#         if user.last_login is not None:
#             if user.auth_token is None:
#                 employee_platform = 'Web'
#             else:
#                 employee_platform = 'APP'
#         else:
#             employee_platform = ''               
         
#         row_num += 1
#         # Define the data for each cell in the row 

#         row = []
#         if 'employee_name' in column_list:
#             row += [ user.first_name + user.middle_name + user.last_name ]

#         if 'employee_role' in column_list:
#             row += [ user.role_name ]
        
#         if 'employee_dep_org' in column_list:
#             row += [ user.department_name + '/' + user.organization_name ] 

#         if 'employee_platform' in column_list:
#             row += [ employee_platform ]

#         if 'employee_last_sign_in' in column_list:
#             row += [ user.last_login ]             
       
#         # row += [ organization.address + ', ' + organization.pincode ]
        
#         # Assign the data for each cell of the row 
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.alignment = wrapped_alignment

#     workbook.save(response)

#     return response

#     #export to pdf non operational user list



#export to excel employee list
@login_required
def exportEmployeeToXlsx(request, columns, search, jobs, depts, roles):
    column_list = columns.split (",")
    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    jobss = jobs.split(",")
    deptss = depts.split(",")
    roless = roles.split(",")

    if search != '0':
        users = users.filter(id = search)
    if depts != '0':
        users = users.filter(department_id__in = deptss)
    if roles != '0':
        users = users.filter(role_id__in = roless)
    if jobs != '0':
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type__in = jobss):
                users = users.exclude(id = user.id)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employees'
    
    # Define the titles for columns
    columns = []
    columns += [ 'Employee ID' ]
    columns += [ 'Name' ]
    columns += [ 'Contact' ]
    columns += [ 'Role' ]    
    columns += [ 'Department' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:

        if user.last_login is not None:
            if user.web_auth_token is None:
                employee_platform = 'Web'
            else:
                employee_platform = 'APP'
        else:
            employee_platform = ''               
         
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        row += [ user.emp_sap_id ]
        row += [ getUserName(user.id) ]
        row += [ user.primary_contact_number ]
        row += [ user.role_name ]
        row += [ user.department_name ]
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response




@login_required
def exportEmployeeToPdf(request, columns, search, jobs, depts, roles):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')

    jobs = jobs.split(",")
    depts = depts.split(",")
    roles = roles.split(",")

    if search != '0':
        users = users.filter(Q(emp_sap_id__icontains = search) | Q(first_name__icontains = search))
    if depts != '0':
        users = users.filter(department_id__in = depts)
    if roles != '0':
        users = users.filter(role_id__in = roles)
    if jobs != '0':
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type__in = jobs):
                users = users.exclude(id = user.id)
    users = SpUsers.objects.all().filter(user_type=1).order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/employee_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Employees.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

@login_required
def getUserMap(request):
    user_coordinates = SpUsers.objects.filter(id=request.GET.get('distributor_id')).values('latitude', 'longitude').first()
    
    context = {}
    context['user_coordinates'] = user_coordinates
    context['periphery']        = request.GET.get('periphery')
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-map.html'
    return render(request, template, context) 

@login_required
def importProductVariant(request):
    # workbook object is created 
    # wb_obj = load_workbook('media/operatonal-users.xlsx')

    # sheet_obj = wb_obj.active 
    # m_row = sheet_obj.max_row 
    
    # # Loop will print all values 
    # # of first column  
    
    # for i in range(1, m_row + 1): 
    #     row = [cell.value for cell in sheet_obj[i]] 
    #     print(row)
    #     template = ProductVariantTemplate()
    #     template.store_name = row[0]
    #     template.role = row[1]
    #     template.save()
            
    return HttpResponse('row')     


def updateUserRole(user_id,params):
    role_permissions = SpRolePermissions.objects.filter(role_id=params.POST['role_id'])
    if len(role_permissions):
        SpUserRolePermissions.objects.filter(user_id=user_id).delete()
        for role_permission in role_permissions:
            user_role_permission = SpUserRolePermissions()
            user_role_permission.user_id = user_id
            user_role_permission.role_id = params.POST['role_id']
            user_role_permission.module_id = role_permission.module_id
            user_role_permission.sub_module_id = role_permission.sub_module_id
            user_role_permission.permission_id = role_permission.permission_id
            user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
            user_role_permission.workflow = role_permission.workflow
            user_role_permission.save()

        
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=params.POST['role_id'])
        if len(role_permission_workflows):
            SpUserRoleWorkflowPermissions.objects.filter(user_id=user_id).delete()
            for role_permission_workflow in role_permission_workflows : 
                user_role_permission_wf = SpUserRoleWorkflowPermissions()
                user_role_permission_wf.user_id = user_id
                user_role_permission_wf.role_id = role_permission_workflow.role_id
                user_role_permission_wf.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_wf.permission_id = role_permission_workflow.permission_id
                user_role_permission_wf.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_wf.level_id = role_permission_workflow.level_id
                user_role_permission_wf.level = role_permission_workflow.level
                user_role_permission_wf.description = role_permission_workflow.description
                user_role_permission_wf.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                user_role_permission_wf.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_wf.status = role_permission_workflow.status
                user_role_permission_wf.save()


def mapProductToUser(user_id):
    product_variants = SpProductVariants.objects.all()
    if len(product_variants):
        
        SpUserProductVariants.objects.filter(user_id=user_id).delete()

        for product_variant in product_variants:
            user_product_variant                            = SpUserProductVariants()
            user_product_variant.user_id                    = user_id
            user_product_variant.product_id                 = product_variant.product_id
            user_product_variant.product_name               = product_variant.product_name
            user_product_variant.product_variant_id         = product_variant.id
            user_product_variant.item_sku_code              = product_variant.item_sku_code
            user_product_variant.variant_quantity           = product_variant.variant_quantity
            user_product_variant.variant_unit_id            = product_variant.variant_unit_id
            user_product_variant.variant_name               = product_variant.variant_name
            user_product_variant.variant_unit_name          = product_variant.variant_unit_name
            user_product_variant.variant_size               = product_variant.variant_size
            user_product_variant.no_of_pouch                = product_variant.no_of_pouch
            user_product_variant.container_size             = product_variant.container_size
            user_product_variant.is_bulk_pack               = product_variant.is_bulk_pack
            user_product_variant.mrp                        = product_variant.mrp
            user_product_variant.sp_distributor             = product_variant.sp_distributor
            user_product_variant.sp_superstockist           = product_variant.sp_superstockist
            user_product_variant.sp_employee                = product_variant.sp_employee
            user_product_variant.container_mrp              = float(product_variant.mrp) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_distributor   = float(product_variant.sp_distributor) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_superstockist = float(product_variant.sp_superstockist) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_employee      = float(product_variant.sp_employee) * float(product_variant.no_of_pouch)
            user_product_variant.valid_from                 = product_variant.valid_from
            user_product_variant.valid_to                   = product_variant.valid_to
            user_product_variant.status                     = product_variant.status
            user_product_variant.save()


@login_required
def viewUserRolePermission(request,user_id):
    if request.method == "POST":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        context = {}
        role_id = getModelColumnById(SpUsers,user_id,'role_id')
        role = SpRoles.objects.get(id=role_id)
        permissions = SpPermissions.objects.filter(status=1)
        organizations = SpOrganizations.objects.filter(status=1)
        departments = SpDepartments.objects.filter(status=1)

        other_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for department in other_departments : 
            department.other_roles = SpRoles.objects.filter(status=1,department_id=department.id).exclude(id=role.id)

        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        
        context['permissions'] = permissions
        context['organizations'] = organizations
        context['modules'] = modules
        context['user_id'] = user_id
        context['role'] = role
        context['other_departments'] = other_departments
        context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
        context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
        context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')

        
        template = 'user-management/user-role-permission.html'
        return render(request,template,context)

@login_required
def updateUserRolePermission(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
       
        permissions = SpPermissions.objects.filter(status=1)
        sub_modules = SpSubModules.objects.filter(status=1)

        SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id']).delete()
        SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id']).delete()

        for sub_module in sub_modules :
            for permission in permissions :
                var_name = 'permission_'+ str(sub_module.id) + '_' + str(permission.id)
                if var_name in request.POST:
                    role_permission = SpUserRolePermissions()
                    role_permission.user_id = request.POST['user_id']
                    role_permission.role_id = request.POST['role_id']
                    role_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                    role_permission.sub_module_id = sub_module.id
                    role_permission.permission_id = permission.id
                    role_permission.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                    role_permission.save()
                    total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                    if request.POST[total_work_flows_var] :

                        role_permission.workflow = request.POST[total_work_flows_var]
                        role_permission.save()

                        total_work_flows = json.loads(request.POST[total_work_flows_var])
                        
                        for total_work_flow in total_work_flows :
                            role_permission_level = SpUserRoleWorkflowPermissions()
                            role_permission_level.user_id = request.POST['user_id']
                            role_permission_level.role_id = request.POST['role_id']
                            role_permission_level.sub_module_id = sub_module.id
                            role_permission_level.permission_id = permission.id
                            role_permission_level.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                            role_permission_level.level_id = total_work_flow['level_id']
                            role_permission_level.level = getModelColumnById(SpWorkflowLevels,total_work_flow['level_id'],'level')
                            role_permission_level.description = total_work_flow['description']
                            if int(total_work_flow['role_id']) > 0 :
                                role_permission_level.workflow_level_dept_id = getModelColumnById(SpRoles,total_work_flow['role_id'],'department_id')
                            else:
                                role_permission_level.workflow_level_dept_id = None
                            role_permission_level.workflow_level_role_id = total_work_flow['role_id']
                            if 'status' in total_work_flow :
                                role_permission_level.status = total_work_flow['status']
                            else:
                                role_permission_level.status = 1

                            role_permission_level.save()

        response['flag']    = True
        response['message'] = "Record has been updated successfully."
        return JsonResponse(response)
        
@login_required
def checkRolePermision(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        if SpRolePermissions.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).exists():
            response['flag'] = True
            response['message'] = "Workflow is already applied to this permission."
        else :
            response['flag'] = False
            response['message'] = "Workflow is not applied to this permission. Please contact administator."
            
        return JsonResponse(response)

@login_required
def saveRolePermisionValidity(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        role_permission = SpRolePermissions.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).first()
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=role_permission.role_id, permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id'])

        SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'],permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()
        
        user_role_permission = SpUserRolePermissions()
        user_role_permission.user_id = request.POST['user_id']
        user_role_permission.role_id = request.POST['role_id']
        user_role_permission.module_id = role_permission.module_id
        user_role_permission.sub_module_id = role_permission.sub_module_id
        user_role_permission.permission_id = role_permission.permission_id
        user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
        user_role_permission.workflow = role_permission.workflow
        user_role_permission.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.save()

        if len(role_permission_workflows) :

            SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'], permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()

            for role_permission_workflow in role_permission_workflows :
                user_role_permission_level = SpUserRoleWorkflowPermissions()
                user_role_permission_level.user_id = request.POST['user_id']
                user_role_permission_level.role_id = request.POST['role_id']
                user_role_permission_level.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_level.permission_id = role_permission_workflow.permission_id
                user_role_permission_level.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_level.level_id = role_permission_workflow.level_id
                user_role_permission_level.level = role_permission_workflow.level
                user_role_permission_level.description = role_permission_workflow.description
                if int(role_permission_workflow.level_id) > 0 :
                    user_role_permission_level.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                else:
                    user_role_permission_level.workflow_level_dept_id = None

                user_role_permission_level.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_level.status = role_permission_workflow.status
                user_role_permission_level.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.save()

        response['flag'] = True
        response['message'] = "Record has been updated successfully."
        response['workflow'] = user_role_permission.workflow
            
        return JsonResponse(response)


# add Employee Biometric
@login_required
def addEmployeeBiometricDetails(request):
    if request.method == "POST":
        last_user_id = request.POST['last_user_id']
        user = SpUsers.objects.get(id=request.POST['last_user_id'])
        user.finger_iso_1        = request.POST['finger_iso_1']
        user.finger_iso_2        = request.POST['finger_iso_2']
        # user.aten_timing         = request.POST['aten_timing']
        user.periphery           = request.POST['periphery']
        # user.fencing_timing      = request.POST['fencing_timing']
        
        if request.POST['attendence_mode'] != "":
            user.attendence_mode     = request.POST['attendence_mode']
        
        if request.POST['geofencing_type']:
            SpBasicDetails.objects.filter(user_id=request.POST['last_user_id']).update(geofencing=request.POST['geofencing_type'])

        for each_shift in request.POST.getlist('working_shift_id[]'):
            working_shift = TblClAllocatedShifts()
            working_shift.user_id = request.POST['last_user_id']
            working_shift.working_shift_id = each_shift
            working_shift.save()

        # user.course_type_id      = request.POST['course_id']
        # user.course_type_name    = getModelColumnById(TblCourseTypes, request.POST['course_id'], 'course_type')
        # user.branch_id           = request.POST['branch_id']
        # user.branch_name         = getModelColumnById(TblBranch, request.POST['branch_id'], 'branch')
        # if getModelColumnById(TblBranch, request.POST['branch_id'], 'total_year'):
        #     user.year_id             = "year_"+request.POST['year_sem_id']
        # else:    
        #     user.semester_id         = "sem_"+request.POST['year_sem_id']
        # user.section_id          = request.POST['section_id']
        # user.reg_no              = request.POST['roll_no']
        # if request.POST['appeared_option'] ==  '1':
        #     user.sbat_id             = request.POST['sbat_id']
        #     user.sbat_percentage     = request.POST['percent']
        # else:
        #     user.sbat_id             = None
        #     user.sbat_percentage     = None
        user.save()
        # user_data1 = TblStudents.objects.get(id=last_user_id)
        # user_data1.is_registered = 1
        # user_data1.save()

        context={}
        context['college']           = 'college'
        context['document_types']    = []
        context['document_lists']    = []
        context['last_user_id']      = last_user_id 
        context['college_id']        = user.organization_id   
        rooms = TblClRoom.objects.filter(college_id=user.organization_id).values('id', 'room').order_by('id')
        context['rooms']                = rooms 
        context['document_types']       = TblClDocumentTypes.objects.filter().order_by('document_name')
        context['document_lists']       = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).order_by('-id') 
        context['document_list_count']  = document_list_count  = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).count() 
        return render(request,"user-management/add-employee-documents.html",context)
    else:
        context={}
        context['working_shifts'] = TblClWorkingShifts.objects.all()
        context['college'] = 'college'
        return render(request,"user-management/add-employee-biometric-details.html",context)


# # edit Employee Biometric
# @login_required
# def editEmployeeBiometricDetails(request, employee_id):
#     if request.method == "POST":
#         user = SpUsers.objects.get(id = request.POST['last_user_id'])
#         user.finger_iso_1        = request.POST['finger_iso_1']
#         user.finger_iso_2        = request.POST['finger_iso_2']
#       # user.aten_timing         = request.POST['aten_timing']
#         user.periphery           = request.POST['periphery']
#       # user.fencing_timing      = request.POST['fencing_timing']
#         user.attendence_mode     = request.POST['attendence_mode']
#         user.save()

#         # try:
#         #     employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
#         # except SpUserDocuments.DoesNotExist:
#         #     employee_documents = None

#         last_user_id = request.POST['last_user_id']

#         context={}
#         context['college']            = 'college'
#         context['document_types']     = []
#         context['document_lists']     = []
#         context['last_user_id']       = last_user_id
#         context['college_id']          = getModelColumnById(SpUsers, last_user_id, 'organization_id')
#         context['employee_documents'] = []
#         return JsonResponse(context)
#     else:
#         user        = SpUsers.objects.get(id=employee_id)
#         context                      = {}
#         context['college']           = 'college'
#         context['last_user_id']      = employee_id
#         context['user']              = user
#         template                     = 'user-management/edit-employee/employee-biometric-details.html'
#         return render(request, template, context) 

# edit Employee Biometric
@login_required
def editEmployeeBiometricDetails(request, employee_id):
    if request.method == "POST":
        user = SpUsers.objects.get(id = request.POST['last_user_id'])
        user.finger_iso_1        = request.POST['finger_iso_1']
        user.finger_iso_2        = request.POST['finger_iso_2']
       # user.aten_timing         = request.POST['aten_timing']
       # user.fencing_timing      = request.POST['fencing_timing']
        user.periphery           = request.POST['periphery']
        if request.POST['attendence_mode'] != "":
            user.attendence_mode     = request.POST['attendence_mode']
        else:
            user.attendence_mode     = None
        user.save()

        if request.POST['geofencing_type'] != "":
            employee_basic_details = SpBasicDetails.objects.filter(user_id=employee_id).update(geofencing=request.POST['geofencing_type'])

        # try:
        #     employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
        # except SpUserDocuments.DoesNotExist:
        #     employee_documents = None

        last_user_id = request.POST['last_user_id']

        TblClAllocatedShifts.objects.filter(user_id = request.POST['last_user_id']).delete()
        for each_shift in request.POST.getlist('working_shift_id[]'):
            working_shift = TblClAllocatedShifts()
            working_shift.user_id = request.POST['last_user_id']
            working_shift.working_shift_id = each_shift
            working_shift.save()

        context={}
        context['college']            = 'college'
        context['document_types']     = []
        context['document_lists']     = []
        context['last_user_id']       = last_user_id
        context['college_id']         = getModelColumnById(SpUsers, last_user_id, 'organization_id')
        context['employee_documents'] = []
        return JsonResponse(context)

    else:
        working_shifts         = TblClWorkingShifts.objects.all()
        allocat_shift          = TblClAllocatedShifts.objects.filter(user_id=employee_id).values_list("working_shift_id", flat=True)
        user                   = SpUsers.objects.get(id=employee_id)
        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)

        context                      = {}
        context['working_shifts']    = working_shifts
        context['allocat_shifts']    = allocat_shift
        context['college']           = 'college'
        context['last_user_id']      = employee_id
        context['user']              = user
        context['employee_basic_details'] = employee_basic_details
        template    = 'user-management/edit-employee/employee-biometric-details.html'
        return render(request, template, context) 


# add student documents
@login_required
def addEmployeeDocuments(request):
    response = {}
    if request.method == "POST":
        try:
            last_user_id = request.POST['last_user_id']
            if TblClEmployeeDocuments.objects.filter(ducument_number=request.POST['document_number']).exists():
                response['error']   = True
                response['message'] = "Document No. already exists."
                return JsonResponse(response)
            elif TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_name=getModelColumnById(TblClDocumentTypes, request.POST['document_type'], 'document_name')).exists():
                response['error']   = True
                response['message'] = "Document already addedd."
                return JsonResponse(response)
            else:
                document                  = TblClEmployeeDocuments()
                document.employee_id       = last_user_id
                document.document_name    = getModelColumnById(TblClDocumentTypes, request.POST['document_type'], 'document_name')
                document.ducument_number  = request.POST['document_number']
                document.is_uploaded      = 0
                document.created_by       = request.user.id
                document.save()

                #-----------------------------notify android block-------------------------------#
                organization_id = getModelColumnById(SpUsers, request.POST['last_user_id'], 'organization_id')
                all_users = SpUsers.objects.filter(organization_id = organization_id, user_type=1).exclude(id=1).values_list("id", flat=True)
                for each_user in all_users:
                    user_id = each_user
                    userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
                    employee_name = getUserName(request.user.id)
    
                    message_title = document.document_name+" document scan request."
                    message_body =  document.document_name+" document scan request has been generated by "+employee_name
                    notification_image = ""
    
                    if userFirebaseToken is not None and userFirebaseToken != "" :
                        registration_ids = []
                        registration_ids.append(userFirebaseToken)
                        data_message = {}
                        data_message['id'] = 1
                        data_message['status'] = 'notification'
                        data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image'] = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

                context={}
                context['college']           = 'college'
                context['college_id']          = getModelColumnById(SpUsers, last_user_id, 'organization_id')
                context['document_lists']    = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).order_by('-id')
                context['document_list_count']  = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).count()
                context['last_user_id']      = last_user_id    
                return render(request,"user-management/employee-documents.html",context)
        except Exception as e:
            response['error']            = True
            response['message']          = str(e)
            return HttpResponse(e)
    else:
        rooms = TblClRoom.objects.filter(college_id=request.GET['college_id']).values('id', 'room').order_by('id')
        context = {}
        context['last_user_id']         = request.GET['employee_id']
        context['college_id']           = getModelColumnById(SpUsers, request.GET['employee_id'], 'organization_id')
        context['rooms']                = rooms 
        context['document_types']       = TblClDocumentTypes.objects.filter().order_by('document_name')
        context['document_lists']       = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).order_by('-id') 
        context['document_list_count']  = document_list_count  = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).count()       
        return render(request,"user-management/add-employee-documents.html",context) 
        
# add student document list
@login_required
def employeeDocumentList(request):
    type                = request.POST['type']
    last_user_id        = request.POST['last_user_id']
    previous_list_count = request.POST['document_list_count']
    context={}
    context['college']              = 'college'
    context['last_user_id']         = last_user_id
    context['college_id']           = getModelColumnById(SpUsers, last_user_id, 'organization_id')
    context['document_lists']       = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).order_by('-id') 
    context['document_list_count']  = document_list_count  = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).count()   
    
    if type == '0' and int(previous_list_count) == int(document_list_count):
        response                     = {}
        response['error']            = True
        return JsonResponse(response)
    else:
        context['is_difference']  = 0   
        return render(request,"user-management/employee-documents.html",context) 

@login_required
def addEmployeeNewFolder(request):
    if request.method == "POST":
        response = {}
        response['employee_id']   = request.POST['employee_id']
        response['college_id']   = getModelColumnById(SpUsers, request.POST['employee_id'], 'organization_id')
        response['room_id']      = request.POST['room_id']
        response['almira_id']    = request.POST['almira_id']
        response['rack_id']      = request.POST['rack_id']
        response['rack_name']    = getModelColumnById(TblClRack, request.POST['rack_id'], 'rack')
        
        employee_name = getModelColumnById(SpUsers, request.POST['employee_id'], 'first_name')
        employee_name = employee_name.lower()
        
        if request.POST['is_auto'] == '1':
            if request.POST['auto_file_name'] == '1':
                file_name = employee_name
            elif request.POST['auto_file_name'] == '2': 
                file_name = request.POST['employee_id']
            else:
                file_name = str(employee_name)+'_'+str(request.POST['employee_id'])
        else:
            file_name = request.POST['file_name']   

        if TblClEmployeeFileFolder.objects.filter(file_name=file_name, employee_id=request.POST['employee_id'], college_id=response['college_id'], room_id=request.POST['room_id'], almira_id=request.POST['almira_id'], rack_id=request.POST['rack_id']).exists():
            response['error']   = True
            response['message'] = "File Name already exists."
        else:    
            file_folder              = TblClEmployeeFileFolder()
            file_folder.employee_id   = request.POST['employee_id']
            file_folder.college_id   = response['college_id']
            file_folder.room_id      = request.POST['room_id']
            file_folder.almira_id    = request.POST['almira_id']
            file_folder.rack_id      = request.POST['rack_id']
            if request.POST['is_auto'] == '1':   
                file_folder.file_name    = file_name
            else:
                file_folder.file_name    = file_name    
            file_folder.created_by   = request.user.id
            file_folder.save()

            college_name = getModelColumnById(SpOrganizations, file_folder.college_id, 'organization_name')
            room_name    = getModelColumnById(TblClRoom, file_folder.room_id, 'room')
            almira_name  = getModelColumnById(TblClAlmirah, file_folder.almira_id, 'almirah')
            rack_name    = getModelColumnById(TblClRack, file_folder.rack_id, 'rack')
            folder='media/documents/'+str(college_name)+'/'+str(room_name)+'/'+str(almira_name)+'/'+str(rack_name)+'/'+str(file_folder.file_name) 
            try:
                os.mkdir(folder)
            except OSError:
                pass
            else:
                pass
            
        return JsonResponse(response)
    else:    
        context = {}
        context['employee_id']   = request.GET['employee_id']
        context['room_id']      = request.GET['room_id']
        context['almira_id']    = request.GET['almira_id']
        context['rack_id']      = request.GET['rack_id']
        template = 'user-management/add-employee-new-folder.html'
        return render(request,template,context)

@login_required
def addEmployeeNewFile(request):
    if request.method == "POST":
        file_path   = getModelColumnById(TblClEmployeeDocuments, request.POST['document_id'], 'document_path')
        baseurl     = settings.BASE_URL
        file_path   = file_path
        
        destination_folder = str(getModelColumnById(TblClRack, request.POST['rack_id'], 'path'))+'/'+str(request.POST['master_name'])
        print(destination_folder)
        shutil.move(file_path, destination_folder)

        document_file           = TblClEmployeeDocuments.objects.get(id=request.POST['document_id'])
        document_file.document_path = None
        document_file.save()
        
        file                    = TblClEmployeeFolderFiles()
        file.employee_id         = request.POST['employee_id']
        file.college_id         = request.POST['college_id']
        file.room_id            = request.POST['room_id']
        file.room_name          = request.POST['room_name']
        file.almira_id          = request.POST['almira_id']
        file.almira_name        = request.POST['almira_name']
        file.rack_id            = request.POST['rack_id']
        file.rack_name          = request.POST['rack_name']
        file.file_id            = request.POST['id']
        file.file_name          = request.POST['master_name']
        file.docket_no          = request.POST['docket_no']
        file.document_id        = request.POST['document_id']
        file.document_name      = request.POST['document_name']
        file.document_no        = request.POST['document_no']
        file.document_group     = request.POST['document_group']
        file.is_expiry    = request.POST['is_expiry']
        if request.POST['is_expiry'] == '1':   
            file.expiry_date    = datetime.strptime(request.POST['expiry_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        else:
            file.expiry_date    = None    
        file_path = file_path.split('/')
        file_path = file_path[-1]    
        file.document_path      = str(destination_folder)+'/'+ str(file_path) 
        file.tags               = request.POST['tags'] 
        file.created_by         = request.user.id
        file.save()

        response = {}
        response['id']           = request.POST['id']
        response['type']         = request.POST['type']
        response['college_id']   = request.POST['college_id']
        response['master_name']  = request.POST['master_name']
        response['employee_id']   = request.POST['employee_id']
        response['room_id']      = request.POST['room_id']
        response['room_name']    = request.POST['room_name']
        response['almira_id']    = request.POST['almira_id']
        response['almira_name']  = request.POST['almira_name']
        response['rack_id']      = request.POST['rack_id']
        response['rack_name']    = request.POST['rack_name']
        return JsonResponse(response)
    else:    
        context = {}
        context['document_id']  = request.GET['document_id']
        context['id']           = request.GET['id']
        context['type']         = request.GET['type']
        context['college_id']   = request.GET['college_id']
        context['master_name']  = request.GET['master_name']
        context['employee_id']   = request.GET['employee_id']
        context['room_id']      = request.GET['room_id']
        context['room_name']    = request.GET['room_name']
        context['almira_id']    = request.GET['almira_id']
        context['almira_name']  = request.GET['almira_name']
        context['rack_id']      = request.GET['rack_id']
        context['rack_name']    = request.GET['rack_name']
        context['document_list_count']  = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).count()
        context['document_details']     = TblClEmployeeDocuments.objects.get(id=request.GET['document_id'], document_path__isnull=False)
        context['document_group_list']  = TblClDocumentGroup.objects.filter(status=1)
        template = 'user-management/create-employee-new-file.html'
        return render(request,template,context)

@login_required
def getEmployeeMasterDetails(request):
    context     = {}
    id          = request.POST['id']
    type        = request.POST['type']
    college_id  = request.POST['college_id']
    last_user_id  = request.POST['last_user_id']
    master_name = request.POST['master_name']
    if type == 'room_list':
        context['master']       = 'room'
        context['rooms']        = TblClRoom.objects.filter(college_id=college_id)
        context['college_id']   = college_id
        context['master_name']  = master_name
        context['last_user_id']  = last_user_id
    elif type == 'room':
        context['master']       = 'almira'
        context['almiras']      = TblClAlmirah.objects.filter(room_id=id)
        context['college_id']   = college_id
        context['master_name']  = master_name
        context['last_user_id']  = last_user_id
    elif type == 'almira':
        context['master']       = 'rack'
        context['racks']      = TblClRack.objects.filter(almira_id=id)
        context['college_id']   = college_id
        context['room_id']      = getModelColumnById(TblClAlmirah, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClAlmirah, id, 'room_name')
        context['almira_id']    = id
        context['master_name']  = getModelColumnById(TblClRoom, context['room_id'], 'room')  
        context['almira_name']  = getModelColumnById(TblClAlmirah, id, 'almirah')
        context['last_user_id']  = last_user_id
    elif type == 'rack':
        context['master']       = 'file'
        
        context['college_id']   = college_id
        context['last_user_id'] = last_user_id
        context['rack_id']      = id
        context['rack_name']    = getModelColumnById(TblClRack, id, 'rack')
        context['room_id']      = getModelColumnById(TblClRack, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClRack, id, 'room_name')
        context['almira_id']    = getModelColumnById(TblClRack, id, 'almira_id') 
        context['almira_name']  = getModelColumnById(TblClRack, id, 'almira_name') 
        context['files']        = TblClEmployeeFileFolder.objects.filter(college_id=college_id, employee_id=last_user_id, room_id=context['room_id'], almira_id=context['almira_id'], rack_id=context['rack_id'])   

        context['master_name']  = '' 
    elif type == 'file':
        context['master']       = 'folder_files'
        
        context['college_id']   = college_id
        context['last_user_id'] = last_user_id
        context['file_id']      = id
        context['file_name']    = getModelColumnById(TblClEmployeeFileFolder, id, 'file_name')
        context['room_id']      = getModelColumnById(TblClEmployeeFileFolder, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClRoom, context['room_id'], 'room')
        context['almira_id']    = getModelColumnById(TblClEmployeeFileFolder, id, 'almira_id') 
        context['almira_name']  = getModelColumnById(TblClAlmirah, context['almira_id'], 'almirah') 
        context['rack_id']      = getModelColumnById(TblClEmployeeFileFolder, id, 'rack_id')
        context['rack_name']    = getModelColumnById(TblClRack, context['rack_id'], 'rack')
        context['folder_files'] = TblClEmployeeFolderFiles.objects.filter(file_id=id, college_id=college_id, employee_id=last_user_id, room_id=context['room_id'], almira_id=context['almira_id'], rack_id=context['rack_id'])    

        context['master_name']  = ''     
    baseurl = settings.BASE_URL
    
    context['baseurl']           = baseurl
    template                     = 'user-management/employee-master-details.html'
    return render(request, template, context)


@login_required
def resetCredential(request,user_id):
    if request.method == "POST":
        response = {}
        try:
            user = SpUsers.objects.get(id=request.POST['user_id'])
            password = make_password(request.POST['new_password'])
            user.password = password
            
            user.save()
            AuthtokenToken.objects.filter(user_id = request.POST['user_id'])
            
            

            if user.id :

                #Save Activity
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'User credentials updated'
                activity    = 'User credentials updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('User Management', 'User Management', heading, activity, request.user.id, user_name, 'updateVehiclePass.png', '1', 'web.png')
                
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,user.id,'firebase_token')
                employee_name = getUserName(user.id)

                message_title = "Password reset"
                message_body = "You password has been changed by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(user.id,'SpUsers','User Management','Password reset',message_title,message_body,notification_image,request.user.id,user_name,user.id,employee_name,'password.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['flag'] = True
                response['user_id'] = request.POST['user_id']
                response['message'] = "Record has been updated successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)
    else:
        context = {}
        context['user']     = SpUsers.objects.get(id=user_id)
        template = 'user-management/reset-user-credential.html'
        return render(request, template, context)


@login_required
def resetUserLocation(request):
    if request.method == "POST":
        response = {}
        try:
            user = SpUsers.objects.get(id=request.POST['user_id'])
            # password = make_password(request.POST['new_password'])
            # user.latitude = None
            # user.longitude = None
            user.device_id = None
            user.save()

            if user.id :

                #Save Activity
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'User location reset'
                activity    = 'User location reset by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('User Management', 'User Management', heading, activity, request.user.id, user_name, 'updateVehiclePass.png', '1', 'web.png')
                
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,user.id,'firebase_token')
                employee_name = getUserName(user.id)

                message_title = "Location reset"
                message_body = "Your location has been reset by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(user.id,'SpUsers','User Management','Location reset',message_title,message_body,notification_image,request.user.id,user_name,user.id,employee_name,'password.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['flag'] = True
                response['user_id'] = request.POST['user_id']
                response['message'] = "Location has been reset successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)




def mapUserLeaves(role_id,user_id):
    try:
        leave_policy_id = SpRoleEntityMapping.objects.get(role_id = role_id , entity_type = "leave_policy")
        leave_policy_id = leave_policy_id.entity_id
    except SpRoleEntityMapping.DoesNotExist:
        leave_policy_id = None
    if leave_policy_id:
        leave_policy_dettail = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id)
        balance  = 0
        for policy_detail in leave_policy_dettail:
            leave_polcy_ledger = SpUserLeavePolicyLedger()
            leave_polcy_ledger.user_id = user_id
            leave_polcy_ledger.leave_policy_id = leave_policy_id
            leave_polcy_ledger.leave_type_id = policy_detail.leave_type_id
            current_month = datetime.today().strftime('%m')
            if int(current_month) == 1:
                month_leave_count = policy_detail.year_leave_count / 12
                leave_polcy_ledger.year_leave_count = policy_detail.year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
                
                leave_polcy_ledger.credit =  policy_detail.year_leave_count
                balance += policy_detail.year_leave_count
                leave_polcy_ledger.balance = balance
            else:   
                current_month = 12 - int(current_month)
                sub_leave_count = policy_detail.year_leave_count / 12
                year_leave_count = sub_leave_count*current_month
                month_leave_count = year_leave_count  / current_month
                leave_polcy_ledger.year_leave_count = year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
                
                leave_polcy_ledger.credit =  year_leave_count
                balance +=year_leave_count
                leave_polcy_ledger.balance = balance
            
            leave_polcy_ledger.consecutive_leave = policy_detail.consecutive_leave 
            leave_polcy_ledger.save
            
            
            

@login_required
def userTrackingReport(request):
    context = {}
    context['users'] = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['page_title'] = "Users Tracking Report"
    context['org_latitude'] = getConfigurationResult('org_latitude')
    context['org_longitude'] = getConfigurationResult('org_longitude')
    template = 'user-management/user-tracking-report.html'
    return render(request, template, context)

@login_required
def ajaxUserTracking(request,user_id):
    if 'track_date' in request.GET and request.GET['track_date'] != "" :
            today                   = request.GET['track_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
        
    tracks      = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today)
    trackss     = [trackss.id for trackss in tracks]
    len_track   = len(trackss)
    
    if tracks:
        distance_travelled  = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
        distance_travelled  = round(distance_travelled['distance_travelled__sum']*0.001,2)
    else:
        distance_travelled = 0
    context = {}
    counter_of_len = 0
    user_tracking_list  = []
    if len_track < 23 :
        for id, user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
            R = 6373.0
            lat1 = radians(float(track.latitude))
            lon1 = radians(float(track.longitude))
            lat2 = radians(float(user_last_data.latitude))
            lon2 = radians(float(user_last_data.longitude))
            dlon = lon2 - lon1
            dlat = lat2 - lat1
            a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            
            if meter_distance > 0.00:
                user_tracking_list.append(trackss[id])
    elif len_track > 100 and distance_travelled > 200:
        for id,user_last in enumerate(trackss):
            counter_of_len  +=  1
            track           =   SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R       = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))   
            distance        = R * c
            meter_distance  = float(distance * 1000)
            if meter_distance > 3000:
                user_tracking_list.append(trackss[id])
    elif len_track > 70 and distance_travelled > 150:
        for id,user_last in enumerate(trackss):
            counter_of_len  +=  1
            track           =   SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R       = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))   
            distance        = R * c
            meter_distance  = float(distance * 1000)
            if meter_distance > 600:
                user_tracking_list.append(trackss[id])
    elif len_track > 40 and distance_travelled > 70:
        for id,user_last in enumerate(trackss):
            counter_of_len  +=  1
            track           =   SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R       = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))   
            distance        = R * c
            meter_distance  = float(distance * 1000)
            if meter_distance > 400:
                user_tracking_list.append(trackss[id])
    elif len_track < 50 and distance_travelled < 20:
        for id,user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
            R = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 150:
                user_tracking_list.append(trackss[id])
    elif len_track < 50 and distance_travelled < 20:
        for id,user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 150:
                user_tracking_list.append(trackss[id])
                
    elif len_track < 100 and distance_travelled < 20:
        for id, user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R = 6373.0
            lat1 = radians(float(track.latitude))
            lon1 = radians(float(track.longitude))
            lat2 = radians(float(user_last_data.latitude))
            lon2 = radians(float(user_last_data.longitude))
            dlon = lon2 - lon1
            dlat = lat2 - lat1
            a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 130:
                user_tracking_list.append(trackss[id])
    elif len_track > 70 and distance_travelled < 50:
        for id, user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R = 6373.0
            lat1 = radians(float(track.latitude))
            lon1 = radians(float(track.longitude))
            lat2 = radians(float(user_last_data.latitude))
            lon2 = radians(float(user_last_data.longitude))
            dlon = lon2 - lon1
            dlat = lat2 - lat1
            a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 500:
                user_tracking_list.append(trackss[id])
    elif len_track > 70 and distance_travelled < 50:
        for id, user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R = 6373.0
            lat1 = radians(float(track.latitude))
            lon1 = radians(float(track.longitude))
            lat2 = radians(float(user_last_data.latitude))
            lon2 = radians(float(user_last_data.longitude))
            dlon = lon2 - lon1
            dlat = lat2 - lat1
            a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 500:
                user_tracking_list.append(trackss[id])
    else:
        for id,user_last in enumerate(trackss):
            counter_of_len += 1
            track = SpUserTracking.objects.get(id=trackss[id])
            if counter_of_len < len_track:
                user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
            else:
                user_last_data = SpUserTracking.objects.get(id=trackss[id])
                
            R = 6373.0
            lat1    = radians(float(track.latitude))
            lon1    = radians(float(track.longitude))
            lat2    = radians(float(user_last_data.latitude))
            lon2    = radians(float(user_last_data.longitude))
            dlon    = lon2 - lon1
            dlat    = lat2 - lat1
            a       = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
            c       = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c
            meter_distance = float(distance * 1000)
            if meter_distance > 350:
                user_tracking_list.append(trackss[id])             
    if len(tracks):
        context['tracks']       = SpUserTracking.objects.filter(id__in = user_tracking_list)
        context['first_track']  = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).first
        context['last_track']   = SpUserTracking.objects.filter(user_id=user_id,created_at__contains=today).last
        
    start_end_day = SpUserAttendance.objects.filter(user_id = user_id,attendance_date_time__icontains = today)
    start_time  = ''
    end_time    = ''
    for days in start_end_day:
        if days.start_time :
            start_time = str(days.start_time)
        if days.end_time:
            end_time = str(days.end_time)      
    context['distance_travelleds']  = distance_travelled
    context['start_time']           = start_time
    context['end_time']             = end_time
    context['user']                 = SpUsers.objects.get(id=user_id)
    template                        = 'user-management/ajax-user-tracking.html'
    return render(request, template, context)    

@login_required
def userTravelSummary(request):
    today       = date.today()
    users   = SpUserTracking.objects.filter().values('user_id').distinct().values('user_id')
    for user in users:
        user['name'] = getUserName(user['user_id'])

    user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).values('user_id').distinct().values('user_id', 'travel_charges')
    for user_tracking in user_tracking_details:
        distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today.strftime("%Y-%m-%d")).aggregate(Sum('distance_travelled'))
        if distance_travelled['distance_travelled__sum']:
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
        else:
            user_tracking['distance_travelled'] = 0
        if user_tracking['distance_travelled'] > 0:
            user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
            user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
        else:
            user_tracking['charges']            = 0
            user_tracking['total_charges']      = 0
        user_tracking['user_name']          = getUserName(user_tracking['user_id'])


    context = {}
    context['today_date']               = today.strftime("%d/%m/%Y")
    context['users']                    = users
    context['user_tracking_details']    = user_tracking_details
    context['month_date']               = date.today().strftime("%m/%Y")
    context['page_title']               = "User Travel Summary"
    template = 'user-management/user-travel-summary.html'
    return render(request, template, context)

@login_required
def ajaxuserTravelSummary(request):
    today                   = request.GET['travel_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    context = {}
    
      
    if request.GET['time_period'] == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        
        if request.GET['user_id']:
            user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            if distance_travelled['distance_travelled__sum']:
                user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                user_tracking['distance_travelled'] = 0
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])
        context['user_tracking_details']    = user_tracking_details       
    else:
        user_tracking_detail = []
        if request.GET['user_id']:
            travel_month = request.GET['travel_month_picker']
            travel_month = travel_month.split('/')
            year  = int(travel_month[1])
            month = int(travel_month[0])
            
            last_day_of_month = calendar.monthrange(year,month)[1]
            if int(month) < 10:
                month = '0'+str(month)
                 
            for x in range(last_day_of_month):
                x = x+1
                if x < 10:
                    x = '0'+str(x)
                travel_date  = str(year)+'-'+str(month)+'-'+str(x)
                travel_dates = str(x)+'/'+str(month)+'/'+str(year)
    
                user_trackings = {}
                
                try:
                    distance_travelled  = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
                except SpUserTracking.DoesNotExist:
                    distance_travelled = None
    
                    
                user_trackings['travel_date']          = travel_dates
                if distance_travelled['distance_travelled__sum']:
                    user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    user_trackings['distance_travelled'] = 0.0
                if user_trackings['distance_travelled'] > 0:
                    travel_charges       = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).values('travel_charges').first()        
                    user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
                    user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
                else:
                    user_trackings['charges']            = 0
                    user_trackings['total_charges']      = 0
                user_trackings['user_name']          = getUserName(request.GET['user_id'])
                user_trackings['user_id']          = request.GET['user_id']
                user_tracking_detail.append(user_trackings)  
        context['user_tracking_details']    = user_tracking_detail
        
    if request.GET['travel_month_picker']:
        context['month_date']               = request.GET['travel_month_picker']
    else:
        context['month_date']               = date.today().strftime("%m/%Y")
    context['time_period']                  = request.GET['time_period']  
    # if request.GET['time_period'] == '2':  
    #     context['user_id']                  = request.GET['user_id']    
    context['travel_date']                  = request.GET['travel_date']    
    template = 'user-management/ajax-user-travel-summary-report.html'
    return render(request, template, context)

#get export user summary
@login_required
def exportUserTravelSummary(request, travel_date, user_id, travel_month_picker, time_period):
    today                   = travel_date
          
    if time_period == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        if user_id!='0':
            user_tracking_details = user_tracking_details.filter(user_id=user_id)
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0   
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_daily_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'User Daily Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
        img.height = 50
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = 'User Daily Travel Summary as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Employee Name' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_details:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['user_name'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By Balinee Milk'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response

    else:
        travel_month = travel_month_picker
        travel_month = travel_month.split('-')
        year  = int(travel_month[1])
        month = int(travel_month[0])
        
        last_day_of_month = calendar.monthrange(year,month)[1]
        if int(month) < 10:
            month = '0'+str(month)
        user_tracking_detail = []     
        for x in range(last_day_of_month):
            x = x+1
            if x < 10:
                x = '0'+str(x)
            travel_date  = str(year)+'-'+str(month)+'-'+str(x)
            travel_dates = str(x)+'/'+str(month)+'/'+str(year)

            user_trackings = {}
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
            user_trackings['travel_date']          = travel_dates
            if distance_travelled['distance_travelled__sum']:
                user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                user_trackings['distance_travelled'] = 0    
            if user_trackings['distance_travelled'] > 0:
                travel_charges       = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).values('travel_charges').first()        
                user_trackings['charges']            = float(travel_charges['travel_charges'])
                user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
            else:
                user_trackings['charges']            = 0
                user_trackings['total_charges']      = 0
            user_trackings['user_name']          = getUserName(user_id)
            user_tracking_detail.append(user_trackings)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_monthly_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'User Monthly Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
        img.height = 50
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        month_name = datetime(int(travel_month[1]),int(travel_month[0]),1).strftime( '%B' )
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = ''+getUserName(user_id)+' Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Date' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_detail:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['travel_date'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By Balinee Milk'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response     

@login_required
def userTrackingReportView(request, user_id, trac_date):
    trac_date = trac_date.replace("-", "/")
    context = {}
    default_user = SpUsers.objects.get(id=user_id)
    user_fname = default_user.first_name
    user_mname = default_user.middle_name
    user_lname = default_user.last_name
    if user_mname:
        default_user_name = user_fname + ' ' + user_mname + ' ' + user_lname
    else:
        default_user_name = user_fname + ' ' + user_lname
    users = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['default_user_name'] = default_user_name
    context['user_id'] = int(user_id)
    context['trac_date'] = trac_date  
    context['users'] = users
    context['page_title'] = "User Tracking "
    context['org_latitude'] = getConfigurationResult('org_latitude')
    context['org_longitude'] = getConfigurationResult('org_longitude')
    context['organization'] = SpOrganizations.objects.all()
    template = 'user-management/user-tracking-report-view.html'
    return render(request, template, context)


@login_required
def addLeaveLedger(request,employee_id):
    employee = SpUsers.objects.filter(id = employee_id)
    leaves = SpLeaveTypes.objects.filter(status = 1)
    context = {}
    context['employee'] = employee
    context['employee_id'] = employee_id
    context['leaves'] = leaves
    templates = "user-management/add-leave-ledger.html"
    return render(request,templates,context)


@login_required
def saveLeaveLedger(request):
    if request.method == "POST":
        emp_id  = request.POST['emp_id']
        leave_type = request.POST['leave_type']
        no_leaves =  request.POST['no_leaves']
        types =  request.POST['type']
        
        leavess = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id).last()
        
        leave_types = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave_type, user_id = emp_id).last()
        if leave_types:
            if float(no_leaves) > float(leave_types.year_leave_count) and int(request.POST['type']) == 0:
                message  = "Debit is less than no. of leaves"
                response = {}
                response['flag'] = False
                response['message'] = message
                response['emp_id'] = emp_id
                return JsonResponse(response)
        else:
            role_id = getModelColumnById(SpUsers,emp_id,'role_id')
            if int(request.POST['type']) == 0:
                message  = "No balance leave"
                response = {}
                response['flag'] = False
                response['message'] = message
                response['emp_id'] = emp_id
                return JsonResponse(response)
            else:
                if SpRoleEntityMapping.objects.filter(role_id=role_id, entity_type = 'leave_policy').exists():
                    led = SpRoleEntityMapping.objects.filter(role_id = role_id, entity_type = 'leave_policy').last()
                    leave_policy_id =led.entity_id
                     
                    leave_ledger = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id, leave_type_id= leave_type).last()
                    # leavess = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id).last()
                    # leave_types = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave_type, user_id = emp_id).last()
                    if leavess:
                        balances = leavess.balance
                    else:
                        balances = 0
                    leave_id = leave_ledger.leave_policy_id
                    leave_typess = leave_ledger.leave_type_id
                    month_leave = leave_ledger.month_leave_count
                    cons_leave = leave_ledger.consecutive_leave
                    year_leave = 0
                    
                    
                    l_ledger = SpUserLeavePolicyLedger()
                    l_ledger.user_id = emp_id
                    l_ledger.leave_policy_id = leave_id
                    l_ledger.leave_type_id = leave_typess
                    l_ledger.month_leave_count = month_leave
                    l_ledger.consecutive_leave = cons_leave
                    l_ledger.credit = no_leaves
                    l_ledger.year_leave_count = (year_leave) + (Decimal(no_leaves))
                    l_ledger.balance = balances + float(no_leaves)
                   
                    l_ledger.save()
                    message  = "Leave Ledger Added Successfully"
                    response = {}
                    response['flag'] = True
                    response['message'] = message
                    response['emp_id'] = emp_id
                    return JsonResponse(response)
                else:
                    message  = "There is no leave policy for this Role."
                    response = {}
                    response['flag'] = False
                    response['message'] = message
                    response['emp_id'] = emp_id
                    return JsonResponse(response)   
        leave_policy_id = leavess.leave_policy_id
        leave_debits = leavess.debit
        leave_credit = leavess.credit
        leave_type_id = leave_types.leave_type_id
        year_leave_counts = leave_types.year_leave_count
        month_leave_count = leavess.month_leave_count
        consecutive_leave = leavess.consecutive_leave
        balances = leavess.balance
        types =  request.POST['type']
        if types == "1":
            credit = no_leaves
        else:
            debit = no_leaves
        
        ledger = SpUserLeavePolicyLedger()
        ledger.user_id = emp_id
        ledger.leave_policy_id = leave_policy_id
        ledger.leave_type_id = leave_type_id
        
        ledger.month_leave_count = month_leave_count
        ledger.consecutive_leave = consecutive_leave
        if types == '1':
            ledger.credit = no_leaves
            ledger.year_leave_count = (year_leave_counts) + (Decimal(no_leaves))
            ledger.balance = balances + float(credit)
        else:
            ledger.debit = no_leaves
            ledger.year_leave_count = (year_leave_counts) - (Decimal(no_leaves))
            ledger.balance = balances - float(debit)
        ledger.save()
        message  = "Leave Ledger Added Successfully"
        response = {}
        response['flag'] = True
        response['message'] = message
        response['emp_id'] = emp_id
        return JsonResponse(response)
 
 
 
@login_required
def ajaxLeaveLedger(request,emp_id):
    if request.method == "POST":
        emp_id  = request.POST['emp_id']
        employee = SpUsers.objects.filter(id = emp_id)
        leave_ledger = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id)
        for leave in leave_ledger:
            leave.leave_policy_name = getModelColumnById(SpLeavePolicies,leave.leave_policy_id,'leave_policy')
            leave.laave_type_name = getModelColumnById(SpLeaveTypes,leave.leave_type_id,'leave_type')
            leave.month_leave_counts = round(leave.month_leave_count,1)
            
    template = 'user-management/ajax-leave-ledger.html'
    context = {}
    context['leave_ledger'] = leave_ledger
    context['employee'] = employee
    context['emp_id'] = emp_id
    return render(request, template,context)
   