import json
import ast
from re import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound, response
from django.core.files.storage import FileSystemStorage
import time

from apps.src.decorators.teachers_api import validate_teachers_api
from ..models import *
from utils import *
from datetime import datetime
from django.shortcuts import render
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


@login_required
def districtOption(request,state_id):
    response = {}
    district_options = '<option value="" selected>Select District</option>'
    tehsil_options = '<option value="" selected>Select Tehsil</option>'
    village_options = '<option value="" selected>Select Village</option>'
    districts = TblNewDistrict.objects.filter(state_id=state_id)

    for district in districts :
        district_options += "<option value="+str(district.id)+">"+district.district_name+"</option>"
        
        tehsils = TblNewTehsil.objects.filter(district_id=district.id)
        for tehsil in tehsils:
            tehsil_options  +=   "<option value="+str(tehsil.id)+">"+tehsil.tehsil_name+"</option>"
            
            villages = TblNewVillage.objects.filter(tehsil_id=tehsil.id)
            for village in villages:
                village_options +=   "<option value="+str(village.id)+">"+village.village_name+"</option>"

    response['district_options']    = district_options
    response['tehsil_options']      = tehsil_options
    response['village_options']     = village_options
    return JsonResponse(response)

@login_required
def tehsilOption(request,district_id):
    response = {}
    tehsil_options = '<option value="" selected>Select Tehsil</option>'
    village_options = '<option value="" selected>Select Village</option>'
        
    tehsils = TblNewTehsil.objects.filter(district_id=district_id)
    for tehsil in tehsils:
        tehsil_options  +=   "<option value="+str(tehsil.id)+">"+tehsil.tehsil_name+"</option>"
        
        villages = TblNewVillage.objects.filter(tehsil_id=tehsil.id)
        for village in villages:
            village_options +=   "<option value="+str(village.id)+">"+village.village_name+"</option>"

    response['tehsil_options']      = tehsil_options
    response['village_options']     = village_options
    return JsonResponse(response)

@login_required
def villageOption(request, tehsil_id):
    response = {}
    village_options = '<option value="" selected>Select Village</option>'
    villages = TblNewVillage.objects.filter(tehsil_id=tehsil_id)
    for village in villages:
        village_options +=   "<option value="+str(village.id)+">"+village.village_name+"</option>"

    response['village_options']     = village_options
    return JsonResponse(response)


@login_required
def enrolledStudents(request):
    state = TblStates.objects.filter().values()
    enrolled = TblEntranceRegistration.objects.filter().order_by('-id').values()
    for each_student in enrolled:
        each_student['address'] = ""
        if each_student['address_hno'] != "" and each_student['address_hno'] is not None:
            each_student['address'] += each_student['address_hno']+ ", "

        if each_student['address_locality'] != "" and each_student['address_locality'] is not None:
            each_student['address'] += each_student['address_locality']+ ", "

        if each_student['address_village'] != "" and each_student['address_village'] is not None:
            each_student['address'] += each_student['address_village']+ ", "

        if each_student['address_tehsil'] != "" and each_student['address_tehsil'] is not None:
            each_student['address'] += each_student['address_tehsil']+ ", "

        if each_student['address_district'] != "" and each_student['address_district'] is not None:
            each_student['address'] += each_student['address_district'] + ", "

        if each_student['address_state'] != "" and each_student['address_state'] is not None:
            each_student['address'] += each_student['address_state']
        
        if each_student['address'] == "":
            each_student['address'] = "-"
        
        each_student['entrance'] = ""
        if each_student['entrance_status'] == 1:
            each_student['entrance'] = "Appeared"
        else:
            each_student['entrance'] = "Not Appeared"

        if TblEntranceOtp.objects.filter(student_contact=each_student['contact']).exists():
            otp_stored = TblEntranceOtp.objects.get(student_contact=each_student['contact'])
            each_student['otp_stored'] = otp_stored.otp
            
    context = {}
    context['page_title']   = "Enrolled Candidates"
    context['total_candidates']   = len(enrolled)
    context['enrolled']     = list(enrolled)
    context['states']       = list(state)

    template = 'entrance/enrolled-list.html'
    return render(request, template, context)

@login_required
def addEntranceCandidate(request):
    if request.method == "GET":
        context = {}
        state = TblStates.objects.filter().values()
        course = TblBranch.objects.filter(college_id = request.user.organization_id)

        context['states']       = list(state)
        context['courses']      = course
        template = "entrance/add-entrance-candidate.html"
        return render(request, template, context)
    
    if request.method == "POST":
        try:
            candidate_name      = request.POST.get('candidate_name')
            contact_no          = request.POST.get('contact')
            course_id           = request.POST.get('course_id')
            email               = request.POST.get('email')
            address_hno         = request.POST.get('hno')
            address_locality    = request.POST.get('locality')
            address_village     = request.POST.get('village_name')
            address_tehsil      = request.POST.get('tehsil_name')
            address_district    = request.POST.get('district_name')
            address_state       = request.POST.get('state_name')

            register = TblEntranceRegistration()
            register.student_name       = candidate_name
            register.contact            = contact_no

            if course_id != '':
                register.course_id      = course_id
                register.course_name    = getModelColumnById(TblBranch, course_id, 'branch')

            register.entrance_status    = 0
            register.email_address      = email
            register.address_hno        = address_hno
            register.address_locality   = address_locality

            if address_village != '':
                register.village_id         = address_village
                register.address_village    = getModelColumnById(TblNewVillage, address_village, 'village_name')

            if address_tehsil != '':
                register.tehsil_id          = address_tehsil
                register.address_tehsil     = getModelColumnById(TblNewTehsil, address_tehsil, 'tehsil_name')

            if address_district != '':
                register.district_id        = address_district
                register.address_district   = getModelColumnById(TblNewDistrict, address_district, 'district_name')

            if address_state != '':
                register.state_id           = address_state
                register.address_district   = getModelColumnById(TblStates, address_state, 'state_name')

            register.registered_by  = request.user.id
            if bool(request.FILES.get('student_image', False)) == True:
                student_image                = request.FILES['student_image']
                folder                      = 'media/entrance_student_images/'
                storage                     = FileSystemStorage(location=folder)
                timestamp                   = int(time.time())
                session_record_filename     = student_image.name
                temp                        = session_record_filename.split('.')
                session_record_filename     = 'entrance/student_image'+str(timestamp)+"."+temp[(len(temp) - 1)]
                student_img_file             = storage.save(session_record_filename, student_image)
                student_img                  = storage.url(student_img_file)

                student_img = student_img.split('media/')[1]
                student_img = "/"+folder+student_img
                register.photo   = student_img
            register.save()
            return render(request,"entrance/enrolled-list.html")

        except Exception as err:
            res = HttpResponse("Registration Number already exists", content_type='application/json')
            res.status_code = 400
            return res


@login_required
def ajaxEnrolledCandidate(request):
    if request.method == "POST":
        filtered_candidates = TblEntranceRegistration.objects

        if request.POST.get('filter_date'):
            filter_date = datetime.strptime(str(request.POST.get('filter_date')), "%d/%m/%Y").strftime("%Y-%m-%d")
        else:
            filter_date = datetime.now().date()

        filtered_candidates = filtered_candidates.filter(created_at__icontains = filter_date)

        if request.POST.get('filter_course'):
            filtered_candidates = filtered_candidates.filter(course_id = request.POST.get('filter_course'))

        if request.POST.get('filter_status'):
            filtered_candidates = filtered_candidates.filter(entrance_status = request.POST.get('filter_status'))

        if request.POST.get('filter_course'):
            filtered_candidates = filtered_candidates.filter(course_id = request.POST.get('filter_course'))
        
        if request.POST.get('filter_search'):
            filtered_candidates = filtered_candidates.filter(student_name__icontains = request.POST.get('filter_search'))

        if request.POST.get('filter_year'):
            year_set    = str(request.POST.get('filter_year')).split("-")
            from_year   = datetime(int(year_set[0]), 3, 31).date()
            to_year     = datetime(int(year_set[1]), 4, 1).date()
            filtered_candidates = filtered_candidates.filter(created_at__range = [from_year, to_year])
        
        filtered_candidates = filtered_candidates.values()
        
        for each_student in filtered_candidates:
            each_student['address'] = ""
            if each_student['address_hno'] != "" and each_student['address_hno'] is not None:
                each_student['address'] += each_student['address_hno']+ ", "

            if each_student['address_locality'] != "" and each_student['address_locality'] is not None:
                each_student['address'] += each_student['address_locality']+ ", "

            if each_student['address_village'] != "" and each_student['address_village'] is not None:
                each_student['address'] += each_student['address_village']+ ", "

            if each_student['address_tehsil'] != "" and each_student['address_tehsil'] is not None:
                each_student['address'] += each_student['address_tehsil']+ ", "

            if each_student['address_district'] != "" and each_student['address_district'] is not None:
                each_student['address'] += each_student['address_district'] + ", "

            if each_student['address_state'] != "" and each_student['address_state'] is not None:
                each_student['address'] += each_student['address_state']
            
            if each_student['address'] == "":
                each_student['address'] = "-"
            
            each_student['entrance'] = ""
            if each_student['entrance_status'] == 1:
                each_student['entrance'] = "Appeared"
            else:
                each_student['entrance'] = "Not Appeared"

            if TblEntranceOtp.objects.filter(student_contact=each_student['contact']).exists():
                otp_stored = TblEntranceOtp.objects.get(student_contact=each_student['contact'])
                each_student['otp_stored'] = otp_stored.otp
            
            # if TblEntranceQuiz.objects.filter(candidate_id = each_student['id']).exists():
            #     quiz_result = TblEntranceQuiz.objects.filter(candidate_id = each_student['id']).values()
            #     if quiz_result.result > 0:
            #         if quiz_result.result == 1:
            #             each_student['result'] = "Passed"
            #         elif quiz_result.result == 2:
            #             each_student['result'] = "Failed"
            #         elif quiz_result.result == 3:
            #             each_student['result'] = "Awaited"
                    
            #     else:
            #         each_student['result'] = "-"
            # else:
            #     each_student['result'] = "-"
            
        context = {}
        context['page_title'] = "Enrolled Students"
        context['total_candidates']   = len(filtered_candidates)
        context['enrolled'] = list(filtered_candidates)

        template = 'entrance/ajax-filter-candidate-list.html'
        return render(request, template, context)


@login_required
def getCandidateDetails(request, candidate_id):
    if request.method == "GET":

        details = TblEntranceRegistration.objects.filter(id = candidate_id).values()
        context = {}
        context['student_detail']   = list(details)

        template = "entrance/candidate-details.html"
        return render(request, template, context)

@login_required
def editRegistration(request, candidate_id):
    if request.method == "GET":
        candidate_details   = TblEntranceRegistration.objects.filter(id = candidate_id).values()
        state               = TblStates.objects.filter().values()
        districts           = TblNewDistrict.objects.filter().values()
        tehsil              = TblNewTehsil.objects.filter().values()
        villages            = TblNewVillage.objects.filter().values()
        course              = TblBranch.objects.filter(college_id = request.user.organization_id).values()

        context = {}
        context['candidate_id']     = candidate_id
        context['student_detail']   = list(candidate_details)
        context['states']           = list(state)
        context['districts']        = list(districts)
        context['tehsils']          = list(tehsil)
        context['villages']         = list(villages)
        context['courses']          = list(course)

        template = "entrance/edit-candidate.html"
        return render(request, template, context)

@login_required
def saveRegistration(request):
    if request.method == "POST":
        try:
            candidate_name      = request.POST.get('candidate_name')
            contact_no          = request.POST.get('contact')
            course_id           = request.POST.get('course_id')
            email               = request.POST.get('email')
            address_hno         = request.POST.get('hno')
            address_locality    = request.POST.get('locality')
            address_village     = request.POST.get('village_name')
            address_tehsil      = request.POST.get('tehsil_name')
            address_district    = request.POST.get('district_name')
            address_state       = request.POST.get('state_name')

            if TblEntranceRegistration.objects.filter(id=request.POST.get('candidate_id')).exists():
                register = TblEntranceRegistration.objects.get(id=request.POST.get('candidate_id'))
                register.student_name       = candidate_name
                register.contact            = contact_no
                if course_id != '':
                    register.course_id          = course_id
                    register.course_name        = getModelColumnById(TblBranch, course_id, 'branch')
                register.entrance_status    = 0
                register.email_address      = email
                register.address_hno        = address_hno
                register.address_locality   = address_locality

                if address_village != '':
                    register.village_id         = address_village
                    register.address_village    = getModelColumnById(TblNewVillage, address_village, 'village_name')

                if address_tehsil != '':
                    register.tehsil_id          = address_tehsil
                    register.address_tehsil     = getModelColumnById(TblNewTehsil, address_tehsil, 'tehsil_name')

                if address_district != '':
                    register.district_id        = address_district
                    register.address_district   = getModelColumnById(TblNewDistrict, address_district, 'district_name')

                if address_state != '':
                    register.state_id           = address_state
                    register.address_state   = getModelColumnById(TblStates, address_state, 'state_name')

                register.registered_by  = request.user.id
                if bool(request.FILES.get('student_image', False)) == True:
                    student_image                = request.FILES['student_image']
                    folder                      = 'media/entrance_student_images/'
                    storage                     = FileSystemStorage(location=folder)
                    timestamp                   = int(time.time())
                    session_record_filename     = student_image.name
                    temp                        = session_record_filename.split('.')
                    session_record_filename     = 'entrance/student_image'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    student_img_file             = storage.save(session_record_filename, student_image)
                    student_img                  = storage.url(student_img_file)

                    student_img = student_img.split('media/')[1]
                    student_img = "/"+folder+student_img
                    register.photo   = student_img
                register.save()
                return render(request,"entrance/enrolled-list.html")

        except Exception as err:
            res = HttpResponse("Registration Number already exists", content_type='application/json')
            res.status_code = 400
            return res


@login_required
def exportEnrolledListtXLS(request, filter_year, filter_course, filter_date, filter_status):
    result_list = TblEntranceRegistration.objects.all()

    if filter_date != '0':
        result_date = datetime.strptime(str(filter_date), "%d-%m-%Y").strftime("%Y-%m-%d")
    else:
        result_date = datetime.now().date()
    result_list = result_list.filter(created_at__icontains=result_date)

    if filter_status != '2':
        result_list = result_list.filter(entrance_status=filter_status)

    if filter_course != '0':
        result_list = result_list.filter(course_id=filter_course)
    
    if filter_year != '0':
        year_set    = str(filter_year).split("-")
        from_year   = datetime(int(year_set[0]), 3, 31).date()
        to_year     = datetime(int(year_set[1]), 4, 1).date()
        result_list = result_list.filter(created_at__range = [from_year, to_year])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=Enrolled-List.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(vertical='top', wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Result'

    # Define the titles for columns
    columns = []
    columns += [ 'S.No' ]
    columns += [ 'Candidate Name' ]
    columns += [ 'Contact' ]
    columns += [ 'Entrance Status' ]
    columns += [ 'Course' ]
    columns += [ 'Address' ]
    columns += [ 'Referral by type' ]
    columns += [ 'Reference by' ]
    columns += [ 'Reference by Branch' ]
    columns += [ 'Reference by Year' ]
    columns += [ 'Registered By' ]
    columns += [ 'Registered On' ]


    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 15

    for sno, each_result in enumerate(result_list, 1):
        row_num += 1
        row = []
        row += [ sno ]
        row += [ each_result.student_name ]
        row += [ each_result.contact ]
        if each_result.entrance_status == 1:
            ent_status = "Appeared"
        else:
            ent_status = "Not Appeared"
        
        address = ""
        if each_result.address_hno != "" and each_result.address_hno is not None:
            address += each_result.address_hno+ ", "

        if each_result.address_locality != "" and each_result.address_locality is not None:
            address += each_result.address_locality+ ", "

        if each_result.address_village != "" and each_result.address_village is not None:
            address += each_result.address_village+ ", "

        if each_result.address_tehsil != "" and each_result.address_tehsil is not None:
            address += each_result.address_tehsil+ ", "

        if each_result.address_district != "" and each_result.address_district is not None:
            address += each_result.address_district + ", "

        if each_result.address_state != "" and each_result.address_state is not None:
            address += each_result.address_state
        
        if address == "":
            address = "-"

        row += [ ent_status ]
        row += [ each_result.course_name ]
        row += [ address ]
        if each_result.referenced_by_name is not None:
            if each_result.is_student == 0:
                row += [ "Student" ]
                row += [ each_result.referenced_by_name ]
                row += [ each_result.referenced_by_branch_name ]
                row += [ each_result.referenced_by_year ]
            elif each_result.is_student == 1:
                row += [ "Other" ]
                row += [ each_result.referenced_by_name ]
                row += [ "-" ]
                row += [ "-" ]
                
        else:
            row += [ "-" ]
            row += [ "-" ]
            row += [ "-" ]
            row += [ "-" ]
        row += [ getUserName(each_result.registered_by) ]
        row += [ each_result.created_at ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    
    workbook.save(response)
    return response







@login_required
def candidateResult(request):
    state = TblStates.objects.filter().values()
    enrolled = TblEntranceQuiz.objects.filter().values()
    for each_student in enrolled:

        if TblEntranceRegistration.objects.filter(id = each_student['candidate_id']).exists():
            candidate_info = TblEntranceRegistration.objects.get(id=each_student['candidate_id'])
            each_student['id']              = candidate_info.id
            each_student['student_name']    = candidate_info.student_name
            each_student['contact']         = candidate_info.contact

            total_questions = len(each_student['question_answers'])
            total_attempted = 0
            total_correct   = 0
            total_skipped   = 0
            total_wrong     = 0
            all_questions   = each_student['question_answers']
            for each_question in all_questions:
                if each_question['answer'] in [1,2,3,4]:
                    total_attempted = total_attempted + 1
                    if int(each_question['answer']) == int(getModelColumnById(TblQuizQuestionnaire, each_question['question_id'], 'answer_key')):
                        total_correct = total_correct + 1
                    else:
                        total_wrong = total_wrong + 1

                elif each_question['answer'] == 5:
                    total_skipped = total_skipped + 1
            
            each_student['total_questions'] = total_questions
            each_student['total_attempted'] = total_attempted
            each_student['total_correct']   = total_correct
            each_student['total_skipped']   = total_skipped
            each_student['total_wrong']     = total_wrong

            percentage = (total_correct / total_questions) * 100
            each_student['percentage'] = round(percentage, 2)

    context = {}
    context['page_title'] = "Candidates Result"
    context['total_results'] = len(enrolled)
    context['enrolled'] = list(enrolled)
    context['state'] = list(state)

    template = 'entrance/result-list.html'
    return render(request, template, context)



@login_required
def ajaxCandidateResult(request):
    enrolled = TblEntranceQuiz.objects

    if request.POST.get('filter_date'):
        filter_date = datetime.strptime(str(request.POST.get('filter_date')), "%d/%m/%Y").strftime("%Y-%m-%d")
        enrolled = enrolled.filter(quiz_start_datetime__icontains = filter_date)
    # else:
    #     enrolled = enrolled.filter(quiz_end_datetime__icontains = datetime.now().date())

    if request.POST.get('filter_percentage'):
        percentage = str(request.POST.get('filter_percentage')).split("-")
        enrolled = enrolled.filter(result__range = [float(percentage[0]), float(percentage[1])])

    enrolled    = enrolled.order_by("-quiz_end_datetime").values()
    for each_student in enrolled:
        if TblEntranceRegistration.objects.filter(id = each_student['candidate_id']).exists():
            candidate_info = TblEntranceRegistration.objects.get(id=each_student['candidate_id'])
            each_student['id']              = candidate_info.id
            each_student['student_name']    = candidate_info.student_name
            each_student['contact']         = candidate_info.contact

            total_questions = len(each_student['question_answers'])
            total_attempted = 0
            total_correct   = 0
            total_skipped   = 0
            total_wrong     = 0
            all_questions   = each_student['question_answers']
            for each_question in all_questions:
                if each_question['answer'] in [1,2,3,4]:
                    total_attempted = total_attempted + 1
                    if int(each_question['answer']) == int(getModelColumnById(TblQuizQuestionnaire, each_question['question_id'], 'answer_key')):
                        total_correct = total_correct + 1
                    else:
                        total_wrong = total_wrong + 1

                elif each_question['answer'] == 5:
                    total_skipped = total_skipped + 1
            
            percentage = (total_correct / total_questions) * 100

            each_student['total_questions'] = total_questions
            each_student['total_attempted'] = total_attempted
            each_student['total_correct']   = total_correct
            each_student['total_skipped']   = total_skipped
            each_student['total_wrong']     = total_wrong
            each_student['percentage']      = round(percentage, 2)

    context                 = {}
    context['page_title']   = "Candidates Result"
    context['total_results'] = len(enrolled)
    context['enrolled']     = list(enrolled)

    template = 'entrance/ajax-filter-candidate-result-list.html'
    return render(request, template, context)


@login_required
def getResultCandidateDetails(request, candidate_id, quiz_date):
    if request.method == "GET":
        context = {}
        if quiz_date != '0':
            result_details = TblEntranceQuiz.objects.filter(candidate_id = candidate_id, quiz_start_datetime__icontains = quiz_date).values()

            for result in result_details:
                candidate_details = TblEntranceRegistration.objects.get(id=candidate_id)
                result['student_name'] = candidate_details.student_name
                result['contact'] = candidate_details.contact
                result['email'] = candidate_details.email_address

                for each_question in result['question_answers']:
                    options                     = TblQuizQuestionnaire.objects.filter(id=each_question['question_id']).values('option_1', 'option_2', 'option_3', 'option_4', 'answer_key')[0]
                    each_question['option_no']    = int(options['answer_key'])
                    each_question['options']    = options['option_'+options['answer_key']]
                    if each_question['answer'] in [1,2,3,4]:
                        each_question['ans']        = options['option_'+str(each_question['answer'])]

            context['student_detail']   = list(result_details)

        template = "entrance/candidate-result-details.html"
        return render(request, template, context)



@login_required
def exportResultXLS(request, filter_date, filter_percentage):

    result_list = TblEntranceQuiz.objects.all()

    if filter_date != '0':
        result_date = datetime.strptime(str(filter_date), "%d-%m-%Y").strftime("%Y-%m-%d")
    # else:
    #     result_date = datetime.now().date()
        result_list = result_list.filter(quiz_end_datetime__icontains=result_date)

    if filter_percentage != '0':
        result = str(filter_percentage).split("-")
        result_list = result_list.filter(result__range=result)
    
    result_list = result_list.order_by("-quiz_end_datetime")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Result-List.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Result'
    
    # Define the titles for columns
    columns = []
    columns += [ 'S.No' ]
    columns += [ 'Candidate Name' ]
    columns += [ 'Contact' ]
    columns += [ 'Quiz Date' ]
    columns += [ 'Status' ]
    columns += [ 'Total Questions' ]
    columns += [ 'Total Attempted' ]
    columns += [ 'Total Skipped' ]
    columns += [ 'Total Correct' ]
    columns += [ 'Total Wrong' ]
    columns += [ 'Percentage' ]
    columns += [ 'Referral by type' ]
    columns += [ 'Reference by' ]
    columns += [ 'Reference by Branch' ]
    columns += [ 'Reference by Year' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 15

    for sno, each_result in enumerate(result_list, 1):
        row_num += 1
        student_details = TblEntranceRegistration.objects.get(id=each_result.candidate_id)
        # Define the data for each cell in the row

        row = []
        row += [ sno ]
        row += [ student_details.student_name ]
        row += [ student_details.contact ]
        row += [ datetime.strptime( str(each_result.quiz_start_datetime), "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y") ]
        if each_result.quiz_end_datetime is not None:
            row += [ 'Completed' ]
        else:
            row += [ 'Incompleted' ]
        
        
        total_questions = len(each_result.question_answers)
        total_attempted = 0
        total_correct   = 0
        total_skipped   = 0
        total_wrong     = 0
        all_questions   = each_result.question_answers
        for each_question in all_questions:
            if each_question['answer'] in [1,2,3,4]:
                total_attempted = total_attempted + 1
                if int(each_question['answer']) == int(getModelColumnById(TblQuizQuestionnaire, each_question['question_id'], 'answer_key')):
                    total_correct = total_correct + 1
                else:
                    total_wrong = total_wrong + 1

            elif each_question['answer'] == 5:
                total_skipped = total_skipped + 1
        percentage = (total_correct / total_questions) * 100

        row += [ total_questions ]
        row += [ total_attempted ]
        row += [ total_skipped ]
        row += [ total_correct ]
        row += [ total_wrong ]
        row += [ round(percentage, 2) ]
        
        if student_details.referenced_by_name is not None:
            if student_details.is_student == 0:
                row += [ "Student" ]
                row += [ student_details.referenced_by_name ]
                row += [ student_details.referenced_by_branch_name ]
                row += [ student_details.referenced_by_year ]
            elif student_details.is_student == 1:
                row += [ "Other" ]
                row += [ student_details.referenced_by_name ]
                row += [ "-" ]
                row += [ "-" ]
                
        else:
            row += [ "-" ]
            row += [ "-" ]
            row += [ "-" ]
            row += [ "-" ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    
    workbook.save(response)
    return response


@login_required
def editCandidate(request):
    if request.method == "POST":
        try:
            candidate_name      = request.POST.get('candidate_name')
            contact_no          = request.POST.get('contact')
            course_id           = request.POST.get('course_id')
            email               = request.POST.get('email')
            address_hno         = request.POST.get('hno')
            address_locality    = request.POST.get('locality')
            address_village     = request.POST.get('village_name')
            address_tehsil      = request.POST.get('tehsil_name')
            address_district    = request.POST.get('district_name')
            address_state       = request.POST.get('state_name')

            if TblEntranceRegistration.objects.filter(id=request.POST.get('candidate_id')).exists():

                register = TblEntranceRegistration.objects.get(id=request.POST.get('candidate_id'))
                register.student_name       = candidate_name
                register.contact            = contact_no
                register.course_id          = course_id
                register.course_name        = getModelColumnById(TblBranch, course_id, 'branch')
                register.email_address      = email
                register.address_hno        = address_hno
                register.address_locality   = address_locality
                register.entrance_status    = register.entrance_status

                if address_village is not None or address_village != "":
                    register.village_id         = address_village
                    register.address_village    = getModelColumnById(TblNewVillage, address_village, 'village_name')

                if address_tehsil is not None or address_tehsil != "":
                    register.tehsil_id          = address_tehsil
                    register.address_tehsil     = getModelColumnById(TblNewTehsil, address_tehsil, 'tehsil_name')

                if address_district is not None or address_district != "":
                    register.district_id        = address_district
                    register.address_district   = getModelColumnById(TblNewDistrict, address_district, 'district_name')

                if address_state is not None or address_state != "":
                    register.state_id           = int(address_state)
                    register.address_state      = getModelColumnById(TblStates, int(address_state), 'state_name')

                register.registered_by  = request.user.id
                if bool(request.FILES.get('student_image', False)) == True:
                    student_image                = request.FILES['student_image']
                    folder                      = 'media/entrance_student_images/'
                    storage                     = FileSystemStorage(location=folder)
                    timestamp                   = int(time.time())
                    session_record_filename     = student_image.name
                    temp                        = session_record_filename.split('.')
                    session_record_filename     = 'entrance/student_image'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    student_img_file             = storage.save(session_record_filename, student_image)
                    student_img                  = storage.url(student_img_file)

                    student_img = student_img.split('media/')[1]
                    student_img = "/"+folder+student_img
                    register.photo   = student_img
                register.save()
                return render(request,"entrance/enrolled-list.html")

        except Exception as err:
            res = HttpResponse("Registration Number already exists", content_type='application/json')
            res.status_code = 400
            return res

@login_required
def sendResultSMS(request):
    if request.method == "POST":
        candidate_ids = request.POST.getlist("candidate_ids")
        candidate_ids = ast.literal_eval(candidate_ids[0])

        try:
            for each_candidate in candidate_ids:
                candidate_id = each_candidate.split(":")[0]
                date = each_candidate.split(":")[1]
                date = datetime.strptime(date, "%Y-%m-%d")
    
                percentage = getModelColumnByColumnId(TblEntranceQuiz, 'candidate_id', candidate_id, "result")
                candidate_name = getModelColumnById(TblEntranceRegistration, candidate_id, 'student_name')
                
                if percentage > 75:
                    grade = "A"
                elif percentage > 65 and percentage < 75:
                    grade = "B"
                elif percentage > 55 and percentage < 65:
                    grade = "C"
                elif percentage > 35 and percentage < 55:
                    grade = "D"
                else:
                    grade = "F"

                if grade == "F":
                    message = "{name}, BIPE प्रवेश परीक्षा 2022, में {percent} अंक के साथ अनुत्तीर्ण होने पर सूचित किया जाता है कि आपको संस्था में प्रवेश के लिए स्वीकृति प्रदान नहीं की गई है.\nपुनः प्रवेश परीक्षा में सम्मिलित होने एवं अधिक जानकारी के लिए 7310077788 पर सम्पर्क करें.\nBGIVNS,".format(name=candidate_name, percent=percentage)
                else:
                    message = "{name}, BIPE प्रवेश परीक्षा 2022, में {percent}% अंक के साथ उत्तीर्ण होने पर Grade {category} की स्कालरशिप के अंतर्गत संस्था में प्रवेश के लिए स्वीकृति प्रदान की गई है.\nप्रवेश पंजीकरण 22 मार्च 2022 से प्रारम्भ है.\nअधिक जानकारी के लिए 7310077788 पर सम्पर्क करें.\nBGIVNS,".format(name=candidate_name, percent=percentage, category=grade)
                
                contact = getModelColumnById(TblEntranceRegistration, candidate_id, 'contact')

                sendSMS('BGIVNS', contact, message)
                # sendSMS('BGIVNS', '8826671267', message)

                update_sms = TblEntranceQuiz.objects.get(candidate_id = candidate_id, quiz_end_datetime__icontains = date.date())
                update_sms.is_notified = 1
                update_sms.save()
    
            return JsonResponse({'message':'All SMS sent successfully'})
        except Error as err:
            return JsonResponse({'message':'Failed to send the SMS '+err})
