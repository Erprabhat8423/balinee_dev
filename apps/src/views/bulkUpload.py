import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound ,Http404
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q,F
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.hashers import make_password
from ..decorators import has_par


@login_required
@has_par(sub_module_id=6,permission='list')
def importBMC(request):
    if request.method == 'POST' :
        bmc_sheet = request.FILES["bmc_sheet"]
        wb = openpyxl.load_workbook(bmc_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        invalid_data  = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['BMC CODE', 'NAME']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/import-bmc')

        bmc_data = []
        for data in excel_data:
            temp = {}
            
            bmc_code = data[0].strip()
            bmc_name = data[1].strip()
            
            if bmc_code == 'None':
                temp['bmc_code_blank_temp'] = "BMC Code is blank"
                invalid_data += 1
            else:
                temp['bmc_code'] = bmc_code
            if  bmc_name == 'None':
                temp['bmc_name_blank_temp'] = "BMC Name is blank"
                invalid_data += 1
            else:
                temp['bmc_name'] = bmc_name
            if SpBMC.objects.filter(bmc_code=bmc_code).exists():
                temp['bmc_code_temp'] = "BMC Already Exists"
                invalid_data += 1
            

            bmc_data.append(temp)
        context = {}
       
        context['excel_data'] = bmc_data
        context['invalid_data'] = invalid_data
        context['page_title']                       = "Upload BMC"
        template = 'bulkUpload/upload-bmc.html'
        return render(request,template,context)
    else:
        context = {}
        context['page_title']                       = "Upload BMC"
        template = 'bulkUpload/upload-bmc.html'
        return render(request,template,context)


from django.utils.timezone import now

def updateBMC(request):
    try:
        bmc_code = request.POST.getlist('bmc_code[]')
        bmc_name = request.POST.getlist('bmc_name[]')
        for idx, code in enumerate(bmc_code):
            if not SpBMC.objects.filter(bmc_code =code).exists():
                obj = SpBMC()
                obj.bmc_code =code
                obj.name = bmc_name[idx]
                obj.save()
        response = {'flag': True, 'message': "Records have been updated successfully."}
    except Exception as e:
        # Log the error
        print(f"Error updating BMC: {e}")
        response = {'flag': False, 'message': "Records not updated"}
    
    return JsonResponse(response)

    

@login_required
@has_par(sub_module_id=6,permission='list')
def importMPP(request):
    if request.method == 'POST' :
        mpp_sheet = request.FILES["mpp_sheet"]
        wb = openpyxl.load_workbook(mpp_sheet)
        worksheet = wb.active
        excel_data = list()
        header_data = list()
        i = 0
        invalid_data  = 0
        for row in worksheet.iter_rows():
            if(i > 0):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            else:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                header_data.append(row_data)    
            i = int(i) + 1
        
        header_list = ['BMC CODE', 'MPP CODE', 'NAME', 'LATITUDE', 'LONGITUDE', 'WEF DATE']
        for header in header_data[0]:
            if header not in header_list:
                messages.error(request, 'Invalid File Format', extra_tags='success')
                return redirect('/import-mpp')

        mpp_data = []
        for data in excel_data:
            temp = {}
            
            bmc_code    = data[0].strip()
            mpp_code    = data[1].strip()
            mpp_name    = data[2].strip()
            latitude    = data[3].strip()
            longitude   = data[4].strip()
            wef_date    = data[5].strip()
            
            if bmc_code == 'None' :
                temp['bmc_code_blank_temp'] = "BMC Code is blank"
                invalid_data += 1
            else:
                temp['bmc_code'] = bmc_code
            if mpp_code == 'None':
                temp['mpp_code_blank_temp'] = "MPP Code is blank"
                invalid_data += 1
            else:
                temp['mpp_code'] = mpp_code
            if mpp_name == 'None':
                temp['mpp_name_blank_temp'] = "BMC Name is blank"
                invalid_data += 1
            else:
                temp['mpp_name'] = mpp_name

            if not SpBMC.objects.filter(bmc_code=bmc_code).exists():
                temp['bmc_code_temp'] = "Invaild BMC Code"
                invalid_data += 1

            if SpMPP.objects.filter(mpp_code=mpp_code).exists():
                temp['mpp_code_temp'] = "MPP Already Exists"
                invalid_data += 1
            temp['latitude']    = latitude
            temp['longitude']   = longitude
            temp['wef_date']    = wef_date
            temp['wef_date']    = wef_date

            mpp_data.append(temp)
        context = {}
       
        context['excel_data']       = mpp_data
        context['invalid_data']     = invalid_data
        context['page_title']       = "Upload MPP"
        template = 'bulkUpload/upload-mpp.html'
        return render(request,template,context)
    else:
        context = {}
        context['page_title']        = "Upload MPP"
        template = 'bulkUpload/upload-mpp.html'
        return render(request,template,context)

def genrateMPPTrCode(bmc_code, mpp_code):

    bmc_code = str(bmc_code)
    get_last_digit_bmc = bmc_code[-3:]
    get_last_digit_mpp = str(mpp_code)

    if len(mpp_code) < 2:
        get_last_digit_mpp = "00" + mpp_code
    elif len(mpp_code) < 3:
        get_last_digit_mpp = "0" + mpp_code
    else:
        get_last_digit_mpp = mpp_code
    mpp_tr_code = get_last_digit_bmc + get_last_digit_mpp

    return mpp_tr_code
    
from django.utils.timezone import now

def updateMPP(request):
    try:
        bmc_code    = request.POST.getlist('bmc_code[]')
        mpp_code    = request.POST.getlist('mpp_code[]')
        mpp_name    = request.POST.getlist('mpp_name[]')
        latitude    = request.POST.getlist('latitude[]')
        longitude   = request.POST.getlist('longitude[]')
        wef_date    = request.POST.getlist('wef_date[]')
        
        if not (len(bmc_code) == len(mpp_code) == len(mpp_name)):
            raise ValueError("Input lists must have the same length.")        
        for idx, code in enumerate(mpp_code):
            if not SpMPP.objects.filter(mpp_code =mpp_code[idx]).exists():
                get_mpp_ta_code = genrateMPPTrCode(bmc_code[idx],mpp_code[idx])
                get_bmc_id  = SpBMC.objects.filter(bmc_code= bmc_code[idx]).first()
                obj = SpMPP()
                obj.bmc_id = get_bmc_id.id
                obj.mpp_code = mpp_code[idx]
                obj.name = mpp_name[idx]
                
                if latitude[idx] != '':
                    obj.latitude = latitude[idx]
                if longitude[idx] != '':
                    obj.longitude = longitude[idx]
                if wef_date[idx] != '' and wef_date[idx] != "None":
                    obj.wef_date = datetime.strptime(wef_date[idx], '%Y-%m-%d %H:%M:%S').date()
                if get_mpp_ta_code:
                    obj.mpp_tr_code = get_mpp_ta_code
                obj.save()
            
        response = {'flag': True, 'message': "Records have been updated successfully."}
    except Exception as e:
        # Log the error
        print(f"Error updating MPP: {e}")
        response = {'flag': False, 'message': "Records not updated"}
    
    return JsonResponse(response)


def mapBMC(request, id):
    if request.method == "POST":
        mpp_list = request.GET.getlist('mpp_list[]')
        SpMPP.objects.filter(id__in = mpp_list).update(employee_id = id)
        response = {'flag': True, 'message': "Records have been updated successfully."}  
    else:
        bmc_list = SpBMC.objects.filter(status = 1)
        context = {}
        context['bmc_list'] = bmc_list
        context['emp_id'] = id
        template = 'bulkUpload/map-bmc-mpp.html'
        return render(request, template, context)
    
def getMppList(request,bmc_list):
    get_mpp_list = SpMPP.objects.filter(status = 1, bmc_id__in =bmc_list )
    response = {}
    options = '<option value="" selected>Select MPP</option>'
    for get_mpp in get_mpp_list :
        options += "<option value="+str(get_mpp.id)+">"+get_mpp.name+"("+get_mpp.mpp_code+")"+"</option>"
    response['options'] = options
    return JsonResponse(response)



@login_required
def exportMPPExcel(request):
    path = 'excel-templates/mpp_upload_sheet.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=mpp_upload_sheet.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Crate Upload excel format exported'
            activity    = 'Crate Upload excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export MPP Upload Sheet', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')
            
            return response
    else:
        raise Http404



@login_required
def exportBMCExcel(request):
    path = 'excel-templates/bmc_upload_sheet.xlsx'
    print(os.path) 
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=bmc_upload_sheet.xlsx'

            #Save Activity
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Crate Upload excel format exported'
            activity    = 'Crate Upload excel format exported by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            saveActivity('Product & Variant Management', 'Export BMC Upload Sheet', heading, activity, request.user.id, user_name, 'generateIndentReport.png', '1', 'web.png')
            
            return response
    else:
        raise Http404
